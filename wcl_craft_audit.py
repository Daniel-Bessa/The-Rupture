#!/usr/bin/env python3
"""
WarcraftLogs Crafted Gear Audit Tool — Midnight Season 1
=========================================================
Pulls player gear from a WarcraftLogs report and identifies crafted items,
spark usage, crest tier, and embellishments. Outputs to .xlsx.

Usage:
    python wcl_craft_audit.py

You will be prompted for:
    - WarcraftLogs Client ID & Secret
    - Report code (the alphanumeric part of a WCL report URL)

Requirements:
    pip install requests openpyxl
"""

import sys
import os
import json
import requests
from datetime import datetime, timezone
from html import escape
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "src"))
from boss_mechanics import BOSS_MECHANICS, BOSS_HAS_INTERRUPTS
from tracked_spells import (HEALTHSTONE_IDS, HEALTH_POT_IDS, COMBAT_POT_IDS,
                             CLASS_DEFENSIVE_IDS, CLASS_EXTERNAL_IDS, SPELL_NAMES)

# ─── Midnight Season 1 Constants ─────────────────────────────────────────────

# Crafting quality bonus IDs (WoW internal — ranks 1 through 5)
# These are the bonus IDs appended to items when they are player-crafted.
CRAFT_QUALITY_BONUS_IDS = {
    10249: 1, 10250: 2, 10251: 3, 10252: 4, 10253: 5,  # TWW-era IDs
    10255: 1, 10256: 2, 10257: 3, 10258: 4, 10259: 5,  # Midnight-era IDs (if changed)
    11109: 1, 11110: 2, 11111: 3, 11112: 4, 11113: 5,  # Alternate range seen in beta
}

# Bonus IDs that indicate an item was crafted via the crafting order / profession system
CRAFTED_INDICATOR_BONUS_IDS = {
    8960,   # Midnight crafted tag
    8791,   # Midnight crafted (often paired with 8960)
    9497,   # TWW-era "Crafted" tag
    9498,   # TWW-era "Crafted" tag (alternate)
    10222,  # Crafting work order
    10343,  # Recrafted
}

# Known embellishment bonus IDs for Midnight Season 1
# Add more as they're discovered — these help confirm an item is crafted
EMBELLISHMENT_BONUS_IDS = {
    # Placeholder — update with actual Midnight embellishment bonus IDs
    # Format: bonus_id: "Embellishment Name"
}

# Item level thresholds for Midnight Season 1 crafted gear
ILVL_TIERS = {
    "Myth (Spark + Myth Crests)":  (272, 285),
    "Hero (Spark + Hero Crests)":  (259, 272),
    "Epic Base (Spark, no crests)": (252, 259),
    "Rare (Veteran Crests)":       (233, 246),
    "Rare (Adventurer Crests)":    (214, 233),
    "Rare Base (no spark)":        (200, 214),
}

# Gear slot mapping from WCL slot IDs
SLOT_NAMES = {
    0: "Head", 1: "Neck", 2: "Shoulder", 3: "Shirt", 4: "Chest",
    5: "Waist", 6: "Legs", 7: "Feet", 8: "Wrist", 9: "Hands",
    10: "Ring 1", 11: "Ring 2", 12: "Trinket 1", 13: "Trinket 2",
    14: "Back", 15: "Main Hand", 16: "Off Hand",
}


def _player_slug(pname: str) -> str:
    return pname.lower().replace(" ", "_")


# ─── WarcraftLogs API ────────────────────────────────────────────────────────

def get_access_token(client_id: str, client_secret: str) -> str:
    """Get OAuth2 access token using client credentials flow."""
    resp = requests.post(
        "https://www.warcraftlogs.com/oauth/token",
        data={"grant_type": "client_credentials"},
        auth=(client_id, client_secret),
    )
    if resp.status_code != 200:
        print(f"[ERROR] Auth failed ({resp.status_code}): {resp.text}")
        sys.exit(1)
    return resp.json()["access_token"]


def query_wcl(token: str, query: str, variables: dict = None) -> dict:
    """Execute a GraphQL query against WarcraftLogs v2 API."""
    resp = requests.post(
        "https://www.warcraftlogs.com/api/v2/client",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={"query": query, "variables": variables or {}},
    )
    if resp.status_code != 200:
        print(f"[ERROR] API request failed ({resp.status_code}): {resp.text}")
        sys.exit(1)
    data = resp.json()
    if "errors" in data:
        msgs = [e.get("message", "") for e in data["errors"]]
        raise RuntimeError(f"GraphQL error: {'; '.join(msgs)}")
    return data["data"]


def fetch_report_info(token: str, report_code: str) -> dict:
    """Fetch basic report metadata (title, fights, actors, etc.)."""
    query = """
    query ($code: String!) {
        reportData {
            report(code: $code) {
                title
                startTime
                endTime
                guild { name server { name region { slug } } }
                fights {
                    id
                    name
                    kill
                    difficulty
                    encounterID
                    startTime
                    endTime
                    bossPercentage
                }
                masterData(translate: true) {
                    actors {
                        id
                        gameID
                        name
                        type
                        subType
                        server
                    }
                    abilities {
                        gameID
                        name
                    }
                }
            }
        }
    }
    """
    return query_wcl(token, query, {"code": report_code})["reportData"]["report"]


def fetch_player_details(token: str, report_code: str, start_time: float = None, end_time: float = None, fight_ids: list = None) -> dict:
    """Fetch player info including gear for the report."""
    # Get player list
    if fight_ids:
        query = """
        query ($code: String!, $fightIDs: [Int]!) {
            reportData {
                report(code: $code) {
                    playerDetails(fightIDs: $fightIDs)
                }
            }
        }
        """
        variables = {"code": report_code, "fightIDs": fight_ids}
    else:
        query = """
        query ($code: String!, $startTime: Float!, $endTime: Float!) {
            reportData {
                report(code: $code) {
                    playerDetails(startTime: $startTime, endTime: $endTime)
                }
            }
        }
        """
        variables = {"code": report_code, "startTime": start_time, "endTime": end_time}
    details = query_wcl(token, query, variables)["reportData"]["report"]["playerDetails"]
    return details


def fetch_combatant_info_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch gear data via the events endpoint for a specific fight."""
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData {
            report(code: $code) {
                events(dataType: CombatantInfo, fightIDs: [$fightID], limit: 500) {
                    data
                    nextPageTimestamp
                }
            }
        }
    }
    """
    variables = {"code": report_code, "fightID": fight_id}
    result = query_wcl(token, query, variables)
    events = result["reportData"]["report"]["events"]
    return events.get("data", [])


def fetch_cast_events(token: str, report_code: str, fight_id: int, start_time: float = None, end_time: float = None) -> list:
    """Fetch all cast events for a fight. Paginates if needed."""
    all_data = []
    next_ts = start_time
    
    for _ in range(10):  # Max 10 pages to avoid infinite loop
        query = """
        query ($code: String!, $fightID: Int!, $startTime: Float, $endTime: Float) {
            reportData {
                report(code: $code) {
                    events(dataType: Casts, fightIDs: [$fightID], startTime: $startTime, endTime: $endTime, hostilityType: Friendlies, limit: 10000) {
                        data
                        nextPageTimestamp
                    }
                }
            }
        }
        """
        variables = {"code": report_code, "fightID": fight_id}
        if next_ts is not None:
            variables["startTime"] = next_ts
        if end_time is not None:
            variables["endTime"] = end_time
        
        result = query_wcl(token, query, variables)
        events = result["reportData"]["report"]["events"]
        all_data.extend(events.get("data", []))
        next_ts = events.get("nextPageTimestamp")
        if next_ts is None:
            break

    return all_data


def fetch_death_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch death events for a fight (friendly players only)."""
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData {
            report(code: $code) {
                events(dataType: Deaths, fightIDs: [$fightID], hostilityType: Friendlies, limit: 100) {
                    data
                }
            }
        }
    }
    """
    result = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    return result["reportData"]["report"]["events"].get("data", [])


def fetch_interrupt_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch interrupt events for a fight (friendly players), with pagination."""
    query = """
    query ($code: String!, $fightID: Int!, $startTime: Float) {
        reportData {
            report(code: $code) {
                events(dataType: Interrupts, fightIDs: [$fightID], hostilityType: Friendlies,
                       limit: 10000, startTime: $startTime) {
                    data
                    nextPageTimestamp
                }
            }
        }
    }
    """
    all_events = []
    variables = {"code": report_code, "fightID": fight_id, "startTime": None}
    while True:
        result = query_wcl(token, query, variables)
        page = result["reportData"]["report"]["events"]
        all_events.extend(page.get("data", []))
        next_ts = page.get("nextPageTimestamp")
        if not next_ts:
            break
        variables = {"code": report_code, "fightID": fight_id, "startTime": next_ts}
    return all_events


def analyze_interrupts(interrupt_events: list, actor_lookup: dict) -> dict:
    """Count interrupts performed by each player.
    Returns {pid: int}
    """
    totals = {}
    for event in interrupt_events:
        pid = event.get("sourceID")
        if pid is None or pid not in actor_lookup:
            continue
        totals[pid] = totals.get(pid, 0) + 1
    return totals


def analyze_boss_mechanics(damage_events: list, actor_lookup: dict, mechanics: list) -> dict:
    """Count per-player hits for each boss mechanic based on spell IDs.
    Returns {pid: {mechanic_label: count}}
    """
    if not mechanics:
        return {}
    spell_to_label = {}
    for mech in mechanics:
        for sid in mech["spell_ids"]:
            spell_to_label[sid] = mech["label"]
    result = {}
    for event in damage_events:
        if event.get("type") != "damage":
            continue
        pid = event.get("targetID")
        if pid not in actor_lookup:
            continue
        aid = event.get("abilityGameID", 0)
        label = spell_to_label.get(aid)
        if label is None:
            continue
        if pid not in result:
            result[pid] = {}
        result[pid][label] = result[pid].get(label, 0) + 1
    return result


def analyze_frontal_failures(damage_events: list, actor_lookup: dict, mechanics: list,
                              fight_start_ms: int = 0, window_ms: int = 2000,
                              cast_events: list = None) -> list:
    """Detect frontal failures: 2+ friendly players hit by the same frontal cast.
    Uses cast event targetID (if available) to identify who had the ability.
    Only checks mechanics with type='frontal'.
    Returns list of {time_str, label, primary_player, others, hit_count}.
    """
    frontal_mechs = [m for m in mechanics if m.get("type") == "frontal"]
    if not frontal_mechs:
        return []

    spell_to_label = {}
    for m in frontal_mechs:
        for sid in m["spell_ids"]:
            spell_to_label[sid] = m["label"]

    # Build cast-target lookup: spell_id -> sorted list of (timestamp, targetID)
    cast_targets = {}  # spell_id -> [(ts, targetID)]
    for event in (cast_events or []):
        aid = event.get("abilityGameID", 0)
        if aid not in spell_to_label:
            continue
        tid = event.get("targetID")
        if tid and tid in actor_lookup:
            cast_targets.setdefault(aid, []).append((event.get("timestamp", 0), tid))
    for v in cast_targets.values():
        v.sort()

    friendly_ids = set(actor_lookup.keys())
    hits = []
    for event in damage_events:
        if event.get("type") != "damage":
            continue
        aid = event.get("abilityGameID", 0)
        if aid not in spell_to_label:
            continue
        pid = event.get("targetID")
        if pid not in friendly_ids:
            continue
        hits.append({"ts": event.get("timestamp", 0), "pid": pid, "label": spell_to_label[aid], "aid": aid})

    if not hits:
        return []

    hits.sort(key=lambda h: h["ts"])

    def _cast_target_at(aid, ts):
        """Find the cast targetID closest before ts for this spell."""
        candidates = cast_targets.get(aid, [])
        result = None
        for c_ts, c_tid in candidates:
            if c_ts <= ts + window_ms:
                result = c_tid
            else:
                break
        return result

    failures = []
    i = 0
    while i < len(hits):
        win_start = hits[i]["ts"]
        label     = hits[i]["label"]
        aid       = hits[i]["aid"]
        pids_hit  = set()
        j = i
        while j < len(hits) and hits[j]["ts"] - win_start <= window_ms and hits[j]["label"] == label:
            pids_hit.add(hits[j]["pid"])
            j += 1
        if len(pids_hit) >= 2:
            elapsed = (win_start - fight_start_ms) / 1000
            m, s    = divmod(int(elapsed), 60)
            time_str = f"{m}:{s:02d}"
            # Prefer cast targetID as primary; fall back to first hit
            primary_pid  = _cast_target_at(aid, win_start) or hits[i]["pid"]
            primary_name = actor_lookup.get(primary_pid, {}).get("name", f"ID-{primary_pid}")
            others = sorted(actor_lookup.get(pid, {}).get("name", f"ID-{pid}")
                            for pid in pids_hit if pid != primary_pid)
            failures.append({"time_str": time_str, "label": label,
                             "primary_player": primary_name, "others": others,
                             "hit_count": len(pids_hit)})
        i = j if j > i else i + 1

    return failures


def analyze_vorasius_tank_swaps(damage_events: list, actor_lookup: dict,
                                 fight_start_ms: int, spec_roles: dict) -> list:
    """Detect Shadowclaw Slam tank swap failures for Vorasius.
    Returns list of {player, time_str, slam_num} when the same tank soaks 3+ consecutive slams.
    """
    SLAM_IDS = {1241808, 1281954, 1281906, 1272328}
    tank_hits = []
    for ev in damage_events:
        if ev.get("type") != "damage":
            continue
        if ev.get("abilityGameID", 0) not in SLAM_IDS:
            continue
        pid = ev.get("targetID")
        if pid not in actor_lookup:
            continue
        if spec_roles.get(pid) != "Tank":
            continue
        tank_hits.append({"ts": ev["timestamp"], "pid": pid})

    if not tank_hits:
        return []

    tank_hits.sort(key=lambda x: x["ts"])

    # Group into discrete slam events (hits within 1500ms = same slam instance)
    slams = [[tank_hits[0]]]
    for hit in tank_hits[1:]:
        if hit["ts"] - slams[-1][-1]["ts"] < 1500:
            slams[-1].append(hit)
        else:
            slams.append([hit])

    # One entry per slam: whichever tank was primarily hit
    slam_seq = [{"ts": s[0]["ts"], "pid": s[0]["pid"]} for s in slams]

    failures = []
    run_start = 0
    for i in range(1, len(slam_seq)):
        if slam_seq[i]["pid"] == slam_seq[i - 1]["pid"]:
            if i - run_start >= 2:  # 3rd+ consecutive hit on same tank
                pid      = slam_seq[i]["pid"]
                elapsed  = (slam_seq[i]["ts"] - fight_start_ms) / 1000
                m, s     = divmod(int(elapsed), 60)
                failures.append({
                    "player":   actor_lookup.get(pid, {}).get("name", f"#{pid}"),
                    "time_str": f"{m}:{s:02d}",
                    "slam_num": i + 1,
                })
        else:
            run_start = i

    return failures


def analyze_mechanic_timestamps(damage_events: list, actor_lookup: dict,
                                 fight_start_ms: int, mechanics: list) -> dict:
    """Track per-hit timestamps + damage for dmg_hits mechanics (used for table cell tooltips).
    Returns {pid: {label: [{"t": "0:23", "dmg": 45000}, ...]}}
    """
    spell_to_label = {}
    for m in mechanics:
        if m.get("type") in ("dmg_hits", "avoid", "soak"):
            for sid in m["spell_ids"]:
                spell_to_label[sid] = m["label"]
    if not spell_to_label:
        return {}
    result = {}
    for ev in damage_events:
        if ev.get("type") != "damage":
            continue
        pid = ev.get("targetID")
        if pid not in actor_lookup:
            continue
        label = spell_to_label.get(ev.get("abilityGameID", 0))
        if label is None:
            continue
        elapsed = (ev["timestamp"] - fight_start_ms) / 1000
        m2, s2  = divmod(int(elapsed), 60)
        dmg     = ev.get("amount", 0) + ev.get("absorbed", 0)
        result.setdefault(pid, {}).setdefault(label, []).append(
            {"t": f"{m2}:{s2:02d}", "dmg": dmg})
    return result


# Spell IDs for CC abilities that stop Fractured Image casts on Salhadaar.
# Includes hard stuns, silences, incapacitates, knockbacks that interrupt.
_SALHADAAR_CC_IDS = {
    # Death Knight
    108194,  # Asphyxiate
    47476,   # Strangulate
    49576,   # Death Grip
    # Demon Hunter
    179057,  # Chaos Nova
    202137,  # Sigil of Silence
    217832,  # Imprison
    # Druid
    33786,   # Cyclone
    99,      # Incapacitating Roar
    5211,    # Mighty Bash
    132469,  # Typhoon
    # Evoker
    360806,  # Sleep Walk
    357214,  # Landslide
    368970,  # Tail Swipe (racial)
    357215,  # Wing Buffet (racial)
    # Hunter
    187650,  # Freezing Trap
    19577,   # Intimidation
    109248,  # Binding Shot
    213691,  # Scatter Shot
    # Mage
    118,     # Polymorph
    113724,  # Ring of Frost
    31661,   # Dragon's Breath
    157997,  # Ice Nova
    157980,  # Supernova
    # Monk
    115078,  # Paralysis
    119381,  # Leg Sweep
    116844,  # Ring of Peace
    # Paladin
    853,     # Hammer of Justice
    115750,  # Blinding Light
    20066,   # Repentance
    # Priest
    8122,    # Psychic Scream
    64044,   # Psychic Horror
    88625,   # Holy Word: Chastise
    # Rogue
    1833,    # Cheap Shot
    408,     # Kidney Shot
    1776,    # Gouge
    2094,    # Blind
    # Shaman
    192058,  # Capacitor Totem
    51514,   # Hex
    # Warlock
    5782,    # Fear
    30283,   # Shadowfury
    6789,    # Mortal Coil
    19647,   # Spell Lock (Felhunter)
    # Warrior
    107570,  # Storm Bolt
    46968,   # Shockwave
    5246,    # Intimidating Shout
}

_SALHADAAR_CC_NAMES = {
    108194: "Asphyxiate",     47476: "Strangulate",      49576: "Death Grip",
    179057: "Chaos Nova",     202137: "Sigil of Silence", 217832: "Imprison",
    33786:  "Cyclone",        99:    "Incap. Roar",        5211:  "Mighty Bash",
    132469: "Typhoon",
    360806: "Sleep Walk",     357214: "Landslide",
    368970: "Tail Swipe",    357215: "Wing Buffet",
    187650: "Freezing Trap",  19577:  "Intimidation",     109248: "Binding Shot",
    213691: "Scatter Shot",
    118:    "Polymorph",      113724: "Ring of Frost",    31661:  "Dragon's Breath",
    157997: "Ice Nova",       157980: "Supernova",
    115078: "Paralysis",      119381: "Leg Sweep",        116844: "Ring of Peace",
    853:    "Hammer of Justice", 115750: "Blinding Light", 20066: "Repentance",
    8122:   "Psychic Scream", 64044:  "Psychic Horror",   88625:  "Holy Word: Chastise",
    1833:   "Cheap Shot",     408:    "Kidney Shot",       1776:   "Gouge",       2094: "Blind",
    192058: "Capacitor Totem", 51514: "Hex",
    5782:   "Fear",           30283:  "Shadowfury",        6789:   "Mortal Coil", 19647: "Spell Lock",
    107570: "Storm Bolt",     46968:  "Shockwave",          5246:   "Int. Shout",
}


def fetch_salhadaar_cc_events(token: str, report_code: str, fight_id: int,
                               fight_start: float, fight_end: float) -> list:
    """Fetch friendly crowd-control cast events targeting Fractured Images."""
    _FRACTURED_IMAGE_GAME_ID = 251791
    # Use a filter to only get CC spell IDs cast by friendlies
    id_list = ",".join(str(x) for x in sorted(_SALHADAAR_CC_IDS))
    filter_expr = f"ability.id in ({id_list}) and target.type = 'NPC'"
    query = """
    query ($code: String!, $fightID: Int!, $start: Float!, $end: Float!, $filter: String!) {
        reportData { report(code: $code) {
            events(dataType: Casts, fightIDs: [$fightID], hostilityType: Friendlies,
                   startTime: $start, endTime: $end, filterExpression: $filter, limit: 10000)
            { data nextPageTimestamp }
        }}
    }
    """
    all_events = []
    cursor = float(fight_start)
    while True:
        result = query_wcl(token, query, {
            "code": report_code, "fightID": fight_id,
            "start": cursor, "end": float(fight_end), "filter": filter_expr,
        })
        page = result["reportData"]["report"]["events"]
        # Only keep events targeting Fractured Image NPCs
        for ev in page.get("data", []):
            all_events.append(ev)
        nxt = page.get("nextPageTimestamp")
        if not nxt:
            break
        cursor = nxt
    return all_events


def analyze_salhadaar_fracture_waves(interrupt_events: list, actor_lookup: dict,
                                      fight_start_ms: int, ability_names: dict = None,
                                      spell_names_map: dict = None,
                                      cc_events: list = None) -> list:
    """Group friendly interrupts + CC on Fractured Images into waves for Fallen-King Salhadaar.
    Filters interrupts to spells with 'fracture' in the interrupted ability name.
    Returns list of {wave_num, wave_start_s, wave_end_s, players: {pid: [{spell, time_str}]}}.
    """
    ability_names   = ability_names  or {}
    spell_names_map = spell_names_map or {}
    cc_events       = cc_events or []

    relevant = []

    # Interrupts of Shadow Fracture
    for ev in interrupt_events:
        extra_id = ev.get("extraAbilityGameID", 0)
        if ability_names:
            interrupted_name = ability_names.get(extra_id, "").lower()
            if "fracture" not in interrupted_name:
                continue
        pid = ev.get("sourceID")
        if pid not in actor_lookup:
            continue
        spell_id   = ev.get("abilityGameID", 0)
        spell_name = spell_names_map.get(spell_id) or ability_names.get(spell_id, f"#{spell_id}")
        ts_s       = (ev["timestamp"] - fight_start_ms) / 1000
        m2, s2     = divmod(int(ts_s), 60)
        relevant.append({"ts_s": ts_s, "pid": pid, "spell": spell_name,
                          "time_str": f"{m2}:{s2:02d}"})

    # CC casts on Fractured Images
    for ev in cc_events:
        if ev.get("type") != "cast":
            continue
        pid = ev.get("sourceID")
        if pid not in actor_lookup:
            continue
        spell_id = ev.get("abilityGameID", 0)
        if spell_id not in _SALHADAAR_CC_IDS:
            continue
        spell_name = _SALHADAAR_CC_NAMES.get(spell_id, ability_names.get(spell_id, f"#{spell_id}"))
        ts_s       = (ev["timestamp"] - fight_start_ms) / 1000
        m2, s2     = divmod(int(ts_s), 60)
        relevant.append({"ts_s": ts_s, "pid": pid, "spell": spell_name,
                          "time_str": f"{m2}:{s2:02d}"})

    if not relevant:
        return []

    relevant.sort(key=lambda x: x["ts_s"])

    # Group into waves: gap > 20s = new wave
    waves_raw = [[relevant[0]]]
    for item in relevant[1:]:
        if item["ts_s"] - waves_raw[-1][-1]["ts_s"] <= 20:
            waves_raw[-1].append(item)
        else:
            waves_raw.append([item])

    result = []
    for wi, items in enumerate(waves_raw, 1):
        players: dict = {}
        for item in items:
            players.setdefault(item["pid"], []).append(
                {"spell": item["spell"], "time_str": item["time_str"]})
        result.append({
            "wave_num":     wi,
            "wave_start_s": items[0]["ts_s"],
            "wave_end_s":   items[-1]["ts_s"],
            "players":      players,
        })
    return result


def detect_salhadaar_wipe_events(boss_cast_events: list, fight_start_ms: int,
                                  ability_names: dict = None) -> list:
    """Detect Void Infusion casts (orb reached the boss) from enemy cast events.
    Returns list of {event_type, time_str}.
    """
    ability_names = ability_names or {}
    flags = []
    for ev in boss_cast_events:
        if ev.get("type") != "cast":
            continue
        aid  = ev.get("abilityGameID", 0)
        name = ability_names.get(aid, "").lower()
        if "void infusion" in name:
            elapsed = (ev["timestamp"] - fight_start_ms) / 1000
            m, s    = divmod(int(elapsed), 60)
            flags.append({"event_type": "Void Infusion", "time_str": f"{m}:{s:02d}"})
    return flags


def fetch_boss_cast_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch enemy cast events for a fight (boss casts), with pagination."""
    query = """
    query ($code: String!, $fightID: Int!, $startTime: Float) {
        reportData {
            report(code: $code) {
                events(dataType: Casts, fightIDs: [$fightID], hostilityType: Enemies,
                       limit: 10000, startTime: $startTime) {
                    data
                    nextPageTimestamp
                }
            }
        }
    }
    """
    all_events = []
    variables = {"code": report_code, "fightID": fight_id, "startTime": None}
    while True:
        result = query_wcl(token, query, variables)
        page = result["reportData"]["report"]["events"]
        all_events.extend(page.get("data", []))
        next_ts = page.get("nextPageTimestamp")
        if not next_ts:
            break
        variables = {"code": report_code, "fightID": fight_id, "startTime": next_ts}
    return all_events


def fetch_damage_taken_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch avoidable damage-taken events for a fight (friendly players only), with pagination."""
    query = """
    query ($code: String!, $fightID: Int!, $startTime: Float) {
        reportData {
            report(code: $code) {
                events(dataType: DamageTaken, fightIDs: [$fightID], hostilityType: Friendlies,
                       limit: 10000, startTime: $startTime) {
                    data
                    nextPageTimestamp
                }
            }
        }
    }
    """
    all_events = []
    variables = {"code": report_code, "fightID": fight_id, "startTime": None}
    while True:
        result = query_wcl(token, query, variables)
        page = result["reportData"]["report"]["events"]
        all_events.extend(page.get("data", []))
        next_ts = page.get("nextPageTimestamp")
        if not next_ts:
            break
        variables = {"code": report_code, "fightID": fight_id, "startTime": next_ts}
    return all_events


def fetch_damage_done_events(token: str, report_code: str, fight_id: int,
                             target_ids: set = None) -> list:
    """Fetch DamageDone events (friendly players dealing damage), with pagination.
    If target_ids is given, only return events where targetID is in that set.
    """
    query = """
    query ($code: String!, $fightID: Int!, $startTime: Float) {
        reportData {
            report(code: $code) {
                events(dataType: DamageDone, fightIDs: [$fightID], hostilityType: Friendlies,
                       limit: 10000, startTime: $startTime) {
                    data
                    nextPageTimestamp
                }
            }
        }
    }
    """
    all_events = []
    variables = {"code": report_code, "fightID": fight_id, "startTime": None}
    while True:
        result = query_wcl(token, query, variables)
        page = result["reportData"]["report"]["events"]
        events = page.get("data", [])
        if target_ids is not None:
            events = [e for e in events if e.get("targetID") in target_ids]
        all_events.extend(events)
        next_ts = page.get("nextPageTimestamp")
        if not next_ts:
            break
        variables = {"code": report_code, "fightID": fight_id, "startTime": next_ts}
    return all_events


def fetch_damage_table_for_target(token: str, report_code: str, fight_id: int,
                                   target_id: int,
                                   start_ms: float = None,
                                   end_ms: float = None) -> dict:
    """Fetch per-player DamageDone stats against a specific target NPC.
    Returns {sourceID: {"dmg": int, "active_ms": int}}.
    Uses the WCL table endpoint (same data as the WCL UI damage-done-to-target view).
    Optional start_ms/end_ms (relative to fight start) restrict the time window.
    """
    query = """
    query ($code: String!, $fightID: Int!, $targetID: Int!, $startTime: Float, $endTime: Float) {
        reportData {
            report(code: $code) {
                table(dataType: DamageDone, fightIDs: [$fightID], targetID: $targetID,
                      startTime: $startTime, endTime: $endTime)
            }
        }
    }
    """
    result = query_wcl(token, query, {"code": report_code, "fightID": fight_id,
                                      "targetID": target_id,
                                      "startTime": start_ms, "endTime": end_ms})
    table  = result["reportData"]["report"]["table"]
    entries = table.get("data", {}).get("entries", []) if isinstance(table, dict) else []
    return {e["id"]: {"dmg": e.get("total", 0), "active_ms": e.get("activeTime", 0)}
            for e in entries if "id" in e}


def fetch_npc_death_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch death events for enemy NPCs in a fight."""
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData {
            report(code: $code) {
                events(dataType: Deaths, fightIDs: [$fightID], hostilityType: Enemies, limit: 1000) {
                    data
                }
            }
        }
    }
    """
    result = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    page = result["reportData"]["report"]["events"]
    return page.get("data", [])


def fetch_uptime_table(token: str, report_code: str, fight_id: int) -> dict:
    """Fetch DPS active time + damage done per player from WCL table endpoint.
    Returns {sourceID: {"activeTime": ms, "total": damage_int}}.
    """
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData {
            report(code: $code) {
                table(dataType: DamageDone, fightIDs: [$fightID])
            }
        }
    }
    """
    result = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    table = result["reportData"]["report"]["table"]
    entries = table.get("data", {}).get("entries", []) if isinstance(table, dict) else []
    return {e["id"]: {"activeTime": e.get("activeTime", 0), "total": e.get("total", 0)}
            for e in entries if "id" in e}


def fetch_add_instance_damage(token: str, report_code: str,
                              fight_start_ms: int, fight_end_ms: int,
                              add_game_id: int) -> list:
    """Fetch per-player damage + active time per add wave (time window).

    Uses absolute report timestamps so pagination works correctly.
    Step 1: collect all add damage events to auto-detect wave windows (gap > 12s).
    Step 2: query WCL table per window to get WCL-computed activeTime + total damage.

    Returns list of {wave_num, wave_start_s, wave_end_s, players: {pid: {dmg, active_ms}}}.
    """
    # Step 1: paginate all add damage events with absolute timestamps
    ev_query = """
    query ($code: String!, $filter: String!, $start: Float!, $end: Float!) {
        reportData { report(code: $code) {
            events(dataType: DamageDone, startTime: $start, endTime: $end,
                   filterExpression: $filter, limit: 1000)
            { data nextPageTimestamp }
        }}
    }
    """
    filter_expr = f"target.id = {add_game_id}"
    all_timestamps = []
    cursor = float(fight_start_ms)
    while True:
        resp = query_wcl(token, ev_query, {
            "code": report_code, "filter": filter_expr,
            "start": cursor, "end": float(fight_end_ms),
        })
        ed = resp["reportData"]["report"]["events"]
        all_timestamps.extend(e["timestamp"] for e in ed["data"])
        nxt = ed.get("nextPageTimestamp")
        if not nxt or nxt >= fight_end_ms:
            break
        cursor = nxt

    if not all_timestamps:
        return []

    # Step 2: detect wave windows (gap > 12s between consecutive events)
    all_timestamps.sort()
    GAP_MS = 12_000
    windows = []
    w_start = all_timestamps[0]
    w_end   = all_timestamps[0]
    for ts in all_timestamps[1:]:
        if ts - w_end > GAP_MS:
            windows.append((w_start, w_end))
            w_start = ts
        w_end = ts
    windows.append((w_start, w_end))

    # Step 3: query WCL table per window (absolute startTime/endTime)
    tbl_query = """
    query ($code: String!, $filter: String!, $start: Float!, $end: Float!) {
        reportData { report(code: $code) {
            table(dataType: DamageDone, startTime: $start, endTime: $end,
                  filterExpression: $filter)
        }}
    }
    """
    results = []
    for i, (ws, we) in enumerate(windows, start=1):
        resp  = query_wcl(token, tbl_query, {
            "code": report_code, "filter": filter_expr,
            "start": float(ws), "end": float(we + 1000),  # +1s buffer
        })
        table   = resp["reportData"]["report"]["table"]
        entries = table.get("data", {}).get("entries", []) if isinstance(table, dict) else []
        players = {
            e["id"]: {"dmg": e.get("total", 0), "active_ms": e.get("activeTime", 0)}
            for e in entries if "id" in e
        }
        results.append({
            "wave_num":     i,
            "wave_start_s": (ws - fight_start_ms) // 1000,
            "wave_end_s":   (we - fight_start_ms) // 1000,
            "players":      players,
        })
    return results


def fetch_healing_table(token: str, report_code: str, fight_id: int) -> dict:
    """Fetch healing done per player from WCL table endpoint.
    Returns {sourceID: total_healing_int}.
    """
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData {
            report(code: $code) {
                table(dataType: Healing, fightIDs: [$fightID])
            }
        }
    }
    """
    result = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    table = result["reportData"]["report"]["table"]
    entries = table.get("data", {}).get("entries", []) if isinstance(table, dict) else []
    return {e["id"]: e.get("total", 0) for e in entries if "id" in e}


def fetch_rankings(token: str, report_code: str, fight_id: int) -> dict:
    """Fetch player parse percentiles for a fight.
    Returns {player_name_lower: rankPercent}.
    """
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData {
            report(code: $code) {
                rankings(fightIDs: [$fightID])
            }
        }
    }
    """
    result = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    rankings = result["reportData"]["report"]["rankings"]
    if not isinstance(rankings, dict):
        return {}
    out = {}
    for fight_entry in rankings.get("data", []):
        for role_data in fight_entry.get("roles", {}).values():
            for char in role_data.get("characters", []):
                name = char.get("name", "").lower()
                pct = char.get("rankPercent", 0)
                if name and pct:
                    out[name] = round(pct)
    return out


def fetch_fight_graph(token: str, report_code: str, fight_id: int) -> dict:
    """Fetch DamageDone, DamageTaken, and Healing time-series graphs for a fight in one call.
    Returns {dps: [(t_s, val), ...], taken: [...], heal: [...]}.
    Each point is (time_seconds_from_fight_start, total_value_in_5s_bucket).
    Returns empty series on any error.
    """
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData {
            report(code: $code) {
                dps:   graph(dataType: DamageDone,  fightIDs: [$fightID], hostilityType: Friendlies)
                taken: graph(dataType: DamageTaken, fightIDs: [$fightID], hostilityType: Friendlies)
                heal:  graph(dataType: Healing,     fightIDs: [$fightID], hostilityType: Friendlies)
            }
        }
    }
    """
    try:
        result = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
        rep    = result["reportData"]["report"]
    except Exception:
        return {"dps": [], "taken": [], "heal": []}

    def _sum_series(graph_raw) -> list:
        """Extract total time-series from WCL graph response → [(t_s, total), ...], time 0-based.
        WCL returns {data: {series: [{pointStart, pointInterval, data: [val, val, ...]}, ...]}}
        Uses the pre-summed 'Total' entry when available, otherwise sums all player entries.
        """
        if not isinstance(graph_raw, dict):
            return []
        inner = graph_raw.get("data", graph_raw)
        if not isinstance(inner, dict):
            return []
        series = inner.get("series", [])
        if not series:
            return []
        # Prefer the pre-summed Total entry
        src = next((e for e in series if e.get("type") == "Total"), None)
        if src is None:
            # Fall back: sum all non-Total entries element-wise
            non_total = [e for e in series if isinstance(e.get("data"), list)]
            if not non_total:
                return []
            src = non_total[0]
            extra = non_total[1:]
            vals = [v or 0 for v in src.get("data", [])]
            for e in extra:
                for i, v in enumerate(e.get("data", [])):
                    if i < len(vals):
                        vals[i] += v or 0
            src = {**src, "data": vals}
        point_start    = src.get("pointStart", 0)
        point_interval = src.get("pointInterval", 1000)
        interval_s     = point_interval / 1000
        data_vals      = src.get("data", [])
        if not data_vals:
            return []
        # Convert cumulative totals → per-second rate for each bucket
        result = []
        prev = 0
        for i, cum in enumerate(data_vals):
            cum = cum or 0
            delta = max(cum - prev, 0)
            rate  = round(delta / interval_s) if interval_s > 0 else 0
            t_s   = round((point_start + i * point_interval) / 1000, 1)
            result.append((t_s, rate))
            if cum > prev:
                prev = cum
        # Trim trailing zero-rate points (fight already over)
        while result and result[-1][1] == 0:
            result.pop()
        return result

    return {
        "dps":   _sum_series(rep.get("dps")),
        "taken": _sum_series(rep.get("taken")),
        "heal":  _sum_series(rep.get("heal")),
    }


def aggregate_damage_taken(damage_events: list, actor_lookup: dict) -> dict:
    """Sum total actual damage taken per player (post-mitigation, all sources).
    Returns {pid: total_damage_int}
    """
    totals = {}
    for event in damage_events:
        if event.get("type") != "damage":
            continue
        pid = event.get("targetID")
        if pid is None or pid not in actor_lookup:
            continue
        totals[pid] = totals.get(pid, 0) + event.get("amount", 0)
    return totals


_ALNDUST_SOAK_IDS   = {1262305, 1246827}   # Alndust Upheaval — players hit = "went down"
_GLOOM_SPELL_ID          = 1245391  # Gloom — Vaelgor & Ezzorak purple ball soak
_GLOOMTOUCHED_DEBUFF_ID  = 1245554  # Gloomtouched — debuff applied when a player soaks Gloom
_VOID_MARKED_DEBUFF_ID   = 1280023  # Void Marked — Imperator Averzian dispel assignment (confirmed from Mythic logs)
_CROWN_CIRCLE_ID    = 1233887              # P3 Void Circle debuff
_CROWN_SILVERSTRIKE_HIT_IDS  = {1233649, 1237729}  # Silverstrike damage events
_CROWN_VOID_STACK_ID         = 1233602             # Void Stacks debuff


def fetch_crown_debuff_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch Silverstrike + P3 circle + Stellar Emission debuff events for a Crown fight."""
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData { report(code: $code) {
            events(
                dataType: Debuffs,
                fightIDs: [$fightID],
                hostilityType: Friendlies,
                filterExpression: "ability.id in (1233887, 1233602, 1237038)",
                limit: 1000
            ) { data }
        }}
    }
    """
    resp = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    return resp["reportData"]["report"]["events"].get("data", [])


def fetch_crown_add_buff_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch Umbral Tether + Corruption Essence buff events on the 3 Crown adds."""
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData { report(code: $code) {
            events(
                dataType: Buffs,
                fightIDs: [$fightID],
                hostilityType: Enemies,
                filterExpression: "ability.id in (1233470, 1233778)",
                limit: 500
            ) { data }
        }}
    }
    """
    resp = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    return resp["reportData"]["report"]["events"].get("data", [])


def fetch_crown_cast_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch Silversunder Catastrophe (1234546) begincast events — marks intermission start."""
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData { report(code: $code) {
            events(
                dataType: Casts,
                fightIDs: [$fightID],
                hostilityType: Enemies,
                filterExpression: "ability.id = 1234546 and type = 'begincast'",
                limit: 50
            ) { data }
        }}
    }
    """
    resp = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    return resp["reportData"]["report"]["events"].get("data", [])


def analyze_crown_mechanics(debuff_events: list, damage_events: list,
                             actor_lookup: dict, spec_roles: dict,
                             fight_start_ms: int, add_buff_events: list = None,
                             cast_events: list = None) -> dict:
    """Analyze Crown of the Cosmos mechanics from raw events.

    Returns:
        silverstrike: list of intermission dicts, each with ordered list of arrow recipients
        p3_circles:   list of circle-set dicts, each with 3 players in leave order
    """
    def pid_name(pid):
        return actor_lookup.get(pid, {}).get("name", f"#{pid}")

    def pid_role(pid):
        return spec_roles.get(pid, "DPS")

    # ── Silverstrike targeting debuff (1233602) — who was ASSIGNED the mechanic ──
    # Applied to exactly 2 players simultaneously, ~6s before the arrow hits.
    assign_applies = sorted(
        [(e["timestamp"] - fight_start_ms, e["targetID"])
         for e in debuff_events
         if e.get("abilityGameID") == _CROWN_VOID_STACK_ID   # 1233602 reused as assign ID
         and e.get("type") == "applydebuff"
         and e.get("targetID") in actor_lookup],
        key=lambda x: x[0]
    )
    # Group into assignment rounds: applies within 500ms of each other = same round
    assign_rounds: list = []   # list of (t_ms, frozenset_of_pids)
    if assign_applies:
        cur_t, _  = assign_applies[0]
        cur_grp   = [assign_applies[0]]
        for t, pid in assign_applies[1:]:
            if t - cur_t > 500:
                assign_rounds.append((cur_t, frozenset(p for _, p in cur_grp)))
                cur_grp = [(t, pid)]
                cur_t   = t
            else:
                cur_grp.append((t, pid))
        assign_rounds.append((cur_t, frozenset(p for _, p in cur_grp)))

    def find_assigned_pids(round_first_t: int) -> frozenset:
        """Return the set of assigned pids whose debuff apply is within 8s of the round."""
        for at, apids in assign_rounds:
            if abs(at - round_first_t) <= 8_000:
                return apids
        return frozenset()

    # ── Stellar Emission (1237038) stack timeline per player ──
    # Debuff is consumed (removedebuff fires) right as the arrow hits — ignore removes
    # that happen within 500ms of a hit (caused BY the hit), use last apply stack count.
    _SE_ID = 1237038
    _se_sorted = sorted(
        [(e["timestamp"] - fight_start_ms, e["targetID"], e["type"], e.get("stack", None))
         for e in debuff_events if e.get("abilityGameID") == _SE_ID],
        key=lambda x: x[0]
    )

    def se_stacks_before_hit(pid: int, t_hit_ms: int) -> int:
        """Stellar Emission stacks on pid just before arrow hit at t_hit_ms.
        Removes within 500ms of t_hit_ms are treated as caused by the hit and ignored."""
        val = 0
        for t, tid, etype, stack in _se_sorted:
            if tid != pid or t > t_hit_ms:
                continue
            if etype == "applydebuff":
                val = 1
            elif etype == "applydebuffstack":
                val = stack if stack is not None else val + 1
            elif etype == "removedebuff" and t < t_hit_ms - 500:
                val = 0   # genuine prior remove (not caused by this hit)
        return val

    # ── Silverstrike hits (damage events for hit IDs) ──
    strike_hits = sorted(
        [(e["timestamp"] - fight_start_ms, e["targetID"], e.get("abilityGameID"))
         for e in damage_events
         if e.get("abilityGameID") in _CROWN_SILVERSTRIKE_HIT_IDS and e.get("targetID") in actor_lookup],
        key=lambda x: x[0]
    )
    # Group into rounds: hits within 5s of each other = same round
    rounds = []
    if strike_hits:
        current = [strike_hits[0]]
        for hit in strike_hits[1:]:
            if hit[0] - current[0][0] > 5_000:
                rounds.append(current)
                current = [hit]
            else:
                current.append(hit)
        rounds.append(current)
    intermissions = rounds  # keep variable name for downstream logic

    # ── Umbral Tether + Corruption Essence on adds ──
    # Add IDs: Morium=256, Demiar=257, Vorelus=260
    _ADD_NAMES = {256: "Morium", 257: "Demiar", 260: "Vorelus"}
    # Expected adds to lose Umbral Tether per round (1-indexed); round 3+ = no adds should need shields
    _IM_EXPECTED = {1: {256, 257}, 2: {260}}

    umbral_removals = []  # (t_ms, add_id)
    ce_removals     = []  # (t_ms, add_id)
    # CE stack timeline per add: {add_id: [(t_ms, stack_count), ...]} sorted by time
    ce_stack_timeline: dict = {aid: [] for aid in _ADD_NAMES}
    _ce_cur = {aid: 0 for aid in _ADD_NAMES}

    for e in sorted((add_buff_events or []), key=lambda x: x["timestamp"]):
        aid  = e.get("abilityGameID")
        tid  = e.get("targetID")
        etype = e.get("type")
        if tid not in _ADD_NAMES:
            continue
        t = e["timestamp"] - fight_start_ms
        if aid == 1233470 and etype == "removebuff":
            umbral_removals.append((t, tid))
        elif aid == 1233778:
            if etype == "removebuff":
                ce_removals.append((t, tid))
                _ce_cur[tid] = 0
            elif etype == "applybuff":
                _ce_cur[tid] = 1
            elif etype in ("applybuffstack", "removebuffstack"):
                _ce_cur[tid] = e.get("stack", _ce_cur[tid])
            ce_stack_timeline[tid].append((t, _ce_cur[tid]))

    def ce_stacks_at(add_id: int, t_ms: int) -> int:
        """Return CE stack count on add_id at time t_ms (last known value ≤ t_ms)."""
        count = 0
        for t, c in ce_stack_timeline.get(add_id, []):
            if t <= t_ms:
                count = c
            else:
                break
        return count

    # CE removal is "normal" if an Umbral Tether removal on the same add happened within 3s
    def ce_is_normal(t_ce, add_id):
        return any(abs(t_ce - t_u) <= 3000 and a == add_id for t_u, a in umbral_removals)

    # Build intermission summaries
    silverstrike = []
    _arrow_hit_cumulative: dict = {}   # pid -> total arrow hits across all rounds so far
    for im_idx, hits in enumerate(intermissions):
        seen = {}
        ordered = []
        first_t       = hits[0][0]
        spell_id      = hits[0][2]
        assigned_pids = find_assigned_pids(first_t)
        for t_ms, pid, _ in hits:
            seen[pid] = seen.get(pid, 0) + 1
            _arrow_hit_cumulative[pid] = _arrow_hit_cumulative.get(pid, 0) + 1
            if assigned_pids:
                is_extra = pid not in assigned_pids
            else:
                is_extra = (t_ms - first_t) > 50
            se_stacks = se_stacks_before_hit(pid, t_ms)
            ordered.append({"t_s": t_ms // 1000, "pid": pid,
                             "name": pid_name(pid), "role": pid_role(pid),
                             "seq": len([h for h in ordered if h["pid"] == pid]) + 1,
                             "is_extra": is_extra,
                             "se_stacks": se_stacks,
                             "arrow_total": _arrow_hit_cumulative[pid]})

        last_t = hits[-1][0]

        # Time window: ±5s around the round's first hit
        im_start = first_t - 5_000
        im_end   = first_t + 5_000

        # Umbral Tether removals that fall in this window
        shields_removed = [_ADD_NAMES[a] for t, a in umbral_removals if im_start <= t <= im_end]
        expected        = {_ADD_NAMES[a] for a in _IM_EXPECTED.get(im_idx + 1, set())}
        correct         = set(shields_removed) == expected

        # CE removals in this window that are NOT paired with an Umbral Tether removal
        wasted_ce = [_ADD_NAMES[a] for t, a in ce_removals
                     if im_start <= t <= im_end and not ce_is_normal(t, a)]

        # CE stacks on each add just before and just after the arrow round
        ce_before = {_ADD_NAMES[a]: ce_stacks_at(a, first_t - 50)  for a in _ADD_NAMES}
        ce_after  = {_ADD_NAMES[a]: ce_stacks_at(a, last_t  + 500) for a in _ADD_NAMES}

        assigned_list = [{"pid": p, "name": pid_name(p), "role": pid_role(p)}
                         for p in assigned_pids]
        silverstrike.append({
            "intermission":    im_idx + 1,
            "spell_id":        spell_id,
            "arrows":          ordered,
            "assigned":        assigned_list,
            "multi_hit":       {pid: cnt for pid, cnt in seen.items() if cnt > 1},
            "shields_removed": shields_removed,
            "expected_shields": sorted(expected),
            "correct":         correct,
            "wasted_ce":       wasted_ce,
            "ce_before":       ce_before,   # {add_name: stack_count} before arrow
            "ce_after":        ce_after,    # {add_name: stack_count} after arrow
        })

    # ── Intermission SE windows (dynamic per-fight timing) ──────────────────────
    # Group consecutive intermission (1237729) rounds into sessions (>60s gap = new intermission)
    _IM_SPELL = 1237729
    all_im_rounds = [r for r in intermissions if r[0][2] == _IM_SPELL]
    im_sessions: list = []
    if all_im_rounds:
        cur_session = [all_im_rounds[0]]
        for r in all_im_rounds[1:]:
            if r[0][0] - cur_session[-1][-1][0] > 60_000:
                im_sessions.append(cur_session)
                cur_session = [r]
            else:
                cur_session.append(r)
        im_sessions.append(cur_session)

    # Silversunder Catastrophe (1234546) begincast times — marks true intermission start
    _silversunder_casts = sorted(
        [(e["timestamp"] - fight_start_ms)
         for e in (cast_events or [])
         if e.get("abilityGameID") == 1234546],
    )

    interm_windows: list = []
    for session in im_sessions:
        first_t = session[0][0][0]
        last_t  = session[-1][-1][0]
        # Intermission start: nearest Silversunder Catastrophe cast before first arrow (within 120s)
        interm_start_t = next(
            (t for t in reversed(_silversunder_casts) if 0 <= first_t - t <= 120_000),
            None
        )
        # Window start: earliest SE applydebuff within 120s before first arrow
        window_start = first_t
        for t, tid, etype, stack in _se_sorted:
            if etype == "applydebuff" and 0 <= first_t - t <= 120_000:
                window_start = min(window_start, t)
        window_end = last_t + 2_000
        # Peak SE stacks per player during the window
        peak_se: dict = {}
        cur_se:  dict = {}
        for t, tid, etype, stack in _se_sorted:
            if t < window_start or t > window_end:
                continue
            if etype == "applydebuff":
                cur_se[tid] = 1
            elif etype == "applydebuffstack":
                cur_se[tid] = stack if stack is not None else cur_se.get(tid, 0) + 1
            elif etype == "removedebuff":
                cur_se[tid] = 0
            v = cur_se.get(tid, 0)
            if v > peak_se.get(tid, 0):
                peak_se[tid] = v
        interm_windows.append({
            "window_start":   window_start,
            "window_end":     window_end,
            "interm_start_t": interm_start_t,  # ms, or None if cast not found
            "interm_end_t":   last_t,
            "peak_se":        peak_se,
        })

    # ── Void stacks (debuff applies without being arrow target) ──
    void_stack_pids = set()
    for e in debuff_events:
        if e.get("abilityGameID") == _CROWN_VOID_STACK_ID and e.get("type") in ("applydebuff", "applydebuffstack"):
            void_stack_pids.add(e.get("targetID"))
    arrow_pids = {pid for hits in intermissions for _, pid, *_ in hits}
    # Players with void stacks but no arrow = leeching/unintended stacks
    no_arrow_stacks = [{"pid": p, "name": pid_name(p)} for p in void_stack_pids - arrow_pids if p in actor_lookup]

    # ── P3 Circles ──
    circle_events = sorted(
        [(e["timestamp"] - fight_start_ms, e["type"], e.get("targetID"))
         for e in debuff_events
         if e.get("abilityGameID") == _CROWN_CIRCLE_ID and e.get("targetID") in actor_lookup],
        key=lambda x: x[0]
    )
    # Build apply/remove pairs per player
    active_circles: dict = {}  # pid -> apply_t
    circle_sets = []
    current_set = []

    for t_ms, etype, pid in circle_events:
        if etype == "applydebuff":
            active_circles[pid] = t_ms
            current_set.append({"pid": pid, "apply_t": t_ms, "remove_t": None,
                                  "name": pid_name(pid), "role": pid_role(pid)})
        elif etype == "removedebuff" and pid in active_circles:
            apply_t = active_circles.pop(pid)
            for entry in reversed(current_set):
                if entry["pid"] == pid and entry["remove_t"] is None:
                    entry["remove_t"] = t_ms
                    entry["hold_ms"]  = t_ms - apply_t
                    break
            # Check if current_set is "complete" (all entries have remove_t) and non-empty
            if current_set and all(e.get("remove_t") is not None for e in current_set):
                # Flush set: sort by remove_t to get leave order
                current_set.sort(key=lambda e: e["remove_t"])
                circle_sets.append(current_set)
                current_set = []

    # Flush any remaining (last set if fight ended with circles active)
    if current_set:
        current_set.sort(key=lambda e: (e.get("remove_t") or float("inf")))
        circle_sets.append(current_set)

    # Flag sets where leave order doesn't match ranged→melee→tank
    p3_circles = []
    for s in circle_sets:
        flags = []
        for i, entry in enumerate(s):
            role = entry["role"]
            if i == 0 and role in ("Tank", "Melee"):
                flags.append(f'{entry["name"]} ({role}) left FIRST — should be ranged')
            elif i == len(s) - 1 and role not in ("Tank",) and len(s) == 3:
                flags.append(f'{entry["name"]} ({role}) left LAST — tank should be last')
        p3_circles.append({"players": s, "flags": flags})

    return {
        "silverstrike":    silverstrike,
        "no_arrow_stacks": no_arrow_stacks,
        "p3_circles":      p3_circles,
        "interm_windows":  interm_windows,
    }

# ── Chimaerus raid group config (hardcoded per report) ───────────────────────
# Group A (odd waves 1,3,5) goes down first; Group B (even waves 2,4,6) goes second.
# Player names must match ROSTER player names exactly.
# Update each week by sending a screenshot of the group assignments.
_CHIMAERUS_RAID_GROUPS: dict = {
    "cRCBrLapZQDgwNyX": {
        # G1(Nope,Doomkry,Mindhacker,Jimy,Madonis) + G3(Zodiacos,Hipe,Kutcher) + G5(Ice,Mostbanned,Nizze,Brunaine,Hypno)
        "group_a": {"Nope", "Doomkry", "Mindhacker", "Jimy", "Madonis",
                    "Zodiacos", "Hipe", "Kutcher",
                    "Ice", "Mostbanned", "Nizze", "Brunaine", "Hypno"},
        # G2(Beldryk,Malheiro,Toshiko,Kaze,Bolters) + G4(Uncleyoinky,Shamishan,Phyxius,Upyeah,Minxy) + G6(Potrenu,Zush,Tinet)
        "group_b": {"Beldryk", "Malheiro", "Toshiko", "Kaze", "Bolters",
                    "Uncleyoinky", "Shamishan", "Phyxius", "Upyeah", "Minxy",
                    "Potrenu", "Zush", "Tinet"},
    },
    "2pYxcGMHZtwqAnRD": {
        # G1(Nope,Doomkry,Mindhacker,Zush,Mostbanned) + G3(Phyxius,Malheiro,Kaze,Hipe,Madonis) + G5(Potrenu,Kutcher,Tinet)
        "group_a": {"Nope", "Doomkry", "Mindhacker", "Zush", "Mostbanned",
                    "Phyxius", "Malheiro", "Kaze", "Hipe", "Madonis",
                    "Potrenu", "Kutcher", "Tinet"},
        # G2(Ice,Gepeto,Upyeah,Beldryk,Toshiko) + G4(Minxy,Hypno,Bolters,Brunaine,Yoruichi) + G6(Zodiacos,Nizze,Pyxius)
        "group_b": {"Ice", "Gepeto", "Upyeah", "Beldryk", "Toshiko",
                    "Minxy", "Hypno", "Bolters", "Brunaine", "Yoruichi",
                    "Zodiacos", "Nizze", "Pyxius"},
    },
    # 6WXZfKwGzgckamPt — split 1
    "6WXZfKwGzgckamPt_1": {
        # G1(Phyxius,Uncleyoinky,Gëpeto,Hypno,Brunaineevo) + G3(Doomkry,?,Toshiko,Nype,Mostlock) + G5(Devert,Nizze,Tinet)
        "group_a": {"Phyxius", "Uncleyoinky", "Gëpeto", "Hypno", "Brunaineevo",
                    "Doomkry", "Toshiko", "Nype", "Mostlock",
                    "Devert", "Nizze", "Tinet"},
        # G2(Nøpæ,Rödinhas,Shamishan,Shaggyzard,Icecöld) + G4(Madonis,Beldrýk,Züsh,Minxymender,Kazeofscales) + G6(Zodiacos,Potrenu,Kutcherdh)
        "group_b": {"Nøpæ", "Rödinhas", "Shamishan", "Shaggyzard", "Icecöld",
                    "Madonis", "Beldrýk", "Züsh", "Minxymender", "Kazeofscales",
                    "Zodiacos", "Potrenu", "Kutcherdh"},
    },
    # 6WXZfKwGzgckamPt — split 2
    "6WXZfKwGzgckamPt_2": {
        # G1(Nøpæ,Allblues,Malheiro,Brunaineevo,Madhmag) + G3(Mindhacker,Holynooblal,Wype,Icecoldleap,Ipala) + G5(Mostbanned,Kutcherdh,Potrenu)
        "group_a": {"Nøpæ", "Allblues", "Malheiro", "Brunaineevo", "Madhmag",
                    "Mindhacker", "Holynooblal", "Wype", "Icecoldleap", "Ipala",
                    "Mostbanned", "Kutcherdh", "Potrenu"},
        # G2(Phyxy,Pumppaladin,Hypnodh,Samdracson,Madonismagus) + G4(Upyeah,Beldryc,Drunkminxy,Züsh,Kazehakase) + G6(Zodiacos,Nizzedk,Pingveryhigh)
        "group_b": {"Phyxy", "Pumppaladin", "Hypnodh", "Samdracson", "Madonismagus",
                    "Upyeah", "Beldryc", "Drunkminxy", "Züsh", "Kazehakase",
                    "Zodiacos", "Nizzedk", "Pingveryhigh"},
    },
    # h8j147qKHy9BMpkL — 2026-03-31 Mythic
    # Groups 1+2 down first, Groups 3+4 down second
    # Post-pull-3 config: Zodiacos swapped into group_b, Madhmag swapped into group_a
    "h8j147qKHy9BMpkL": {
        # G1(Nøpæ,Shamishan,Hypno,Upyeah,Nizze) + G2(Bolters/Devert,Potrenu,Minxy,Jimy/Madhmag,Hipe/Wype)
        "group_a": {"Nope", "Shamishan", "Hypno", "Upyeah", "Nizze",
                    "Bolters", "Potrenu", "Minxy", "Jimy", "Hipe"},
        # G3(Phyxius,Zush,Beldryk,Kutcher,Madonis) + G4(Uncleyoinky,Toshiko,Tinet,Ice,Zodiacos)
        "group_b": {"Phyxius", "Zush", "Beldryk", "Kutcher", "Madonis",
                    "Uncleyoinky", "Toshiko", "Tinet", "Ice", "Zodiacos"},
    },
}
_CHIMAERUS_BOSS_GAME_ID       = 256116
_CHIMAERUS_SMALL_ADD_GAME_IDS = {245555, 245575}  # Swarming Shade, Haunting Essence

def analyze_alndust_groups(damage_events: list, player_pids: set,
                            fight_start_ms: int = 0,
                            report_code: str = "",
                            actor_lookup: dict = None,
                            split_num: int = 1) -> list:
    """Detect who soaked each Alndust Upheaval wave (went 'down') on Chimaerus.
    If a hardcoded group config exists for this report (in _CHIMAERUS_RAID_GROUPS),
    uses it to determine expected groups per wave. Otherwise falls back to N-2 reference.
    """
    hits = []
    for e in damage_events:
        if e.get("abilityGameID") in _ALNDUST_SOAK_IDS and e.get("targetID") in player_pids:
            t_ms = e.get("timestamp", 0) - fight_start_ms
            hits.append((t_ms, e["targetID"]))
    if not hits:
        return []
    hits.sort()
    # Group into waves: gap > 10 s between events = new wave
    waves = [[hits[0]]]
    for h in hits[1:]:
        if h[0] - waves[-1][-1][0] <= 10_000:
            waves[-1].append(h)
        else:
            waves.append([h])

    # Build PID→player-name map and hardcoded group sets (PIDs) if config exists
    group_a_pids: set = set()
    group_b_pids: set = set()
    use_hardcoded = False
    split_key = f"{report_code}_{split_num}"
    lookup_key = split_key if split_key in _CHIMAERUS_RAID_GROUPS else report_code
    if report_code and actor_lookup and lookup_key in _CHIMAERUS_RAID_GROUPS:
        cfg = _CHIMAERUS_RAID_GROUPS[lookup_key]
        for pid, actor in actor_lookup.items():
            char = actor.get("name", "")
            pname, _ = lookup_roster(char)
            if pname in cfg["group_a"]:
                group_a_pids.add(pid)
            elif pname in cfg["group_b"]:
                group_b_pids.add(pid)
        use_hardcoded = True

    # Pre-pass: detect which wave indices are cycle resets (same group as previous wave)
    # Mythic pattern: G1, G2, G1(intermission), G1(new cycle), G2 ...
    # The "new cycle" wave has high overlap with the previous wave (consecutive same group).
    cycle_reset_indices: set = set()
    _prev_down: set = set()
    for wi, wave in enumerate(waves):
        _ds = set(pid for _, pid in wave)
        if wi > 0 and _prev_down:
            jaccard = len(_ds & _prev_down) / max(len(_ds | _prev_down), 1)
            if jaccard > 0.5:
                cycle_reset_indices.add(wi)
        _prev_down = _ds
    # The wave just before each cycle reset is the "intermission" wave
    intermission_indices: set = {wi - 1 for wi in cycle_reset_indices if wi > 0}

    result = []
    consecutive_ups: dict = {pid: 0 for pid in player_pids}
    # Cycle position tracking: 1 = group_a expected, 2 = group_b expected
    cycle_pos   = 1
    last_at_pos: dict = {}   # cycle_pos → last result index where that pos was used (fallback)

    for wi, wave in enumerate(waves):
        wave_num  = wi + 1
        down_pids = list(dict.fromkeys(pid for _, pid in wave))
        down_set  = set(down_pids)
        up_pids   = [pid for pid in sorted(player_pids) if pid not in down_set]
        t_s       = wave[0][0] // 1000
        is_cycle_reset    = wi in cycle_reset_indices
        is_intermission   = wi in intermission_indices

        # On a cycle reset, restart cycle position back to 1
        if is_cycle_reset:
            cycle_pos = 1

        if use_hardcoded:
            expected    = group_a_pids if cycle_pos == 1 else group_b_pids
            wrong_group = group_b_pids if cycle_pos == 1 else group_a_pids
            missed_pids    = [p for p in expected if p not in down_set and p in player_pids]
            wrong_pids     = [p for p in down_set if p in wrong_group]
            double_up_pids = [p for p in up_pids if consecutive_ups.get(p, 0) >= 1 and p in expected]
        elif is_cycle_reset or cycle_pos not in last_at_pos:
            # First time at this position or post-intermission reset — no reference yet
            missed_pids    = []
            wrong_pids     = []
            double_up_pids = []
        else:
            # Compare to the last wave where this cycle_pos was active
            ref = result[last_at_pos[cycle_pos]]
            expected       = set(ref["down_pids"])
            missed_pids    = [p for p in expected if p not in down_set]
            wrong_pids     = [p for p in down_set if p not in expected
                               and consecutive_ups.get(p, 0) < 2]
            double_up_pids = [p for p in up_pids if consecutive_ups.get(p, 0) >= 1]

        last_at_pos[cycle_pos] = wi
        # Advance cycle position for next wave
        cycle_pos = 2 if cycle_pos == 1 else 1

        for pid in player_pids:
            consecutive_ups[pid] = 0 if pid in down_set else consecutive_ups.get(pid, 0) + 1

        result.append({
            "wave": wave_num, "t_s": t_s,
            "down_pids": down_pids, "up_pids": up_pids,
            "missed_pids": missed_pids, "wrong_pids": wrong_pids,
            "double_up_pids": double_up_pids,
            "is_intermission": is_intermission,
            "is_cycle_reset":  is_cycle_reset,
        })
    return result


def fetch_gloom_debuff_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch Gloomtouched (1245554) applydebuff events for a Vaelgor & Ezzorak fight."""
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData { report(code: $code) {
            events(
                dataType: Debuffs,
                fightIDs: [$fightID],
                hostilityType: Friendlies,
                filterExpression: "ability.id = 1245554 and type = 'applydebuff'",
                limit: 500
            ) { data }
        }}
    }
    """
    resp = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    return resp["reportData"]["report"]["events"].get("data", [])


def analyze_gloom_soaks(debuff_events: list, player_pids: set,
                         fight_start_ms: int = 0) -> list:
    """Detect who soaked each Gloom cast on Vaelgor & Ezzorak via Gloomtouched debuff."""
    hits = []
    for e in debuff_events:
        if e.get("abilityGameID") == _GLOOMTOUCHED_DEBUFF_ID and e.get("targetID") in player_pids:
            t_ms = e.get("timestamp", 0) - fight_start_ms
            hits.append((t_ms, e["targetID"]))
    if not hits:
        return []
    hits.sort()
    # Group into waves: gap > 20 s = new Gloom cast
    waves = [[hits[0]]]
    for h in hits[1:]:
        if h[0] - waves[-1][-1][0] <= 20_000:
            waves[-1].append(h)
        else:
            waves.append([h])
    result = []
    for wi, wave in enumerate(waves):
        seen: set = set()
        hit_pids: list = []
        for _, pid in wave:
            if pid not in seen:
                hit_pids.append(pid)
                seen.add(pid)
        t_s = wave[0][0] // 1000
        result.append({"wave": wi + 1, "t_s": t_s, "hit_pids": hit_pids})
    return result


def fetch_void_marked_events(token: str, report_code: str, fight_id: int) -> list:
    """Fetch Void Marked (1280023) applydebuff + removedebuff events for an Averzian fight."""
    query = """
    query ($code: String!, $fightID: Int!) {
        reportData { report(code: $code) {
            events(
                dataType: Debuffs,
                fightIDs: [$fightID],
                hostilityType: Friendlies,
                filterExpression: "ability.id = 1280023",
                limit: 500
            ) { data }
        }}
    }
    """
    resp = query_wcl(token, query, {"code": report_code, "fightID": fight_id})
    return resp["reportData"]["report"]["events"].get("data", [])


def analyze_void_marked(debuff_events: list, player_pids: set,
                         fight_start_ms: int = 0) -> list:
    """Pair each Void Marked application with its removal/dispel.
    Returns list of dicts: target_pid, t_s, dispeller_pid (None=expired), dispel_t_s.
    """
    apps    = []  # (t_ms, target_pid)
    removes = []  # (t_ms, target_pid, source_pid)
    for e in debuff_events:
        if e.get("abilityGameID") != _VOID_MARKED_DEBUFF_ID:
            continue
        t_ms   = e.get("timestamp", 0) - fight_start_ms
        etype  = e.get("type", "")
        target = e.get("targetID")
        source = e.get("sourceID")
        if etype == "applydebuff" and target in player_pids:
            apps.append((t_ms, target))
        elif etype == "removedebuff" and target in player_pids:
            removes.append((t_ms, target, source))

    apps.sort()
    removes.sort()

    result = []
    used_removes = set()
    for app_t, target_pid in apps:
        # Find the earliest matching removedebuff for this target after the application
        dispel_t_ms    = None
        dispeller_pid  = None
        for ri, (rem_t, rem_target, rem_source) in enumerate(removes):
            if ri in used_removes:
                continue
            if rem_target == target_pid and rem_t >= app_t:
                dispel_t_ms   = rem_t
                # source != target and source in players = external dispel
                dispeller_pid = rem_source if (rem_source != target_pid and rem_source in player_pids) else None
                used_removes.add(ri)
                break
        result.append({
            "t_s":          app_t // 1000,
            "target_pid":   target_pid,
            "dispel_t_s":   dispel_t_ms // 1000 if dispel_t_ms is not None else None,
            "dispeller_pid": dispeller_pid,
        })
    return result


def analyze_chimaerus_horror_waves(
    token: str,
    report_code: str,
    fight_id: int,
    alndust_groups: list,
    horror_actor_ids: set,        # set of masterData actor IDs for all Horror instances
    actor_lookup: dict,
    fight_start_ms: int,
    fight_end_ms: int,
) -> list:
    """Analyze per-wave Colossal Horror performance on Chimaerus.

    For each wave window, queries the WCL table endpoint per Horror actor with
    startTime/endTime to handle actor reuse across waves.

    Returns list of wave dicts:
      wave, t_s, down_pids, up_pids,
      kill_time_s (seconds from wave start to Horror death, or None),
      per_player: {pid: {phase, dmg, active_ms, boss_dmg}}
    """
    if not alndust_groups or not horror_actor_ids:
        return []

    def rel_ms(ev):
        return ev.get("timestamp", 0) - fight_start_ms

    # Build wave windows (ms, fight-relative — used for rel_ms comparisons).
    # Also build report-relative windows for WCL table API startTime/endTime.
    windows = []
    windows_abs = []  # report-relative ms (fight_start_ms + fight-relative)
    for i, wave in enumerate(alndust_groups):
        start = wave["t_s"] * 1000
        end   = alndust_groups[i + 1]["t_s"] * 1000 if i + 1 < len(alndust_groups) else float("inf")
        windows.append((start, end))
        end_abs = (fight_start_ms + end) if end != float("inf") else fight_end_ms
        windows_abs.append((fight_start_ms + start, end_abs))

    # Get Horror NPC death events for kill_time_s tracking.
    # Actors may be reused across waves (same actor ID for wave 1 and wave 3, etc.),
    # so we track (actor_id, wave_index) -> death_ms using sorted death events.
    npc_deaths = fetch_npc_death_events(token, report_code, fight_id)
    wave_kill_ms: dict = {}   # wave_index -> death ms (relative to fight start)
    for ev in sorted(npc_deaths, key=rel_ms):
        tid = ev.get("targetID")
        if tid not in horror_actor_ids:
            continue
        t = rel_ms(ev)
        for wi, (ws, we) in enumerate(windows):
            if ws <= t < we and wi not in wave_kill_ms:
                wave_kill_ms[wi] = t
                break

    # Fetch per-player table stats per wave using time-windowed queries.
    # Querying each (actor_id, wave_window) handles actor reuse across waves.
    wave_stats: dict = {}  # wave_index -> {pid: {dmg, active_ms}}
    for wi, (ws_abs, we_abs) in enumerate(windows_abs):
        end_param = we_abs if we_abs != float("inf") else None
        for actor_id in horror_actor_ids:
            tbl = fetch_damage_table_for_target(
                token, report_code, fight_id, actor_id,
                start_ms=ws_abs, end_ms=end_param,
            )
            for pid, stats in tbl.items():
                if wi not in wave_stats:
                    wave_stats[wi] = {}
                if pid not in wave_stats[wi]:
                    wave_stats[wi][pid] = {"dmg": 0, "active_ms": 0}
                wave_stats[wi][pid]["dmg"]       += stats["dmg"]
                wave_stats[wi][pid]["active_ms"] += stats["active_ms"]

    result = []
    for i, wave in enumerate(alndust_groups):
        ws_ms, we_ms = windows[i]
        we_ms_clamped = we_ms if we_ms != float("inf") else (fight_end_ms - fight_start_ms)
        wave_dur_ms   = we_ms_clamped - ws_ms
        down_pids    = set(wave["down_pids"])
        up_pids      = set(wave["up_pids"])

        # Horror kill time: seconds from wave start to Horror death
        kill_time_s = None
        if i in wave_kill_ms:
            kill_time_s = round((wave_kill_ms[i] - ws_ms) / 1000, 1)

        per_player = {}
        stats_for_wave = wave_stats.get(i, {})
        for pid in down_pids | up_pids:
            phase  = "down" if pid in down_pids else "up"
            pstats = stats_for_wave.get(pid, {"dmg": 0, "active_ms": 0})
            per_player[pid] = {
                "phase":     phase,
                "dmg":       pstats["dmg"],
                "active_ms": pstats["active_ms"],
            }

        result.append({
            "wave":          wave["wave"],
            "t_s":           wave["t_s"],
            "wave_dur_ms":   wave_dur_ms,
            "down_pids":     wave["down_pids"],
            "up_pids":       wave["up_pids"],
            "kill_time_s":   kill_time_s,
            "per_player":    per_player,
        })

    return result


def analyze_avoidable_damage(damage_events: list, actor_lookup: dict,
                             fight_start_ms: int = 0, ability_names: dict = None,
                             player_max_hp: dict = None, player_roles: dict = None) -> dict:
    """Count enemy-source damage-taken hits per player (proxy for avoidable damage).
    Returns {pid: {"hits": int, "big_hits": int, "details": [{"ability", "amount_k", "time"}]}}
    big_hits threshold: 50% HP for Tanks, 10% HP for everyone else.
    Melee auto-attacks (abilityGameID == 1) are excluded from details.
    """
    ability_names = ability_names or {}
    player_max_hp  = player_max_hp  or {}
    player_roles   = player_roles   or {}
    friendly_ids   = set(actor_lookup.keys())
    results = {}
    for event in sorted(damage_events, key=lambda e: e.get("timestamp", 0)):
        if event.get("type") != "damage":
            continue
        pid = event.get("targetID")
        if pid is None or pid not in actor_lookup:
            continue
        if event.get("sourceID") in friendly_ids:
            continue
        ability_id = event.get("abilityGameID", 0)
        if ability_id == 1:  # skip melee auto-attacks
            continue
        total_hit = event.get("unmitigatedAmount", 0) or event.get("amount", 0)
        if total_hit == 0:
            continue
        ability_name = ability_names.get(ability_id, f"#{ability_id}")
        ts = event.get("timestamp", 0)
        elapsed_ms = ts - fight_start_ms
        time_str = f"{int(elapsed_ms // 60000)}:{int((elapsed_ms % 60000) // 1000):02d}"
        amount_k = f"{total_hit / 1000:.1f}k"
        entry = results.setdefault(pid, {"hits": 0, "big_hits": 0, "details": []})
        entry["hits"] += 1
        max_hp = player_max_hp.get(pid, 0)
        role = player_roles.get(pid, "DPS")
        threshold = 0.50 if role == "Tank" else 0.10
        if max_hp > 0 and total_hit >= max_hp * threshold:
            entry["big_hits"] += 1
            entry["details"].append({"ability": ability_name, "amount_k": amount_k, "time": time_str})
    return results


# How many deaths before we stop counting (wipe cascade filter)
WIPE_DEATH_THRESHOLD = 4


def analyze_deaths(death_events: list, fight_start_ms: int, ability_names: dict = None,
                   fight_end_ms: int = 0, damage_events: list = None) -> dict:
    """Analyze death events, ignoring deaths after the Nth death (wipe cascade).
    Returns {targetID: [{"time", "ability", "fight_pct", "timeline"}, ...]}
    timeline = last 5s of damage hits before death: [{"ability", "amount_k", "sec_before"}, ...]
    """
    results = {}
    total_deaths = 0
    ability_names = ability_names or {}

    # Build per-player damage index for fast 5s window lookup
    dmg_by_player = {}
    for ev in (damage_events or []):
        if ev.get("type") != "damage":
            continue
        pid = ev.get("targetID")
        if pid is None:
            continue
        dmg_by_player.setdefault(pid, []).append(ev)

    for event in sorted(death_events, key=lambda e: e.get("timestamp", 0)):
        if total_deaths >= WIPE_DEATH_THRESHOLD:
            break
        target_id = event.get("targetID")
        if target_id is None:
            continue
        total_deaths += 1
        ts_ms = event.get("timestamp", 0)
        relative_sec = (ts_ms - fight_start_ms) / 1000.0
        minutes = int(relative_sec // 60)
        seconds = int(relative_sec % 60)
        time_str = f"{minutes}:{seconds:02d}"
        killing_id = event.get("killingAbilityGameID", 0)
        killing_name = ability_names.get(killing_id, f"ID:{killing_id}" if killing_id else "Unknown")
        fight_elapsed = ts_ms - fight_start_ms
        pct = round(fight_elapsed / fight_end_ms * 100) if fight_end_ms > 0 else 0

        # Last 5 seconds of damage events before this death
        timeline = []
        window_start = ts_ms - 5000
        for dev in dmg_by_player.get(target_id, []):
            dts = dev.get("timestamp", 0)
            if window_start <= dts <= ts_ms:
                daid = dev.get("abilityGameID", 0)
                dname = ability_names.get(daid, f"#{daid}")
                amt = dev.get("amount", 0)
                sec_before = (ts_ms - dts) / 1000.0
                timeline.append({"ability": dname, "amount_k": f"{amt/1000:.0f}k" if amt >= 1000 else str(amt), "sec_before": sec_before})
        timeline.sort(key=lambda x: x["sec_before"], reverse=True)  # oldest first

        results.setdefault(target_id, []).append({
            "time": time_str, "ability": killing_name, "fight_pct": pct, "timeline": timeline,
        })

    return results


# ─── Tracked Spells — loaded from tracked_spells.py ─────────────────────────

# (imported at top of file from tracked_spells.py)


def classify_spell(ability_game_id: int) -> str | None:
    """Classify a spell by its game ID into a category, or None if not tracked."""
    if ability_game_id in HEALTHSTONE_IDS:
        return "Healthstone"
    if ability_game_id in HEALTH_POT_IDS:
        return "Health"
    if ability_game_id in COMBAT_POT_IDS:
        return "Combat Pot"
    if ability_game_id in CLASS_DEFENSIVE_IDS:
        return "Defensive"
    if ability_game_id in CLASS_EXTERNAL_IDS:
        return "External"
    return None


def analyze_fight_casts(cast_events: list, fight_start_time: float, actors: dict) -> dict:
    """Analyze cast events for tracked spells. Returns {sourceID: [{spell, category, timestamp}]}."""
    results = {}
    for event in cast_events:
        if event.get("type") != "cast":
            continue
        ability_id = event.get("abilityGameID")
        if ability_id is None:
            continue
        category = classify_spell(ability_id)
        if category is None:
            continue

        source_id = event.get("sourceID")
        if source_id is None:
            continue
        
        # Timestamp relative to fight start (in seconds)
        ts_ms = event.get("timestamp", 0)
        relative_sec = (ts_ms - fight_start_time) / 1000.0
        minutes = int(relative_sec // 60)
        seconds = int(relative_sec % 60)
        time_str = f"{minutes}:{seconds:02d}"
        
        if source_id not in results:
            results[source_id] = []
        results[source_id].append({
            "spell": SPELL_NAMES.get(ability_id, f"ID:{ability_id}"),
            "category": category,
            "time": time_str,
            "timestamp_ms": ts_ms,
        })
    
    return results


# ─── Gear Analysis ───────────────────────────────────────────────────────────

def detect_craft_quality(bonus_ids: list) -> int | None:
    """Return crafting quality rank (1-5) if item has a craft quality bonus ID."""
    for bid in bonus_ids:
        if bid in CRAFT_QUALITY_BONUS_IDS:
            return CRAFT_QUALITY_BONUS_IDS[bid]
    return None


def is_crafted(bonus_ids: list) -> bool:
    """Check if an item is crafted based on bonus IDs."""
    if any(bid in CRAFTED_INDICATOR_BONUS_IDS for bid in bonus_ids):
        return True
    if detect_craft_quality(bonus_ids) is not None:
        return True
    return False


def classify_ilvl_tier(ilvl: int, quality: int) -> str:
    """Classify an item's crafting tier based on ilvl."""
    if quality >= 4:  # Epic
        if ilvl >= 272:
            return "Myth (Spark + Myth Crests)"
        elif ilvl >= 259:
            return "Hero (Spark + Hero Crests)"
        elif ilvl >= 252:
            return "Epic Base (Spark, no crests)"
        else:
            return f"Epic (ilvl {ilvl})"
    else:  # Rare
        if ilvl >= 233:
            return "Rare (Veteran Crests)"
        elif ilvl >= 214:
            return "Rare (Adventurer Crests)"
        else:
            return "Rare Base (no spark)"


def estimate_spark_usage(ilvl: int, quality: int, slot: int) -> str:
    """Estimate spark usage based on ilvl and quality."""
    if quality < 4:  # Rare items don't use sparks
        return "No"
    if ilvl >= 252:
        if slot in (15, 16):  # Weapons
            return "Yes (2H = 4 sparks)" if ilvl >= 252 else "Yes (2 sparks)"
        return "Yes (2 sparks)"
    return "No"


def analyze_players(actors: list, combatant_events: list) -> list:
    """Analyze all players' gear using masterData actors + combatant events."""
    records = []
    
    # Build actor lookup from masterData: id -> actor info
    actor_lookup = {}
    for actor in actors:
        actor_lookup[actor["id"]] = {
            "name": actor.get("name", "Unknown"),
            "class": actor.get("subType", actor.get("type", "Unknown")),
            "server": actor.get("server", "Unknown"),
        }
    
    # Build gear lookup from combatant events: sourceID -> gear list
    gear_lookup = {}
    for event in combatant_events:
        source_id = event.get("sourceID")
        gear = event.get("gear", [])
        if source_id is not None and gear:
            # Keep the most complete gear entry per player
            if source_id not in gear_lookup or len(gear) > len(gear_lookup.get(source_id, [])):
                gear_lookup[source_id] = gear
    
    print(f"[DEBUG] Actors from masterData: {len(actor_lookup)}")
    print(f"[DEBUG] Players with gear from events: {len(gear_lookup)}")
    
    # For each player that has gear events, analyze their gear
    for source_id, gear in gear_lookup.items():
        actor = actor_lookup.get(source_id, {"name": f"Unknown-{source_id}", "class": "Unknown", "server": "Unknown"})
        player_has_crafted = False
        
        for idx, item in enumerate(gear):
            if not item or not isinstance(item, dict):
                continue
            
            item_id = item.get("id", 0)
            if item_id == 0:
                continue
            
            ilvl = item.get("itemLevel", 0)
            quality = item.get("quality", 0)
            bonus_ids = item.get("bonusIDs", []) or []
            slot = idx
            
            if is_crafted(bonus_ids):
                player_has_crafted = True
                craft_rank = detect_craft_quality(bonus_ids)
                tier = classify_ilvl_tier(ilvl, quality)
                spark = estimate_spark_usage(ilvl, quality, slot)
                
                records.append({
                    "player": actor["name"],
                    "class": actor["class"],
                    "server": actor["server"],
                    "slot": SLOT_NAMES.get(slot, f"Slot {slot}"),
                    "item_id": item_id,
                    "item_level": ilvl,
                    "quality": "Epic" if quality >= 4 else "Rare" if quality >= 3 else "Uncommon",
                    "craft_rank": f"Rank {craft_rank}" if craft_rank else "Unknown",
                    "tier": tier,
                    "spark_used": spark,
                    "bonus_ids": ", ".join(str(b) for b in bonus_ids),
                })
        
        if not player_has_crafted:
            records.append({
                "player": actor["name"],
                "class": actor["class"],
                "server": actor["server"],
                "slot": "—",
                "item_id": 0,
                "item_level": 0,
                "quality": "—",
                "craft_rank": "—",
                "tier": "NO CRAFTED GEAR FOUND",
                "spark_used": "—",
                "bonus_ids": "",
            })
    
    return records
    
    return records
    
    return records


# ─── HTML Output ─────────────────────────────────────────────────────────────

WOWHEAD_SCRIPTS = """\
<script>const whTooltips = {colorLinks: true, iconizeLinks: true, renameLinks: true};</script>
<script src="https://wow.zamimg.com/js/tooltips.js"></script>"""


def _ilvl_color(ilvl: int) -> str:
    if ilvl >= 259: return "#ffd700"   # max crafted — gold
    if ilvl >= 246: return "#a335ee"   # upper tier — epic purple
    if ilvl >= 239: return "#1eff00"   # mid tier — uncommon green
    return "#9d9d9d"                   # low — gray


def _build_gear_html(records: list):
    """Build gear table content (header, rows, stats) with mains/alts separation."""
    chars: dict = {}
    for rec in records:
        p = rec["player"]
        if p not in chars:
            _, main_alt = lookup_roster(p)
            chars[p] = {"class": rec["class"], "items": [], "is_alt": main_alt == "Alt"}
        if rec["item_id"] != 0:
            chars[p]["items"].append(rec)

    max_items     = max((len(c["items"]) for c in chars.values()), default=2)
    max_items     = max(max_items, 2)
    total_players = len(chars)
    total_crafted = sum(len(c["items"]) for c in chars.values())
    no_craft      = sum(1 for c in chars.values() if not c["items"])
    total_cols    = 1 + max_items * 5

    def _row(pname, pdata):
        cls_color = CLASS_COLORS.get(pdata["class"], "#ccc")
        row_class = "no-craft" if not pdata["items"] else ""
        row = f'<tr class="{row_class}"><td class="player-cell" style="color:{cls_color}">{escape(pname)}</td>'
        for i in range(max_items):
            div = " divider" if i > 0 else ""
            if i < len(pdata["items"]):
                item      = pdata["items"][i]
                iid       = item["item_id"]
                ilvl      = item["item_level"]
                ic        = _ilvl_color(ilvl)
                spark_cls = "spark-yes" if "Yes" in str(item["spark_used"]) else ""
                spark_txt = "Yes" if "Yes" in str(item["spark_used"]) else "No"
                item_link = f'<a href="https://www.wowhead.com/item={iid}" data-wowhead="item={iid}" target="_blank">#{iid}</a>'
                row += f'<td class="item-cell{div}">{item_link}</td>'
                row += f'<td>{escape(item["slot"])}</td>'
                row += f'<td class="center" style="color:{ic};font-weight:600">{ilvl}</td>'
                row += f'<td>{escape(item["craft_rank"])}</td>'
                row += f'<td class="center {spark_cls}">{spark_txt}</td>'
            else:
                row += f'<td class="empty{div}">—</td>' + '<td class="empty">—</td>' * 4
        row += "</tr>"
        return row

    sort_key = lambda x: (-len(x[1]["items"]), x[0].lower())
    mains = {p: d for p, d in chars.items() if not d["is_alt"]}
    alts  = {p: d for p, d in chars.items() if d["is_alt"]}

    rows = "".join(_row(p, d) for p, d in sorted(mains.items(), key=sort_key))
    if alts:
        rows += f'<tr class="section-sep"><td colspan="{total_cols}">── Alts ──</td></tr>'
        rows += "".join(_row(p, d) for p, d in sorted(alts.items(), key=sort_key))

    gear_header = '<th class="player-header">Player</th>'
    for i in range(max_items):
        div = " divider" if i > 0 else ""
        gear_header += f'<th class="{div}">Item {i+1}</th><th>Slot</th><th>Ilvl</th><th>Rank</th><th>Spark?</th>'

    return gear_header, rows, max_items, total_players, total_crafted, no_craft

CLASS_COLORS = {
    "DeathKnight": "#C41E3A", "DemonHunter": "#A330C9", "Druid": "#FF7C0A",
    "Evoker": "#33937F", "Hunter": "#AAD372", "Mage": "#3FC7EB",
    "Monk": "#00FF98", "Paladin": "#F48CBA", "Priest": "#FFFFFF",
    "Rogue": "#FFF468", "Shaman": "#0070DD", "Warlock": "#8788EE",
    "Warrior": "#C69B6D",
}

# WoW spec IDs → role  (used to detect per-fight spec swaps for accurate table grouping)
SPEC_ROLES: dict = {
    # Death Knight
    250: "Tank",   251: "DPS",    252: "DPS",
    # Demon Hunter
    577: "DPS",    581: "Tank",
    # Druid
    102: "DPS",    103: "DPS",    104: "Tank",   105: "Healer",
    # Evoker
    1467: "DPS",   1468: "Healer", 1473: "DPS",
    # Hunter
    253: "DPS",    254: "DPS",    255: "DPS",
    # Mage
    62: "DPS",     63: "DPS",     64: "DPS",
    # Monk
    268: "Tank",   269: "DPS",    270: "Healer",
    # Paladin
    65: "Healer",  66: "Tank",    70: "DPS",
    # Priest
    256: "Healer", 257: "Healer", 258: "DPS",
    # Rogue
    259: "DPS",    260: "DPS",    261: "DPS",
    # Shaman
    262: "DPS",    263: "DPS",    264: "Healer",
    # Warlock
    265: "DPS",    266: "DPS",    267: "DPS",
    # Warrior
    71: "DPS",     72: "DPS",     73: "Tank",
}


def _build_crown_mechanics_html(w: dict, actor_lookup: dict) -> str:
    """Render Silverstrike + P3 circle mechanics for a Crown of the Cosmos wipe pane."""
    from html import escape as _esc

    cm = w.get("crown_mechanics", {})
    if not cm:
        return ""

    def norm(raw_name):
        player, _ = lookup_roster(raw_name)
        return player

    def pcolor(raw_name):
        player = norm(raw_name)
        for a in actor_lookup.values():
            if lookup_roster(a.get("name", ""))[0] == player:
                return CLASS_COLORS.get(a.get("subType", ""), "#ccc")
        return "#ccc"

    # ── Helper: build one arrow round block ──────────────────────────────────
    def _arrow_block(im, show_shields):
        t_s   = im["arrows"][0]["t_s"] if im["arrows"] else 0
        t_lbl = f'{t_s // 60}:{t_s % 60:02d}'

        # Use debuff-assigned players as the primary list; fall back to non-extra hits
        assigned_raw = im.get("assigned", [])
        if assigned_raw:
            main_list   = assigned_raw
            also_hit    = [a for a in im["arrows"] if a["pid"] not in {x["pid"] for x in assigned_raw}]
        else:
            main_list   = [a for a in im["arrows"] if not a.get("is_extra", False)]
            also_hit    = [a for a in im["arrows"] if     a.get("is_extra", False)]

        def _player_row(a, extra=False):
            player_name = norm(a["name"])
            is_multi    = im["multi_hit"].get(a["pid"], 0) > 1
            row_cls     = ' class="cm-arrow-extra"' if extra else (' class="cm-arrow-multi"' if is_multi else "")
            seq_span    = f' <span class="cm-seq">×{a["seq"]}</span>' if a.get("seq", 1) > 1 else ""
            return (f'<tr{row_cls}>'
                    f'<td class="cm-name" style="color:{pcolor(a["name"])}">{_esc(player_name)}{seq_span}</td>'
                    f'<td class="cm-role">{_esc(a["role"])}</td>'
                    f'</tr>')

        # Main rows: debuff-assigned targets
        rows = "".join(_player_row(a) for a in main_list)

        shield_row = ""
        stacks_row = ""
        if show_shields:
            shield_cells = ""
            stacks_cells = ""
            for add_name in ["Demiar", "Morium", "Vorelus"]:
                exp = im.get("expected_shields", [])
                rem = im.get("shields_removed", [])
                if add_name not in exp:
                    shield_cells += f'<td class="cm-add-na">—</td>'
                elif add_name in rem:
                    shield_cells += f'<td class="cm-add-ok">✓</td>'
                else:
                    shield_cells += f'<td class="cm-add-miss">✗</td>'

                bef = im.get("ce_before", {}).get(add_name, 0)
                aft = im.get("ce_after",  {}).get(add_name, 0)
                if bef == 0 and aft == 0:
                    stacks_cells += f'<td class="cm-add-na">—</td>'
                elif aft > bef:
                    stacks_cells += f'<td class="cm-stacks-bad">{bef}→{aft}</td>'
                elif aft < bef:
                    stacks_cells += f'<td class="cm-stacks-ok">{bef}→{aft}</td>'
                else:
                    stacks_cells += f'<td class="cm-stacks-same">{bef}</td>'

            shield_row = f'<tr class="cm-shield-row"><td colspan="2" class="cm-shield-label">Shield removed</td>{shield_cells}</tr>'
            stacks_row = f'<tr class="cm-shield-row"><td colspan="2" class="cm-shield-label">CE stacks</td>{stacks_cells}</tr>'

        thead = '<tr><th>Player</th><th>Role</th><th>Demiar</th><th>Morium</th><th>Vorelus</th></tr>' if show_shields else '<tr><th>Player</th><th>Role</th></tr>'

        # Extra hits section (bounced / unintended)
        extra_html = ""
        if also_hit:
            extra_rows = "".join(_player_row(a, extra=True) for a in also_hit)
            extra_html = (f'<div class="cm-extra-section">'
                          f'<div class="cm-extra-hdr">Also hit</div>'
                          f'<table class="cm-table"><tbody>{extra_rows}</tbody></table>'
                          f'</div>')

        wasted = im.get("wasted_ce", [])
        wasted_html = f'<div class="cm-multiwarn">⚠ Wasted arrow on CE: {", ".join(wasted)}</div>' if wasted else ""
        return (f'<div class="cm-block">'
                f'<div class="cm-label cm-t">@ {t_lbl}</div>'
                f'<table class="cm-table"><thead>{thead}</thead>'
                f'<tbody>{rows}{shield_row}{stacks_row}</tbody></table>'
                f'{extra_html}{wasted_html}</div>')

    # ── Split silverstrike rounds into 4 phase buckets (chronological order) ──
    _SPELL_STAGE = 1233649  # Stage One / Stage Two: Silverstrike
    stage_buckets:  list = [[], []]   # index 0 = Stage One, 1 = Stage Two
    interm_buckets: list = [[], []]   # index 0 = Intermission 1, 1 = Intermission 2
    prev_spell = None
    stage_idx  = -1
    interm_idx = -1
    for im in cm.get("silverstrike", []):
        s = im.get("spell_id")
        if s != prev_spell:
            prev_spell = s
            if s == _SPELL_STAGE:
                stage_idx = min(stage_idx + 1, 1)
            else:
                interm_idx = min(interm_idx + 1, 1)
        if s == _SPELL_STAGE and 0 <= stage_idx <= 1:
            stage_buckets[stage_idx].append(im)
        elif s != _SPELL_STAGE and 0 <= interm_idx <= 1:
            interm_buckets[interm_idx].append(im)

    p3_circles = cm.get("p3_circles", [])
    sections   = ""

    def _phase_section(label, blocks_html, is_interm=False):
        """Wrap a phase label + content in a collapsible <details> panel."""
        cls = "cm-phase cm-interm" if is_interm else "cm-phase"
        return (f'<details class="{cls}">'
                f'<summary class="cm-phase-divider">{_esc(label)}</summary>'
                f'<div class="cm-blocks">{blocks_html}</div>'
                f'</details>')

    # ── Helper: intermission summary table ───────────────────────────────────
    def _interm_table(rounds: list, window_data: dict = None) -> str:  # noqa: ARG001
        """Table: rows=ALL raid players, one hits column showing SE+arrow# per hit.
        Format per hit: {se_stacks}+{cumulative_arrow_number}, e.g. '6+1 || 2+2 || 8+3'
        """
        if not rounds:
            return ""

        # spec_roles keys may be stored as strings in JSON cache — normalise to int
        spec_r = {int(k): v for k, v in w.get("spec_roles", {}).items()}

        # All raid pids if we have spec_roles, else just those hit
        if spec_r:
            all_pids = list(spec_r.keys())
        else:
            seen_set: set = set()
            all_pids = []
            for im in rounds:
                for a in im["arrows"]:
                    if a["pid"] not in seen_set:
                        seen_set.add(a["pid"])
                        all_pids.append(a["pid"])

        # Collect ordered hits per pid (chronological across all rounds)
        pid_hits: dict = {}   # pid -> [{"se": int, "total": int, "t_s": int}, ...]
        for im in rounds:
            for a in im["arrows"]:
                pid = a["pid"]
                pid_hits.setdefault(pid, []).append({
                    "se":    a.get("se_stacks", 0),
                    "total": a.get("arrow_total", 0),
                    "t_s":   a.get("t_s", 0),
                })

        # Sort: hit players first (most hits → top), then alphabetical
        sorted_pids = sorted(
            all_pids,
            key=lambda pid: (0 if pid in pid_hits else 1, -len(pid_hits.get(pid, [])))
        )

        thead = ('<tr>'
                 '<th class="cm-name">Player</th>'
                 '<th class="cm-role">Role</th>'
                 '<th class="cm-im-hits">Hits  <span style="font-weight:400;color:#555">(SE stacks + arrow #)</span></th>'
                 '</tr>')

        rows = ""
        for pid in sorted_pids:
            raw_name = actor_lookup.get(pid, {}).get("name", f"#{pid}")
            name     = norm(raw_name)
            role     = spec_r.get(pid, "DPS")
            hits = pid_hits.get(pid, [])

            if not hits:
                hits_cell = '<td class="cm-im-empty">—</td>'
            else:
                parts = []
                for h in hits:
                    se  = h["se"]
                    tot = h["total"]
                    se_cls = "cm-im-bad" if se == 0 else "cm-im-ok"
                    parts.append(
                        f'<span class="{se_cls}">{se}</span>'
                        f'<span class="cm-im-sep">+</span>'
                        f'<span class="cm-im-seq">{tot}</span>'
                    )
                sep = '<span class="cm-im-div"> || </span>'
                hits_cell = f'<td class="cm-im-hits-cell">{sep.join(parts)}</td>'

            rows += (f'<tr>'
                     f'<td class="cm-name" style="color:{pcolor(raw_name)}">{_esc(name)}</td>'
                     f'<td class="cm-role">{_esc(role)}</td>'
                     f'{hits_cell}</tr>')

        hits_table = f'<table class="cm-table cm-im-table"><thead>{thead}</thead><tbody>{rows}</tbody></table>'

        # ── Right: simple arrow count table (all players, sorted by count desc) ──
        count_sorted = sorted(all_pids, key=lambda pid: -len(pid_hits.get(pid, [])))
        count_rows = ""
        for pid in count_sorted:
            raw_name = actor_lookup.get(pid, {}).get("name", f"#{pid}")
            name     = norm(raw_name)
            n        = len(pid_hits.get(pid, []))
            if n == 0:
                count_cell = '<td class="cm-im-empty" style="text-align:center">—</td>'
            elif n > 1:
                count_cell = f'<td class="cm-im-stacked" style="text-align:center">×{n}</td>'
            else:
                count_cell = '<td style="text-align:center;color:#3fb950;font-weight:700">1</td>'
            count_rows += (f'<tr>'
                           f'<td class="cm-name" style="color:{pcolor(raw_name)}">{_esc(name)}</td>'
                           f'{count_cell}</tr>')
        count_thead = '<tr><th class="cm-name">Player</th><th class="cm-im-rnd">Arrows Hit</th></tr>'
        count_table = f'<table class="cm-table cm-im-table"><thead>{count_thead}</thead><tbody>{count_rows}</tbody></table>'

        return f'<div style="display:flex;gap:20px;align-items:flex-start">{hits_table}{count_table}</div>'

    # ── Fixed 5-phase layout — all dividers always rendered ───────────────────
    def _phase_timer(bucket: list) -> str:
        """Return ' — MM:SS – MM:SS' from first to last arrow in a bucket, or ''."""
        all_arrows = [a for im in bucket for a in im.get("arrows", [])]
        if not all_arrows:
            return ""
        t0 = min(a["t_s"] for a in all_arrows)
        t1 = max(a["t_s"] for a in all_arrows)
        return f"  \u2014  {t0//60}:{t0%60:02d} \u2013 {t1//60}:{t1%60:02d}"

    blocks = "".join(_arrow_block(im, show_shields=True)  for im in stage_buckets[0])
    sections += _phase_section(f"Stage One: The Void\u2019s Spire{_phase_timer(stage_buckets[0])}", blocks)

    def _iw_timer(iw: dict) -> str:
        """Format intermission timer from stored interm_start_t / interm_end_t."""
        s = iw.get("interm_start_t")
        e = iw.get("interm_end_t")
        if s is None or e is None:
            return ""
        s //= 1000; e //= 1000
        return f"  \u2014  {s//60}:{s%60:02d} \u2013 {e//60}:{e%60:02d}"

    _iw = cm.get("interm_windows", [])
    sections += _phase_section(
        f"Intermission: Crushing Singularity (WORK IN PROGRESS){_iw_timer(_iw[0]) if len(_iw) > 0 else ''}",
        _interm_table(interm_buckets[0], _iw[0] if len(_iw) > 0 else None),
        is_interm=True)

    blocks = "".join(_arrow_block(im, show_shields=False) for im in stage_buckets[1])
    sections += _phase_section(f"Stage Two: The Severed Rift (WORK IN PROGRESS){_phase_timer(stage_buckets[1])}", blocks)

    sections += _phase_section(
        f"Intermission: Shattering Singularity (WORK IN PROGRESS){_iw_timer(_iw[1]) if len(_iw) > 1 else ''}",
        _interm_table(interm_buckets[1], _iw[1] if len(_iw) > 1 else None),
        is_interm=True)

    # ── Stage Three: The End of the End — Circles ─────────────────────────────
    p3_blocks = ""
    for ci, circle_set in enumerate(p3_circles):
        rows = ""
        for i, p in enumerate(circle_set["players"]):
            bad = (i == 0 and p["role"] in ("Tank", "Melee")) or \
                  (i == len(circle_set["players"]) - 1 and p["role"] not in ("Tank",) and len(circle_set["players"]) == 3)
            row_cls = ' class="cm-circle-bad"' if bad else ""
            hold_s  = f'{p.get("hold_ms", 0)//1000}s' if p.get("hold_ms") else "?"
            rows += (f'<tr{row_cls}>'
                     f'<td class="cm-t">#{i+1}</td>'
                     f'<td class="cm-name" style="color:{pcolor(p["name"])}">{_esc(norm(p["name"]))}</td>'
                     f'<td class="cm-role">{_esc(p["role"])}</td>'
                     f'<td class="cm-t">{hold_s}</td>'
                     f'</tr>')
        flag_html = "".join(f'<div class="cm-multiwarn">⚠ {_esc(fl)}</div>' for fl in circle_set["flags"])
        p3_blocks += (f'<div class="cm-block">'
                      f'<div class="cm-label cm-t">Set {ci+1}</div>'
                      f'<table class="cm-table"><thead>'
                      f'<tr><th>#</th><th>Player</th><th>Role</th><th>Held</th></tr>'
                      f'</thead><tbody>{rows}</tbody></table>'
                      f'{flag_html}</div>')
    sections += _phase_section("Stage Three: The End of the End (WORK IN PROGRESS)", p3_blocks, is_interm=False)

    return f'<div class="cm-section"><div class="cm-section-title">Crown Mechanics</div>{sections}</div>'


_CROWN_MECHANICS_CSS = """
.cm-section { margin-top: 14px; }
.cm-section-title { font-size: 11px; font-weight: 700; color: #8b949e; text-transform: uppercase;
  letter-spacing: .06em; margin-bottom: 8px; }
.cm-phase { margin: 6px 0; }
.cm-phase-divider { list-style: none; display: block; cursor: pointer; user-select: none;
  background: #16213e; color: #7289DA; padding: 8px 10px;
  font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px;
  white-space: nowrap; position: relative; padding-left: 24px; }
.cm-phase-divider::before { content: "▼"; position: absolute; left: 8px;
  font-size: 8px; top: 50%; transform: translateY(-50%); opacity: 0.6;
  transition: transform .15s; }
.cm-phase:not([open]) > .cm-phase-divider::before { content: "▶"; transform: translateY(-50%); }
.cm-phase.cm-interm > .cm-phase-divider { color: #e57373; }
.cm-blocks { display: flex; flex-wrap: wrap; gap: 14px; padding: 6px 0 2px; }
.cm-block { flex: 0 0 auto; }
.cm-label { font-size: 11px; font-weight: 600; color: #58a6ff; margin-bottom: 5px; }
.cm-table { border-collapse: collapse; font-size: 12px; }
.cm-table th { background: #0d1117; color: #8b949e; font-size: 10px; font-weight: 600;
  text-transform: uppercase; padding: 3px 8px; border: 1px solid #30363d; text-align: left; }
.cm-table td { padding: 2px 8px; border: 1px solid #21262d; }
.cm-t { color: #8b949e; min-width: 38px; }
.cm-arrow { color: #58a6ff; font-weight: 700; min-width: 46px; }
.cm-name { font-weight: 600; min-width: 100px; }
.cm-role { color: #8b949e; min-width: 52px; }
.cm-seq { color: #d29922; font-size: 10px; }
.cm-arrow-rico td { opacity: 0.5; }
.cm-arrow-multi td { background: #3d1a1a !important; }
.cm-arrow-extra td { background: #2a1f00 !important; }
.cm-extra-section { margin-top: 5px; }
.cm-extra-hdr { font-size: 10px; font-weight: 700; color: #d29922; text-transform: uppercase;
  letter-spacing: .05em; padding: 2px 0; border-top: 1px solid #444; margin-bottom: 2px; }
.cm-circle-bad td { background: #3d1a1a; color: #f85149; }
.cm-multiwarn { margin-top: 5px; font-size: 11px; color: #f85149; }
.cm-shield-row td { border-top: 1px solid #444; font-size: 11px; text-align: center; }
.cm-shield-label { color: #8b949e; text-align: left !important; font-weight: 600; }
.cm-add-ok   { color: #3fb950; font-weight: 700; }
.cm-add-miss { color: #f85149; font-weight: 700; }
.cm-add-na   { color: #444; }
.cm-stacks-bad  { color: #f85149; font-weight: 700; font-size: 11px; }
.cm-stacks-ok   { color: #3fb950; font-weight: 700; font-size: 11px; }
.cm-stacks-same { color: #8b949e; font-size: 11px; }
/* ── Intermission summary table ── */
.cm-im-table { border-collapse: collapse; font-size: 12px; }
.cm-im-table th { background: #0d1117; color: #8b949e; font-size: 10px; font-weight: 600;
  text-transform: uppercase; padding: 3px 8px; border: 1px solid #30363d; text-align: left; }
.cm-im-table td { padding: 3px 8px; border: 1px solid #21262d; }
.cm-im-rnd       { min-width: 60px; }
.cm-im-hits      { min-width: 200px; }
.cm-im-hits-cell { white-space: nowrap; }
.cm-im-empty { text-align: center; color: #444; }
.cm-im-ok    { color: #3fb950; font-weight: 700; }
.cm-im-bad   { color: #f85149; font-weight: 700; }
.cm-im-seq   { color: #8b949e; }
.cm-im-sep   { color: #58a6ff; margin: 0 1px; }
.cm-im-div   { color: #444; margin: 0 4px; }
.cm-im-stacked     { text-align: center; color: #f85149; font-weight: 700; }
"""


def _build_boss_html(boss_data: dict, actor_lookup: dict, id_prefix: str = "0", wipe_data: dict = None) -> dict:
    """Build HTML for each boss tab. Returns {boss_name: html_string}."""
    ROLE_ORDER = {"Tank": 0, "Healer": 1, "DPS": 2}
    wipe_data  = wipe_data or {}
    results      = {}

    for boss_idx, (boss_name, fights) in enumerate(boss_data.items()):
        table_id       = f"boss-tbl-{id_prefix}-{boss_idx}"
        boss_name_base = boss_name.rsplit(" (", 1)[0]
        mech_defs      = BOSS_MECHANICS.get(boss_name_base, [])
        has_interrupts  = boss_name_base in BOSS_HAS_INTERRUPTS
        has_void_marked = "Averzian" in boss_name_base
        total_cols      = 8 + len(mech_defs) + (1 if has_interrupts else 0) + (1 if has_void_marked else 0)

        all_pids_set = set()
        for fight in fights:
            all_pids_set.update(fight.get("deaths", {}).keys())
            all_pids_set.update(fight.get("all_player_ids", set()))

        def _get_role(pid, fight=None):
            """Return the role for a player in a specific fight, falling back to roster."""
            char_name = actor_lookup.get(pid, {}).get("name", "")
            pname, _  = lookup_roster(char_name)
            if fight is not None:
                spec_role = fight.get("spec_roles", {}).get(pid)
                if spec_role:
                    return spec_role
            return PLAYER_ROLES.get(pname, "DPS")

        def pid_sort(pid):
            char_name = actor_lookup.get(pid, {}).get("name", "")
            pname, _  = lookup_roster(char_name)
            # Use spec from first available fight for stable sort order
            all_fight_list = list(fights) + list(wipe_data.get(boss_name, []))
            spec_role = next(
                (f.get("spec_roles", {}).get(pid) for f in all_fight_list if f.get("spec_roles", {}).get(pid)),
                None
            )
            role = spec_role or PLAYER_ROLES.get(pname, "DPS")
            return (ROLE_ORDER.get(role, 2), pname.lower())

        sorted_pids = sorted(all_pids_set, key=pid_sort)

        def _hfmt(n):
            return (f"{n/1_000_000_000:.1f}B" if n >= 1_000_000_000
                    else (f"{n/1_000_000:.1f}M" if n >= 1_000_000 else f"{n//1000}k"))

        def _overview_row(label, fight):
            sp_dmg   = sum(v.get("total", 0) if isinstance(v, dict) else 0 for v in fight.get("uptime_map", {}).values())
            sp_heal  = sum(v for v in fight.get("healing_map", {}).values() if isinstance(v, int))
            sp_taken = sum(v for v in fight.get("dmg_taken", {}).values() if isinstance(v, int))
            sp_dur_s = fight.get("fight_dur_ms", 0) / 1000 or 1
            sp_dps   = sp_dmg / sp_dur_s
            sp_m, sp_s = divmod(int(sp_dur_s), 60)
            inner = (f'<span class="sov-title">{label}</span>'
                     f'<span class="sov-dur">{sp_m}:{sp_s:02d}</span>'
                     f'<span class="sov-item"><span class="sov-lbl">⚡ Raid DPS</span> <b>{_hfmt(sp_dps)}</b></span>'
                     f'<span class="sov-item"><span class="sov-lbl">⚔ Dmg Done</span> <b>{_hfmt(sp_dmg) if sp_dmg else "—"}</b></span>'
                     f'<span class="sov-item"><span class="sov-lbl">💚 Healing</span> <b>{_hfmt(sp_heal) if sp_heal else "—"}</b></span>'
                     f'<span class="sov-item"><span class="sov-lbl">🛡 Dmg Taken</span> <b>{_hfmt(sp_taken) if sp_taken else "—"}</b></span>')
            # Per-mechanic totals (hits + damage)
            mts_all = fight.get("mechanic_timestamps", {})
            for m in mech_defs:
                lbl = m["label"]
                total_hits = 0
                total_dmg  = 0
                for pid_hits in mts_all.values():
                    for h in pid_hits.get(lbl, []):
                        if isinstance(h, dict):
                            total_hits += 1
                            total_dmg  += h.get("dmg", 0)
                if total_hits:
                    color = "#81c784" if m["type"] == "soak" else "#e57373"
                    inner += (f'<span class="sov-item" style="border-left:1px solid #2a2a4a;padding-left:12px">'
                              f'<span class="sov-lbl">{escape(lbl)}</span>'
                              f'<b style="color:{color}">{total_hits}×</b>'
                              f' <span style="color:#aaa">{_hfmt(total_dmg)}</span></span>')
            return f'<tr class="split-ov-row"><td colspan="{total_cols}"><div class="split-ov-bar">{inner}</div></td></tr>'

        # ── Per-pull banner builder (frontal + avoid for any list of fights) ──

        def _build_banners(fights_list: list) -> str:
            """Build Frontal Failures + Avoidable Hits HTML for a given list of fights."""
            out = ""
            # Frontal failures
            all_frontal = [f for fight in fights_list for f in fight.get("frontal_failures", [])]
            if all_frontal:
                rows = ""
                for f in all_frontal:
                    others_str = ", ".join(escape(p) for p in f["others"])
                    rows += (f'<div class="ff-item">'
                             f'<span class="ff-time">{escape(f["time_str"])}</span> '
                             f'<span class="ff-label">{escape(f["label"])}</span> → '
                             f'<span class="ff-primary">{escape(f["primary_player"])}</span>'
                             + (f' <span class="ff-sep">||</span> <span class="ff-players">{others_str}</span>' if others_str else '')
                             + '</div>')
                out += (f'<div class="frontal-failures">'
                        f'<div class="ff-header" onclick="this.classList.toggle(\'open\');this.nextElementSibling.classList.toggle(\'open\')">'
                        f'<span class="ff-title">⚠ Frontal Failures</span><span class="ff-chevron">▼</span></div>'
                        f'<div class="ff-body open">{rows}</div></div>')
            # Avoidable hits
            avoid_defs_local = [m for m in mech_defs if m.get("type") == "avoid"]
            if avoid_defs_local:
                avoid_labels = {m["label"] for m in avoid_defs_local}
                label_hits: dict = {}
                for fight in fights_list:
                    for pid, label_counts in fight.get("mechanics_data", {}).items():
                        for lbl, cnt in label_counts.items():
                            if lbl in avoid_labels:
                                label_hits.setdefault(lbl, {})[pid] = label_hits.setdefault(lbl, {}).get(pid, 0) + cnt
                if label_hits:
                    rows = ""
                    for m in avoid_defs_local:
                        lbl = m["label"]
                        if lbl not in label_hits:
                            continue
                        parts = " · ".join(
                            f'<span class="av-player">{escape(actor_lookup.get(pid, {}).get("name", f"?{pid}"))}</span>'
                            f'<span class="av-count"> ×{cnt}</span>'
                            for pid, cnt in sorted(label_hits[lbl].items(), key=lambda x: -x[1])
                        )
                        rows += f'<div class="av-item"><span class="av-label">{escape(lbl)}</span> → {parts}</div>'
                    if rows:
                        out += (f'<div class="avoid-failures">'
                                f'<div class="av-header" onclick="this.classList.toggle(\'open\');this.nextElementSibling.classList.toggle(\'open\')">'
                                f'<span class="av-title">⚡ Avoidable Hits</span><span class="av-chevron">▼</span></div>'
                                f'<div class="av-body open">{rows}</div></div>')
            return out

        def _build_chart(*_) -> str:
            return ""  # parked — see ideas/timeline_chart.html

        # ── Shared table renderer (kills + wipes) ──
        def _render_table(fights_list, tbl_id, pids=None):
            if pids is None:
                pids = sorted_pids

            def _sth(col_idx, lbl, css=""):
                return (f'<th class="{css}" data-sortable="1" style="cursor:pointer;user-select:none"'
                        f' onclick="sortBossTable(\'{tbl_id}\',{col_idx},this)">'
                        f'{lbl} <span class="sort-arrow">▼</span></th>')

            t  = f'<div class="table-wrap"><table id="{tbl_id}" class="detail-col-hidden"><thead><tr>'
            t += _sth(0, 'Player', 'player-header')
            t += '<th class="death-h">Killed by (time)</th>'
            t += _sth(2, 'Ext', 'ext-h')
            t += _sth(3, 'Def', 'def-h')
            t += _sth(4, 'Dmg Taken', 'dmg-h')
            t += _sth(5, 'Uptime %', 'uptime-h')
            if has_interrupts:
                t += _sth(6, 'Interrupts', 'interrupt-h')
            for mi, m in enumerate(mech_defs):
                mci = (7 if has_interrupts else 6) + mi
                css = "mech-soak-h" if m["type"] == "soak" else "mech-bad-h"
                tip = escape(m.get("name", m["label"])).replace("'", "&#39;")
                t += (f'<th class="{css}" data-sortable="1" style="cursor:pointer;user-select:none"'
                      f' data-htip=\'{tip}\' onmouseenter="showHTip(this)" onmouseleave="hideHTip()"'
                      f' onclick="sortBossTable(\'{tbl_id}\',{mci},this)">'
                      f'{escape(m["label"])} <span class="sort-arrow">▼</span></th>')
            if has_void_marked:
                t += '<th class="detail-h" style="min-width:140px"><span>Void Marked</span></th>'
            else:
                t += '<th>Notes</th>'
            t += '</tr></thead><tbody>'

            for fi, fight in enumerate(fights_list, 1):
                row_lbl = fight.get("_row_label") or f"Split {fi}"
                t += _overview_row(row_lbl, fight)
                current_role = None
                for pid in pids:
                    deaths_map = fight.get("deaths", {})
                    if pid not in fight.get("all_player_ids", set()) and pid not in deaths_map:
                        continue
                    actor     = actor_lookup.get(pid, {})
                    char_name = actor.get("name", f"ID-{pid}")
                    cls       = actor.get("subType", "Unknown")
                    cls_color = CLASS_COLORS.get(cls, "#ccc")
                    pname, _  = lookup_roster(char_name)
                    role      = _get_role(pid, fight)

                    if role != current_role:
                        current_role = role
                        t += f'<tr class="role-sep" data-role="{current_role}"><td colspan="{total_cols}">── {current_role}s ──</td></tr>'

                    dmg_taken_map = fight.get("dmg_taken", {})
                    uptime_map    = fight.get("uptime_map", {})
                    interrupts    = fight.get("interrupts", {})
                    fight_dur     = fight.get("fight_dur_ms", 0)
                    fight_mechs   = fight.get("mechanics_data", {})

                    death_list = deaths_map.get(pid, [])
                    d_raw      = dmg_taken_map.get(pid, 0)
                    dmg_str    = (f"{d_raw/1_000_000:.1f}M" if d_raw >= 1_000_000
                                  else (f"{d_raw/1000:.0f}k" if d_raw > 0 else "—"))
                    active     = (uptime_map.get(pid, {}).get("activeTime", 0)
                                  if isinstance(uptime_map.get(pid), dict) else 0)
                    uptime_str = (f"{min(active / fight_dur * 100, 100):.0f}%"
                                  if fight_dur > 0 and active > 0 else "—")
                    ext_list  = fight.get("external_casts", {}).get(pid, [])
                    ext_count = len(ext_list)
                    if ext_count > 0:
                        ext_tip = "<br>".join(
                            f'<span style="color:#c8e6c9">{escape(e["spell"])}</span>'
                            f' <span style="color:#7289DA">{escape(e["time"])}</span>'
                            for e in ext_list
                        )
                        ext_tip_attr = ext_tip.replace("'", "&#39;")
                        ext_cell = (f'<td class="center ext-h" style="cursor:help"'
                                    f' data-htip=\'{ext_tip_attr}\''
                                    f' onmouseenter="showHTip(this)" onmouseleave="hideHTip()">'
                                    f'<span class="ext-num">{ext_count}</span></td>')
                    else:
                        ext_cell = '<td class="center ext-h">—</td>'

                    def_list   = fight.get("defensive_casts", {}).get(pid, [])
                    def_count  = len(def_list)
                    if def_count > 0:
                        def_tip = "<br>".join(
                            f'<span style="color:#a0c4ff">{escape(d["spell"])}</span>'
                            f' <span style="color:#7289DA">{escape(d["time"])}</span>'
                            for d in def_list
                        )
                        def_tip_attr = def_tip.replace("'", "&#39;")
                        def_cell = (f'<td class="center def-h" style="cursor:help"'
                                    f' data-htip=\'{def_tip_attr}\''
                                    f' onmouseenter="showHTip(this)" onmouseleave="hideHTip()">'
                                    f'<span class="def-num">{def_count}</span></td>')
                    else:
                        def_cell = '<td class="center def-h">—</td>'

                    int_count  = interrupts.get(pid, 0)
                    int_str    = str(int_count) if int_count > 0 else "—"
                    int_style  = (' style="background:#1A3A1A"' if int_count > 0
                                  else (' style="background:#5D1A1A"' if has_interrupts else ""))

                    death_count  = len(death_list)
                    killed_str   = ("<br>".join(
                        f'{escape(d["ability"])} @ {escape(d["time"])} ({d.get("fight_pct", 0)}%)'
                        for d in death_list) or "—")

                    death_tip_html = ""
                    if death_list:
                        parts = []
                        for d in death_list:
                            tl    = d.get("timeline", [])
                            block = f"<b style='color:#e57373'>☠ {escape(d['ability'])}</b> @ {escape(d['time'])} ({d.get('fight_pct',0)}%)"
                            if tl:
                                trows = []
                                for tt in tl:
                                    ik    = abs(tt["sec_before"]) < 0.05
                                    col   = "#ff6b6b" if ik else "#aaa"
                                    lbl   = " ← killing blow" if ik else f"-{tt['sec_before']:.1f}s"
                                    trows.append(f"<span style='color:{col}'>{escape(tt['ability'])}: {escape(tt['amount_k'])} &nbsp;<span style='color:#666;font-size:11px'>{lbl}</span></span>")
                                block += "<br>" + "<br>".join(trows)
                            parts.append(block)
                        death_tip_html = ("<hr style='border-color:#333;margin:6px 0'>").join(parts)

                    row_cls = "boss-death-row" if death_count > 0 else ""
                    t += f'<tr class="{row_cls}" data-role="{role}" data-class="{cls}" data-player="{escape(pname.lower())}">'
                    slug = pname.lower()
                    t += f'<td class="player-cell"><a href="players/player_{slug}.html" class="pname" style="color:{cls_color}">{escape(pname)}</a></td>'
                    if death_count > 0 and death_tip_html:
                        tip_attr = death_tip_html.replace("'", "&#39;")
                        t += f'<td class="death-h" data-htip=\'{tip_attr}\' onmouseenter="showHTip(this)" onmouseleave="hideHTip()" style="cursor:help">{killed_str}</td>'
                    else:
                        t += f'<td class="death-h">{killed_str}</td>'
                    t += ext_cell
                    t += def_cell
                    t += f'<td class="center dmg-h">{dmg_str}</td>'
                    t += f'<td class="center uptime-h">{uptime_str}</td>'
                    if has_interrupts:
                        t += f'<td class="center interrupt-h"{int_style}>{int_str}</td>'
                    pid_mechs = fight_mechs.get(pid, {})
                    pid_mts   = fight.get("mechanic_timestamps", {}).get(pid, {})
                    for m in mech_defs:
                        cnt   = pid_mechs.get(m["label"], 0)
                        hits  = pid_mts.get(m["label"], [])
                        if cnt:
                            bg = "#1A3D1A" if m["type"] == "soak" else "#5D1A1A"
                            def _fmt_dmg(v):
                                if v >= 1_000_000: return f"{v/1_000_000:.1f}M"
                                if v >= 1_000:     return f"{v/1_000:.0f}k"
                                return str(v)
                            if hits:
                                total_dmg = sum(h["dmg"] for h in hits if isinstance(h, dict))
                                tip_html  = "<br>".join(
                                    f'<span style="color:#a0b4ff">{escape(h["t"])}</span>'
                                    f' <span style="color:#e57373">{_fmt_dmg(h["dmg"])}</span>'
                                    if isinstance(h, dict) else escape(h)
                                    for h in hits
                                )
                                tip_attr  = tip_html.replace("'", "&#39;")
                                cell_txt  = f'{cnt} <span style="color:#aaa;font-size:11px">({_fmt_dmg(total_dmg)})</span>'
                                t += (f'<td class="center" style="background:{bg};cursor:help"'
                                      f' data-htip=\'{tip_attr}\' onmouseenter="showHTip(this)"'
                                      f' onmouseleave="hideHTip()">{cell_txt}</td>')
                            else:
                                t += f'<td class="center" style="background:{bg}">{cnt}</td>'
                        else:
                            t += '<td class="center">—</td>'
                    if has_void_marked:
                        vm_list = [mv for mv in fight.get("void_marked", []) if mv["target_pid"] == pid]
                        if vm_list:
                            parts = []
                            for mv in vm_list:
                                tm2, ts2 = divmod(mv["t_s"], 60)
                                t_str2   = f"{tm2}:{ts2:02d}"
                                if mv["dispeller_pid"] is not None:
                                    d_name = escape(actor_lookup.get(mv["dispeller_pid"], {}).get("name", f"#{mv['dispeller_pid']}"))
                                    parts.append(f'<span style="color:#e57373">{t_str2}</span>'
                                                 f' <span class="vm-arrow">&#8594;</span>'
                                                 f' <span style="color:#81c784">{d_name}</span>')
                                else:
                                    parts.append(f'<span style="color:#e57373">{t_str2}</span>'
                                                 f' <span class="vm-arrow">&#8594;</span>'
                                                 f' <span class="vm-expired">expired</span>')
                            t += f'<td class="detail-cell">{"<br>".join(parts)}</td>'
                        else:
                            t += '<td class="center">—</td>'
                    else:
                        t += '<td></td>'
                    t += '</tr>'

            t += '</tbody></table></div>'
            return t

        def _build_alndust_panel(fights_list: list) -> str:
            """Render the Alndust Upheaval up/down grouping panel for Chimaerus."""
            waves = []
            for fd in fights_list:
                if fd.get("alndust_groups"):
                    waves = fd["alndust_groups"]
                    break
            if not waves:
                return ""

            def _pid_chip(pid, missed=False, wrong=False, double_up=False):
                info = actor_lookup.get(pid, {})
                name = info.get("name", f"#{pid}")
                if missed:
                    return f'<span class="ag-chip ag-miss" title="Missed soak">{escape(name)} ✗</span>'
                if wrong:
                    return f'<span class="ag-chip ag-wrong" title="Wrong group soaked">{escape(name)} ?</span>'
                if double_up:
                    return f'<span class="ag-chip ag-double-up" title="Up 2+ waves in a row">{escape(name)} ⚠</span>'
                return f'<span class="ag-chip">{escape(name)}</span>'

            rows = ""
            for w in waves:
                wave_num     = w["wave"]
                m, s         = divmod(w["t_s"], 60)
                t_str        = f"{m}:{s:02d}"
                missed       = set(w.get("missed_pids", []))
                wrong        = set(w.get("wrong_pids",  []))
                double_up    = set(w.get("double_up_pids", []))

                down_chips = " ".join(
                    _pid_chip(p, wrong=(p in wrong)) for p in w["down_pids"]
                )
                up_chips = " ".join(
                    _pid_chip(p, missed=(p in missed), double_up=(p in double_up))
                    for p in w["up_pids"]
                )
                warnings = []
                if missed:
                    warnings.append(f'<span class="ag-miss-count">⚠ {len(missed)} missed</span>')
                if double_up:
                    warnings.append(f'<span class="ag-double-up-count">⚠ {len(double_up)} up again</span>')
                label_extra = (" " + " ".join(warnings)) if warnings else ""
                rows += (
                    f'<div class="ag-wave">'
                    f'<div class="ag-wave-label">Wave {wave_num} <span class="ag-time">({t_str})</span>{label_extra}</div>'
                    f'<div class="ag-group ag-down"><span class="ag-badge ag-badge-down">⬇ Down</span> {down_chips}</div>'
                    f'<div class="ag-group ag-up"><span class="ag-badge ag-badge-up">⬆ Up</span> {up_chips}</div>'
                    f'</div>'
                )
            return (
                f'<div class="alndust-panel">'
                f'<div class="ag-header" onclick="this.classList.toggle(\'open\');this.nextElementSibling.classList.toggle(\'open\')">'
                f'<span class="ag-title">⬆⬇ Alndust Upheaval — Intermission Groups</span>'
                f'<span class="ag-chevron">▼</span></div>'
                f'<div class="ag-body open">{rows}</div></div>'
            )

        def _build_gloom_panel(fights_list: list) -> str:
            """Render per-wave Gloomfield soak list for Vaelgor & Ezzorak."""
            soaks = []
            for fd in fights_list:
                if fd.get("gloom_soaks"):
                    soaks = fd["gloom_soaks"]
                    break
            if not soaks:
                return ""
            rows = ""
            for w in soaks:
                m, s   = divmod(w["t_s"], 60)
                t_str  = f"{m}:{s:02d}"
                chips  = " ".join(
                    f'<span class="ag-chip">{escape(actor_lookup.get(p, {}).get("name", f"#{p}"))}</span>'
                    for p in w["hit_pids"]
                )
                rows += (
                    f'<div class="ag-wave">'
                    f'<div class="ag-wave-label">Soak {w["wave"]} <span class="ag-time">({t_str})</span></div>'
                    f'<div class="ag-group ag-down">{chips}</div>'
                    f'</div>'
                )
            return (
                f'<div class="alndust-panel">'
                f'<div class="ag-header" onclick="this.classList.toggle(\'open\');this.nextElementSibling.classList.toggle(\'open\')">'
                f'<span class="ag-title">&#127761; Gloomfield — Who Got Hit</span>'
                f'<span class="ag-chevron">&#9660;</span></div>'
                f'<div class="ag-body open">{rows}</div></div>'
            )

        def _group_void_marked_waves(marks: list) -> list:
            """Group Void Marked events into waves (gap > 5s = new wave). Returns list of waves."""
            if not marks:
                return []
            sorted_marks = sorted(marks, key=lambda x: x["t_s"])
            waves = [[sorted_marks[0]]]
            for m in sorted_marks[1:]:
                if m["t_s"] - waves[-1][-1]["t_s"] <= 5:
                    waves[-1].append(m)
                else:
                    waves.append([m])
            return waves

        def _build_void_marked_panel(fights_list: list) -> str:
            """Render Void Marked waves — grouped by simultaneous application (~5s window)."""
            marks = []
            for fd in fights_list:
                if fd.get("void_marked"):
                    marks = fd["void_marked"]
                    break

            def _colored_name(pid):
                a     = actor_lookup.get(pid, {})
                name  = escape(a.get("name", f"#{pid}"))
                color = CLASS_COLORS.get(a.get("subType", ""), "#ccc")
                return f'<span style="color:{color};font-weight:600">{name}</span>'

            def _ts(s):
                m, sec = divmod(s, 60)
                return f"{m}:{sec:02d}"

            if not marks:
                body = '<div style="color:#555;font-size:12px;padding:8px 4px;">No Void Marked events recorded.</div>'
            else:
                waves = _group_void_marked_waves(marks)

                body = ""
                for wi, wave in enumerate(waves, 1):
                    chips = ""
                    for m in sorted(wave, key=lambda x: (x["dispel_t_s"] is None, x["dispel_t_s"] or 0)):
                        apply_str = _ts(m["t_s"])
                        if m["dispel_t_s"] is not None:
                            remove_str = _ts(m["dispel_t_s"])
                            right = f'<span class="ag-time">{remove_str}</span> <span style="color:#81c784;font-size:11px">Removed</span>'
                        else:
                            right = '<span style="color:#555;font-size:11px">—</span>'
                        chips += (f'<span style="display:inline-flex;align-items:center;gap:5px;'
                                  f'background:rgba(255,255,255,0.04);border:1px solid #2a2a4a;'
                                  f'border-radius:6px;padding:2px 10px;margin:2px 4px 2px 0;font-size:12px">'
                                  f'{_colored_name(m["target_pid"])} <span class="ag-time">({apply_str})</span>'
                                  f' <span class="vm-arrow">&#8594;</span> {right}</span>')
                    wm2, ws2 = divmod(wave[0]["t_s"], 60)
                    body += (
                        f'<div class="ag-wave">'
                        f'<div class="ag-wave-label">Wave {wi} <span class="ag-time">({wm2}:{ws2:02d})</span></div>'
                        f'<div style="display:flex;flex-wrap:wrap;padding:2px 0">{chips}</div>'
                        f'</div>'
                    )

            return (
                f'<div class="alndust-panel">'
                f'<div class="ag-header open" onclick="this.classList.toggle(\'open\');this.nextElementSibling.classList.toggle(\'open\')">'
                f'<span class="ag-title">&#128165; Void Marked — Dispel Log</span>'
                f'<span class="ag-chevron">&#9660;</span></div>'
                f'<div class="ag-body open">{body}</div></div>'
            )

        def _build_add_damage_panel(fights_list: list) -> str:
            """Render per-player, per-add damage + active-time table (one column per add instance)."""
            instances: list = []
            fight_dur_ms    = 0
            vm_marks: list  = []
            for fd in fights_list:
                if fd.get("add_damage"):
                    instances    = fd["add_damage"]
                    fight_dur_ms = fd.get("fight_dur_ms", 0)
                    vm_marks     = fd.get("void_marked", [])
                    break

            if not instances or not isinstance(instances, list) or not isinstance(instances[0], dict) \
                    or "wave_num" not in instances[0]:
                body = '<div style="color:#555;font-size:12px;padding:8px 4px;">No add damage recorded.</div>'
            else:
                def fmt(v):
                    if v >= 1_000_000: return f"{v/1_000_000:.1f}M"
                    if v >= 1_000:     return f"{v/1_000:.0f}k"
                    return str(int(v)) if v else "—"

                def _add_label(wave):
                    ws = wave["wave_start_s"]
                    m, s = divmod(ws, 60)
                    return f'Add {wave["wave_num"]} <span class="ag-time">({m}:{s:02d})</span>'

                # Collect all player IDs as ints (JSON loads dict keys as strings)
                all_pids_set: set = set()
                for wave in instances:
                    all_pids_set.update(int(k) for k in wave["players"].keys())
                sorted_pids = sorted(all_pids_set,
                                     key=lambda p: sum(w["players"].get(str(p), w["players"].get(p, {})).get("dmg", 0) for w in instances),
                                     reverse=True)

                n_waves   = len(instances)
                total_col = 1 + n_waves * 2  # Player=0, then pairs, total at end

                th_style  = 'style="border-left:1px solid #2a2a4a;cursor:pointer;user-select:none"'
                thead = '<tr><th rowspan="2">Player</th>'
                for wave in instances:
                    thead += f'<th colspan="2" style="text-align:center;border-left:1px solid #2a2a4a">{_add_label(wave)}</th>'
                thead += (f'<th rowspan="2" data-col="{total_col}" onclick="sortAddTable(this)" '
                          f'style="border-left:1px solid #2a2a4a;cursor:pointer;user-select:none">'
                          f'Total<span class="sort-arrow"></span></th></tr>')
                thead += '<tr>'
                for i, _ in enumerate(instances):
                    dmg_col    = 1 + i * 2
                    uptime_col = 2 + i * 2
                    thead += (f'<th data-col="{dmg_col}" onclick="sortAddTable(this)" {th_style}>'
                              f'Dmg<span class="sort-arrow"></span></th>'
                              f'<th data-col="{uptime_col}" onclick="sortAddTable(this)" {th_style}>'
                              f'Uptime<span class="sort-arrow"></span></th>')
                thead += '</tr>'

                rows = ""
                for pid in sorted_pids:
                    actor     = actor_lookup.get(pid, {})
                    name      = escape(actor.get("name", f"#{pid}"))
                    color     = CLASS_COLORS.get(actor.get("subType", ""), "#ccc")
                    total_dmg = sum(w["players"].get(str(pid), w["players"].get(pid, {})).get("dmg", 0) for w in instances)
                    rows += f'<tr><td data-val="0"><span style="color:{color};font-weight:600">{name}</span></td>'
                    for wave in instances:
                        pdata    = wave["players"].get(str(pid), wave["players"].get(pid))
                        wave_dur = max((wave["wave_end_s"] - wave["wave_start_s"]) * 1000, 1)
                        if pdata and pdata["dmg"]:
                            active_ms = pdata.get("active_ms", 0)
                            uptime_p  = min(active_ms / wave_dur * 100, 100)
                            rows += (f'<td class="add-dmg-val" data-val="{pdata["dmg"]}" style="border-left:1px solid #2a2a4a">{fmt(pdata["dmg"])}</td>'
                                     f'<td class="add-dmg-val" data-val="{uptime_p:.1f}">{uptime_p:.0f}%</td>')
                        else:
                            rows += '<td class="add-dmg-val" data-val="0" style="border-left:1px solid #2a2a4a;color:#444">—</td><td class="add-dmg-val" data-val="0" style="color:#444">—</td>'
                    rows += f'<td class="add-dmg-val" data-val="{total_dmg}" style="border-left:1px solid #2a2a4a;font-weight:600">{fmt(total_dmg)}</td></tr>'

                body = f'<table class="add-dmg-table"><thead>{thead}</thead><tbody>{rows}</tbody></table>'

            return (
                f'<div class="alndust-panel">'
                f'<div class="ag-header" onclick="this.classList.toggle(\'open\');this.nextElementSibling.classList.toggle(\'open\')">'
                f'<span class="ag-title">&#128165; Abyssal Voidshaper — Add Damage</span>'
                f'<span class="ag-chevron">&#9660;</span></div>'
                f'<div class="ag-body open">{body}</div></div>'
            )


        def _build_horror_waves_table(fights_list: list) -> str:
            """Build per-wave Colossal Horror damage + uptime table for Chimaerus."""
            horror_waves = []
            for fd in fights_list:
                if fd.get("horror_waves"):
                    horror_waves = fd["horror_waves"]
                    break
            if not horror_waves:
                return ""

            def fmt_dmg(v):
                if v >= 1_000_000: return f"{v / 1_000_000:.2f}M"
                if v >= 1_000:     return f"{v / 1_000:.1f}k"
                if v > 0:          return str(v)
                return "—"

            def fmt_dps(dmg, wave_dur_ms):
                if not wave_dur_ms or not dmg:
                    return "—"
                dps = dmg / (wave_dur_ms / 1000)
                if dps >= 1_000_000: return f"{dps / 1_000_000:.2f}M"
                if dps >= 1_000:     return f"{dps / 1_000:.1f}k"
                return str(int(dps))

            def fmt_active(active_ms, wave_dur_ms):
                if not wave_dur_ms or not active_ms:
                    return "—"
                pct = min(100.0, active_ms / wave_dur_ms * 100)
                return f"{pct:.1f}%"

            # Determine group membership: Group 1 = players who went DOWN on wave 1
            group1 = set(horror_waves[0]["down_pids"]) if horror_waves else set()
            group2 = set(horror_waves[1]["down_pids"]) if len(horror_waves) > 1 else set()

            # Build ordered player list: group1 first, then group2, then others
            ordered = []
            seen = set()
            for pid in horror_waves[0].get("down_pids", []):
                if pid not in seen: ordered.append(pid); seen.add(pid)
            for pid in horror_waves[0].get("up_pids", []):
                if pid not in seen: ordered.append(pid); seen.add(pid)
            for wave in horror_waves:
                for pid in wave.get("down_pids", []) + wave.get("up_pids", []):
                    if pid not in seen: ordered.append(pid); seen.add(pid)

            # Separate group1 and group2 ordered lists
            g1_players = [p for p in ordered if p in group1]
            g2_players = [p for p in ordered if p in group2]
            other_players = [p for p in ordered if p not in group1 and p not in group2]

            num_waves = len(horror_waves)
            tbl_id = "hw-tbl-main"

            # Build header: Wave N with 4 sub-cols each (Phase, DMG, DPS, Active)
            t = '<div class="horror-waves-wrap">'
            t += '<div class="hw-title">⚔ Colossal Horror — Wave Breakdown</div>'
            t += f'<table class="boss-table hw-tbl" id="{tbl_id}"><thead>'
            t += '<tr><th class="hw-player-col" rowspan="2">Player</th>'
            for w in horror_waves:
                t += f'<th colspan="4" class="hw-wave-hdr">Wave {w["wave"]}</th>'
            t += '</tr><tr>'
            for wi, w in enumerate(horror_waves):
                t += f'<th class="hw-sub">Phase</th>'
                t += (f'<th class="hw-sub hw-sortable" style="cursor:pointer;user-select:none"'
                      f' onclick="hwSort({wi},\'dmg\',this)">DMG <span class="sort-arrow">▼</span></th>')
                t += (f'<th class="hw-sub hw-sortable" style="cursor:pointer;user-select:none"'
                      f' onclick="hwSort({wi},\'dps\',this)">DPS <span class="sort-arrow">▼</span></th>')
                t += (f'<th class="hw-sub hw-sortable" style="cursor:pointer;user-select:none"'
                      f' onclick="hwSort({wi},\'active\',this)">Active <span class="sort-arrow">▼</span></th>')
            t += '</tr></thead><tbody>'

            def player_row(pid, group_id=""):
                info  = actor_lookup.get(pid, {})
                name  = info.get("name", f"#{pid}")
                cls   = info.get("subType", "")
                color = CLASS_COLORS.get(cls, "#ccc")
                row = f'<tr data-group="{group_id}">'
                row += f'<td class="hw-name" style="color:{color}">{escape(name)}</td>'
                for w in horror_waves:
                    pp       = w["per_player"].get(pid, {})
                    phase    = pp.get("phase", "")
                    dmg      = pp.get("dmg", 0)
                    active   = pp.get("active_ms", 0)
                    dur      = w.get("wave_dur_ms", 0)
                    ph_html  = ('<span class="hw-phase-down">↓ Down</span>' if phase == "down"
                                else '<span class="hw-phase-up">↑ Up</span>' if phase == "up"
                                else "—")
                    dps_val  = int(dmg / (dur / 1000)) if dur else 0
                    row += f'<td class="hw-phase-cell">{ph_html}</td>'
                    row += f'<td class="hw-dmg-cell" data-val="{dmg}">{fmt_dmg(dmg)}</td>'
                    row += f'<td class="hw-dmg-cell" data-val="{dps_val}">{fmt_dps(dmg, dur)}</td>'
                    row += f'<td class="hw-dmg-cell" data-val="{active}">{fmt_active(active, dur)}</td>'
                row += '</tr>'
                return row

            num_cols = 1 + num_waves * 4
            def group_header(label, group_id):
                return (f'<tr class="hw-group-hdr" data-group-hdr="{group_id}">'
                        f'<td colspan="{num_cols}">{label}</td></tr>')

            if g1_players:
                t += group_header("Group 1", "g1")
                for pid in g1_players:
                    t += player_row(pid, group_id="g1")
            if g2_players:
                t += group_header("Group 2", "g2")
                for pid in g2_players:
                    t += player_row(pid, group_id="g2")
            for pid in other_players:
                t += player_row(pid, group_id="other")

            t += '</tbody></table></div>'

            # Sort JS
            t += """
<script>
(function(){
  var hwAsc = {};
  window.hwSort = function(waveIdx, col, th) {
    var key = waveIdx + '_' + col;
    hwAsc[key] = !hwAsc[key];
    var tbl = document.getElementById('hw-tbl-main');
    var tbody = tbl.querySelector('tbody');
    var colOffset = 1 + waveIdx * 4 + (col === 'dmg' ? 1 : col === 'dps' ? 2 : 3);
    var asc = hwAsc[key];

    // Collect groups in order: [{hdr: trEl or null, rows: [trEl, ...]}, ...]
    var groups = [];
    var currentGroup = null;
    Array.from(tbody.querySelectorAll('tr')).forEach(function(r) {
      if (r.hasAttribute('data-group-hdr')) {
        currentGroup = { hdr: r, rows: [] };
        groups.push(currentGroup);
      } else if (r.hasAttribute('data-group')) {
        if (!currentGroup) { currentGroup = { hdr: null, rows: [] }; groups.push(currentGroup); }
        currentGroup.rows.push(r);
      }
    });

    // Sort rows within each group
    groups.forEach(function(g) {
      g.rows.sort(function(a, b) {
        var av = parseInt(a.cells[colOffset] ? a.cells[colOffset].getAttribute('data-val') || 0 : 0);
        var bv = parseInt(b.cells[colOffset] ? b.cells[colOffset].getAttribute('data-val') || 0 : 0);
        return asc ? av - bv : bv - av;
      });
      // Re-append: header first, then sorted rows
      if (g.hdr) tbody.appendChild(g.hdr);
      g.rows.forEach(function(r) { tbody.appendChild(r); });
    });

    tbl.querySelectorAll('.sort-arrow').forEach(function(s){ s.textContent = '▼'; });
    if (th) th.querySelector('.sort-arrow').textContent = asc ? '▲' : '▼';
  };
})();
</script>
"""
            return t

        def _build_vorasius_tank_swap_banner(fights_list: list) -> str:
            """Orange warning for Shadowclaw Slam tank swap failures."""
            all_swaps = [ev for fd in fights_list for ev in fd.get("vorasius_tank_swaps", [])]
            if not all_swaps:
                return ""
            items = ""
            for ev in all_swaps:
                items += (f'<div class="ts-item">'
                          f'<span class="ts-player">{escape(ev["player"])}</span>'
                          f' <span class="ts-time">({escape(ev["time_str"])})</span>'
                          f" didn't swap — Slam #{ev['slam_num']}"
                          f'</div>')
            return (f'<div class="tank-swap-warn">'
                    f'<div class="ts-header">'
                    f'<span class="ts-title">&#9888;&#65039; Tank Swap Missed — Shadowclaw Slam</span>'
                    f'</div>'
                    f'<div class="ts-body">{items}</div></div>')

        def _build_salhadaar_wipe_banner(fights_list: list) -> str:
            """Red banner if Void Infusion (orb reached boss) was detected."""
            all_flags = [ev for fd in fights_list for ev in fd.get("salhadaar_wipe_events", [])]
            if not all_flags:
                return ""
            items = ""
            for ev in all_flags:
                items += (f'<div class="sw-item">'
                          f'<span class="sw-time">{escape(ev["time_str"])}</span>'
                          f' <span class="sw-event">&#128308; {escape(ev["event_type"])}</span>'
                          f' — orb reached the boss'
                          f'</div>')
            return (f'<div class="salhadaar-wipe-warn">'
                    f'<div class="sw-header">'
                    f'<span class="sw-title">&#128308; Wipe Condition Detected</span>'
                    f'</div>'
                    f'<div class="sw-body">{items}</div></div>')

        def _build_salhadaar_interrupt_table(fights_list: list) -> str:
            """Shadow Fracture interrupt wave table for Fallen-King Salhadaar.
            Rows = players who interrupted, columns = waves, cells = spell + timestamp."""
            waves = next((fd["salhadaar_fracture_waves"]
                          for fd in fights_list if fd.get("salhadaar_fracture_waves")), [])
            if not waves:
                return ""

            all_pids: set = set()
            for wave in waves:
                all_pids.update(int(k) if isinstance(k, str) else k for k in wave["players"])
            if not all_pids:
                return ""

            sorted_int_pids = sorted(all_pids, key=pid_sort)

            def _ts2(s):
                m2, s2 = divmod(int(s), 60)
                return f"{m2}:{s2:02d}"

            t  = '<div class="alndust-panel">'
            t += ('<div class="ag-header open" onclick="this.classList.toggle(\'open\');'
                  'this.nextElementSibling.classList.toggle(\'open\')">'
                  '<span class="ag-title">&#128683; Shadow Fracture — Interrupt Log</span>'
                  '<span class="ag-chevron">&#9660;</span></div>')
            t += '<div class="ag-body open"><div style="overflow-x:auto">'
            t += '<table class="boss-table" style="width:100%;border-collapse:collapse"><thead><tr>'
            t += '<th style="text-align:left;padding:5px 10px;color:#aaa;white-space:nowrap">Player</th>'
            for wave in waves:
                ws_str = _ts2(wave["wave_start_s"])
                we_str = _ts2(wave["wave_end_s"])
                t += (f'<th style="text-align:center;padding:5px 8px;color:#aaa;'
                      f'border-left:1px solid #2a2a4a;white-space:nowrap">'
                      f'Wave {wave["wave_num"]}'
                      f'<br><span style="color:#556;font-size:11px;font-weight:400">'
                      f'{ws_str}–{we_str}</span></th>')
            t += '</tr></thead><tbody>'

            for pid in sorted_int_pids:
                actor = actor_lookup.get(pid, {})
                name  = escape(actor.get("name", f"#{pid}"))
                color = CLASS_COLORS.get(actor.get("subType", ""), "#ccc")
                t += (f'<tr><td style="padding:4px 10px;white-space:nowrap">'
                      f'<span style="color:{color};font-weight:600">{name}</span></td>')
                for wave in waves:
                    pdata = wave["players"].get(pid, wave["players"].get(str(pid), []))
                    if pdata:
                        parts = " ".join(
                            f'<span style="white-space:nowrap">'
                            f'<span style="color:#a0c4ff">{escape(c["spell"])}</span>'
                            f' <span class="ag-time">({escape(c["time_str"])})</span>'
                            f'</span>'
                            for c in pdata
                        )
                        t += (f'<td style="text-align:center;border-left:1px solid #2a2a4a;'
                              f'padding:4px 8px">{parts}</td>')
                    else:
                        t += '<td style="text-align:center;border-left:1px solid #2a2a4a;color:#333">—</td>'
                t += '</tr>'

            t += '</tbody></table></div></div></div>'
            return t

        # ── Pull selector + panes ──
        boss_wipes = wipe_data.get(boss_name, [])  # oldest → newest

        # Build kill pane content: banners + table + chart (use first kill for chart)
        kill_banners = _build_banners(fights)
        kill_table   = _render_table(fights, table_id)
        kill_chart   = _build_chart(fights[0] if fights else {}, f"{table_id}-kill")
        kill_alndust         = _build_alndust_panel(fights)              if "Chimaerus"  in boss_name else ""
        kill_horror_waves    = _build_horror_waves_table(fights)         if "Chimaerus"  in boss_name else ""
        kill_gloom           = _build_gloom_panel(fights)                if "Vaelgor"    in boss_name else ""
        kill_void_marked     = _build_void_marked_panel(fights)          if "Averzian"   in boss_name else ""
        kill_add_damage      = _build_add_damage_panel(fights)           if "Averzian"   in boss_name else ""
        kill_vorasius_swap   = _build_vorasius_tank_swap_banner(fights)  if "Vorasius"   in boss_name else ""
        kill_sal_wipe        = _build_salhadaar_wipe_banner(fights)      if "Salhadaar"  in boss_name else ""
        kill_sal_interrupts  = _build_salhadaar_interrupt_table(fights)  if "Salhadaar"  in boss_name else ""
        kill_content = (kill_sal_wipe + kill_vorasius_swap + kill_banners + kill_table + kill_chart
                        + kill_horror_waves + kill_alndust + kill_gloom
                        + kill_void_marked + kill_add_damage + kill_sal_interrupts)

        if boss_wipes:
            total_wipes = len(boss_wipes)
            # Annotate wipes with display labels (oldest = Wipe 1)
            for wi, w in enumerate(boss_wipes):
                bpct  = w.get("boss_pct", 0)
                dur_s = w.get("fight_dur_ms", 0) // 1000
                wm, ws = divmod(dur_s, 60)
                w["_row_label"] = f"💀 Wipe {wi + 1} — {bpct}% boss HP · {wm}:{ws:02d}"

            kill_dur_s = fights[0].get("fight_dur_ms", 0) // 1000 if fights else 0
            km, ks     = divmod(kill_dur_s, 60)
            pull_btns  = f'<button class="pull-btn active" onclick="switchPull(this,\'{table_id}\')">⚔ Kill · {km}:{ks:02d}</button>'

            wipe_panes = ""
            for wi, w in enumerate(reversed(boss_wipes)):  # newest first in selector
                actual_wipe_num = total_wipes - wi          # 1-based, newest = total_wipes
                bpct    = w.get("boss_pct", 0)
                dur_s   = w.get("fight_dur_ms", 0) // 1000
                wm, ws  = divmod(dur_s, 60)
                wtbl_id = f"wipe-{table_id}-{wi}"
                pull_btns += f'<button class="pull-btn wipe" onclick="switchPull(this,\'{wtbl_id}\')">Wipe {actual_wipe_num} · {bpct}% · {wm}:{ws:02d}</button>'
                wipe_pids = sorted(
                    all_pids_set | {p for p in (set(w.get("all_player_ids", [])) | set(w.get("deaths", {}).keys())) if p in actor_lookup},
                    key=pid_sort
                )
                wipe_banners         = _build_banners([w])
                wipe_table           = _render_table([w], wtbl_id, pids=wipe_pids)
                wipe_chart           = _build_chart(w, wtbl_id)
                wipe_sal_wipe        = _build_salhadaar_wipe_banner([w])     if "Salhadaar" in boss_name else ""
                wipe_vorasius_swap   = _build_vorasius_tank_swap_banner([w]) if "Vorasius"  in boss_name else ""
                wipe_sal_interrupts  = _build_salhadaar_interrupt_table([w]) if "Salhadaar" in boss_name else ""
                wipe_panes += (f'<div class="pull-pane" id="pane-{wtbl_id}">'
                               f'{wipe_sal_wipe}{wipe_vorasius_swap}{wipe_banners}'
                               f'{wipe_table}{wipe_chart}{wipe_sal_interrupts}</div>')

            pull_selector = f'<div class="pull-selector">{pull_btns}</div>'
            kill_pane     = f'<div class="pull-pane active" id="pane-{table_id}">{kill_content}</div>'
            results[boss_name] = pull_selector + kill_pane + wipe_panes
        else:
            results[boss_name] = kill_content

    # ── Wipe-only bosses (not killed this raid) ──
    for boss_name, boss_wipes in (wipe_data or {}).items():
        if boss_name in results or not boss_wipes:
            continue
        boss_name_base = boss_name.rsplit(" (", 1)[0]
        mech_defs      = BOSS_MECHANICS.get(boss_name_base, [])
        has_interrupts = boss_name_base in BOSS_HAS_INTERRUPTS
        table_id  = f"{id_prefix}-wipeonly-{boss_name_base.lower().replace(' ', '-').replace('&', 'and')}"
        total     = len(boss_wipes)

        for wi, w in enumerate(boss_wipes):
            bpct  = w.get("boss_pct", 0)
            dur_s = w.get("fight_dur_ms", 0) // 1000
            wm, ws = divmod(dur_s, 60)
            w["_row_label"] = f"💀 Wipe {wi + 1} — {bpct}% boss HP · {wm}:{ws:02d}"

        pull_btns  = ""
        wipe_panes = ""
        for wi, w in enumerate(reversed(boss_wipes)):
            actual_wipe_num = total - wi
            bpct    = w.get("boss_pct", 0)
            dur_s   = w.get("fight_dur_ms", 0) // 1000
            wm, ws  = divmod(dur_s, 60)
            wtbl_id = f"wipe-{table_id}-{wi}"
            active  = "active" if wi == 0 else ""
            btn_cls = "pull-btn wipe active" if wi == 0 else "pull-btn wipe"
            pull_btns += f'<button class="{btn_cls}" onclick="switchPull(this,\'{wtbl_id}\')">Wipe {actual_wipe_num} · {bpct}% · {wm}:{ws:02d}</button>'
            wipe_pids = sorted(
                {p for p in (set(w.get("all_player_ids", [])) | set(w.get("deaths", {}).keys())) if p in actor_lookup},
                key=pid_sort
            )
            wipe_banners        = _build_banners([w])
            wipe_table          = _render_table([w], wtbl_id, pids=wipe_pids)
            wipe_chart          = _build_chart(w, wtbl_id)
            crown_html          = _build_crown_mechanics_html(w, actor_lookup) if "Crown"     in boss_name_base else ""
            wo_sal_wipe         = _build_salhadaar_wipe_banner([w])            if "Salhadaar" in boss_name_base else ""
            wo_vorasius_swap    = _build_vorasius_tank_swap_banner([w])        if "Vorasius"  in boss_name_base else ""
            wo_sal_interrupts   = _build_salhadaar_interrupt_table([w])        if "Salhadaar" in boss_name_base else ""
            wipe_panes  += (f'<div class="pull-pane {active}" id="pane-{wtbl_id}">'
                            f'{wo_sal_wipe}{wo_vorasius_swap}{wipe_banners}'
                            f'{wipe_table}{wipe_chart}{crown_html}{wo_sal_interrupts}</div>')

        pull_selector = f'<div class="pull-selector">{pull_btns}</div>'
        results[boss_name] = pull_selector + wipe_panes

    return results


_FILTER_BAR_CSS = """/* ── Filter bar ── */
.filter-bar { display: flex; flex-wrap: wrap; align-items: center; gap: 8px; margin-bottom: 16px; padding: 10px 14px; background: #0d1525; border-radius: 8px; border: 1px solid #2a2a4a; }
.filter-label { font-size: 11px; color: #555; text-transform: uppercase; letter-spacing: 0.5px; white-space: nowrap; margin-right: 2px; }
.filter-tag { background: #16213e; color: #888; border: 1px solid #2a2a4a; padding: 3px 10px; border-radius: 20px; font-size: 12px; cursor: pointer; transition: all 0.15s; white-space: nowrap; }
.filter-tag:hover { color: #bbb; border-color: #4a5568; }
.filter-tag.active { background: #2a3a6a; color: #7289DA; border-color: #7289DA; font-weight: 600; }
.filter-divider { width: 1px; height: 20px; background: #2a2a4a; margin: 0 4px; flex-shrink: 0; }
.chip-input-area { position: relative; }
.chip-input-wrap { display: inline-flex; align-items: center; flex-wrap: wrap; gap: 4px; background: #16213e; border: 1px solid #2a2a4a; border-radius: 6px; padding: 4px 8px; min-width: 180px; cursor: text; }
.chip-input-wrap:focus-within { border-color: #7289DA; }
.chip { display: inline-flex; align-items: center; gap: 4px; background: #2a3a6a; color: #a0b4ff; border-radius: 20px; padding: 2px 8px 2px 10px; font-size: 12px; }
.chip-remove { background: none; border: none; color: #7289DA; cursor: pointer; font-size: 13px; padding: 0; line-height: 1; }
.chip-remove:hover { color: #e57373; }
.chip-text-input { background: none; border: none; color: #e0e0e0; font-size: 12px; outline: none; min-width: 90px; padding: 1px 0; }
.chip-text-input::placeholder { color: #555; }
.chip-dropdown { position: absolute; top: 100%; left: 0; min-width: 180px; background: #1a2233; border: 1px solid #4a5568; border-radius: 6px; z-index: 100; margin-top: 2px; box-shadow: 0 4px 12px #0008; }
.chip-dd-item { padding: 6px 12px; font-size: 12px; color: #e0e0e0; cursor: pointer; }
.chip-dd-item:hover { background: #2a3a6a; color: #a0b4ff; }
.filter-clear { background: none; border: 1px solid #3a2a2a; color: #666; padding: 3px 10px; border-radius: 20px; font-size: 11px; cursor: pointer; white-space: nowrap; }
.filter-clear:hover { color: #e57373; border-color: #e57373; }
"""

_FILTER_BAR_JS = """
const _CLASS_COLORS_FB = {DeathKnight:'#C41E3A',DemonHunter:'#A330C9',Druid:'#FF7C0A',Evoker:'#33937F',Hunter:'#AAD372',Mage:'#3FC7EB',Monk:'#00FF98',Paladin:'#F48CBA',Priest:'#FFFFFF',Rogue:'#FFF468',Shaman:'#0070DD',Warlock:'#8788EE',Warrior:'#C69B3A'};
function buildClassTags(splitId) {
  const tab = document.getElementById('tab-' + splitId);
  const ct = document.getElementById('class-tags-' + splitId);
  if (!tab || !ct || ct.children.length) return;
  const classes = [...new Set([...tab.querySelectorAll('tr[data-class]')].map(r => r.dataset.class))].sort();
  classes.forEach(cls => {
    const btn = document.createElement('button');
    btn.className = 'filter-tag'; btn.dataset.ftype = 'class'; btn.dataset.fval = cls;
    btn.style.cssText = 'border-left:3px solid ' + (_CLASS_COLORS_FB[cls] || '#aaa') + ';';
    btn.textContent = cls;
    btn.onclick = () => { btn.classList.toggle('active'); applyFilters(splitId); };
    ct.appendChild(btn);
  });
}
function toggleTag(btn, splitId) { btn.classList.toggle('active'); applyFilters(splitId); }
function addChip(splitId, value) {
  value = value.trim().toLowerCase();
  if (!value) return;
  const ct = document.getElementById('chips-' + splitId);
  if (!ct || [...ct.querySelectorAll('.chip')].some(c => c.dataset.val === value)) return;
  const chip = document.createElement('span');
  chip.className = 'chip'; chip.dataset.val = value;
  chip.textContent = value;
  const rm = document.createElement('button');
  rm.className = 'chip-remove'; rm.title = 'Remove'; rm.textContent = '\u00d7';
  rm.onclick = () => { chip.remove(); applyFilters(splitId); };
  chip.appendChild(rm);
  ct.appendChild(chip); applyFilters(splitId);
}
function chipKey(e, splitId) {
  const inp = e.target;
  if ((e.key === 'Enter' || e.key === ',') && inp.value.trim()) {
    e.preventDefault(); addChip(splitId, inp.value); inp.value = ''; hideDropdown(splitId);
  } else if (e.key === 'Backspace' && !inp.value) {
    const ct = document.getElementById('chips-' + splitId);
    if (ct && ct.lastElementChild) { ct.lastElementChild.remove(); applyFilters(splitId); }
  } else if (e.key === 'Escape') { hideDropdown(splitId); }
}
function chipSuggest(inp, splitId) {
  const val = inp.value.toLowerCase().trim();
  const dd = document.getElementById('chip-dd-' + splitId);
  if (!val || !dd) { if (dd) dd.style.display = 'none'; return; }
  const tab = document.getElementById('tab-' + splitId);
  if (!tab) return;
  const names = [...new Set([...tab.querySelectorAll('tr[data-player]')].map(r => r.dataset.player))].filter(n => n.includes(val)).slice(0, 8);
  if (!names.length) { dd.style.display = 'none'; return; }
  dd.innerHTML = '';
  names.forEach(n => {
    const item = document.createElement('div');
    item.className = 'chip-dd-item'; item.textContent = n;
    item.onmousedown = (ev) => { ev.preventDefault(); addChip(splitId, n); inp.value = ''; hideDropdown(splitId); };
    dd.appendChild(item);
  });
  dd.style.display = 'block';
}
function hideDropdown(splitId) { const dd = document.getElementById('chip-dd-' + splitId); if (dd) dd.style.display = 'none'; }
function clearFilters(splitId) {
  const fb = document.getElementById('fb-' + splitId);
  if (fb) fb.querySelectorAll('.filter-tag.active').forEach(b => b.classList.remove('active'));
  const ct = document.getElementById('chips-' + splitId);
  if (ct) ct.innerHTML = '';
  const inp = document.getElementById('chip-in-' + splitId);
  if (inp) inp.value = '';
  hideDropdown(splitId); applyFilters(splitId);
}
function applyFilters(splitId) {
  const fb = document.getElementById('fb-' + splitId);
  const tab = document.getElementById('tab-' + splitId);
  if (!fb || !tab) return;
  const activeRoles   = new Set([...fb.querySelectorAll('[data-ftype="role"].active')].map(b => b.dataset.fval));
  const activeClasses = new Set([...fb.querySelectorAll('[data-ftype="class"].active')].map(b => b.dataset.fval));
  const ct = document.getElementById('chips-' + splitId);
  const activeNames = ct ? [...ct.querySelectorAll('.chip')].map(c => c.dataset.val) : [];
  const hasFilters = activeRoles.size || activeClasses.size || activeNames.length;
  tab.querySelectorAll('table').forEach(tbl => {
    const tbody = tbl.tBodies[0]; if (!tbody) return;
    const roleVis = {};
    [...tbody.rows].forEach(row => {
      if (row.classList.contains('split-ov-row') || row.classList.contains('section-sep') || row.classList.contains('role-sep')) return;
      const role = row.dataset.role; if (!role) return;
      let show = true;
      if (hasFilters) {
        const roleOk  = !activeRoles.size   || activeRoles.has(role);
        const classOk = !activeClasses.size || activeClasses.has(row.dataset.class);
        const nameOk  = !activeNames.length || activeNames.some(n => (row.dataset.player || '').includes(n));
        show = roleOk && classOk && nameOk;
      }
      row.style.display = show ? '' : 'none';
      if (show) roleVis[role] = true;
    });
    [...tbody.rows].forEach(row => {
      if (row.classList.contains('role-sep'))
        row.style.display = (!hasFilters || roleVis[row.dataset.role]) ? '' : 'none';
    });
  });
}
function sortAddTable(th) {
  const col = parseInt(th.dataset.col);
  const table = th.closest('table');
  const tbody = table.querySelector('tbody');
  const rows  = [...tbody.querySelectorAll('tr')];
  const asc   = th.dataset.dir !== 'asc';
  // Clear indicators
  table.querySelectorAll('th[data-col]').forEach(h => { h.dataset.dir = ''; h.querySelector('.sort-arrow') && (h.querySelector('.sort-arrow').textContent = ''); });
  th.dataset.dir = asc ? 'asc' : 'desc';
  const arrow = th.querySelector('.sort-arrow');
  if (arrow) arrow.textContent = asc ? ' ▲' : ' ▼';
  rows.sort((a, b) => {
    const va = parseFloat(a.cells[col]?.dataset.val ?? 0);
    const vb = parseFloat(b.cells[col]?.dataset.val ?? 0);
    return asc ? va - vb : vb - va;
  });
  rows.forEach(r => tbody.appendChild(r));
}
"""


def _filter_bar_html(split_id: str) -> str:
    si = split_id.replace("'", "\\'")
    return (
        f'<div class="filter-bar" id="fb-{split_id}">'
        f'<span class="filter-label">Role</span>'
        f'<button class="filter-tag" data-ftype="role" data-fval="Tank" onclick="toggleTag(this,\'{si}\')">🛡 Tank</button>'
        f'<button class="filter-tag" data-ftype="role" data-fval="Healer" onclick="toggleTag(this,\'{si}\')">💚 Healer</button>'
        f'<button class="filter-tag" data-ftype="role" data-fval="DPS" onclick="toggleTag(this,\'{si}\')">⚔ DPS</button>'
        f'<div class="filter-divider"></div>'
        f'<span class="filter-label">Class</span>'
        f'<span class="class-tags" id="class-tags-{split_id}"></span>'
        f'<div class="filter-divider"></div>'
        f'<div class="chip-input-area">'
        f'<div class="chip-input-wrap" onclick="this.querySelector(\'.chip-text-input\').focus()">'
        f'<span class="chips-container" id="chips-{split_id}"></span>'
        f'<input class="chip-text-input" id="chip-in-{split_id}" placeholder="Player name\u2026"'
        f' onkeydown="chipKey(event,\'{si}\')" oninput="chipSuggest(this,\'{si}\')"'
        f' onblur="setTimeout(()=>hideDropdown(\'{si}\'),150)">'
        f'</div>'
        f'<div class="chip-dropdown" id="chip-dd-{split_id}" style="display:none"></div>'
        f'</div>'
        f'<button class="filter-clear" onclick="clearFilters(\'{si}\')">✕ Clear all</button>'
        f'</div>'
    )


def write_raid_html(day_data: dict, output_path: str) -> None:
    """Write a single raid day as a standalone HTML page (Gear + Split tabs)."""
    ri         = day_data["report_info"]
    rc         = day_data["report_code"]
    date_str   = datetime.fromtimestamp(ri["startTime"] / 1000, tz=timezone.utc).strftime("%Y-%m-%d") if ri.get("startTime") else ""
    diff_label = day_data.get("difficulty", "")
    title      = ri.get("title", rc) + (f" — {diff_label}" if diff_label else "")

    # Gear lives in gear_normal.html — not shown on individual raid pages.
    show_gear = False

    gear_tab_html = ""
    gear_tab_btn  = ""

    if show_gear:
        gear_header, gear_rows, _, total_players, total_crafted, no_craft = _build_gear_html(day_data["records"])
        gear_inner = f"""
  <div class="stats">
    <div class="stat-box"><div class="num">{total_players}</div><div class="label">Players</div></div>
    <div class="stat-box"><div class="num">{total_crafted}</div><div class="label">Crafted Items</div></div>
    <div class="stat-box warn"><div class="num">{no_craft}</div><div class="label">No Crafted Gear</div></div>
  </div>
  <div class="search-box"><input type="text" placeholder="Search player..." onkeyup="filterGear(this)"></div>
  <div class="table-wrap">
    <table id="gear-table"><thead><tr>{gear_header}</tr></thead><tbody>{gear_rows}</tbody></table>
  </div>"""
        gear_tab_btn  = '<button class="tab-btn" onclick="switchTab(\'gear\',this)">⚙ Gear</button>\n'
        gear_tab_html = f'<div id="tab-gear" class="tab-content">{gear_inner}\n</div>'

    # ── Build Boss tabs (one tab per boss) ──
    actor_lookup = {a["id"]: a for a in day_data["actors"]}

    # Merge all fights per boss (split_num is preserved inside each fight dict)
    merged_boss_data: dict = {}
    for boss_name, fights in day_data["boss_data"].items():
        merged_boss_data.setdefault(boss_name, []).extend(fights)

    boss_htmls = _build_boss_html(merged_boss_data, actor_lookup, id_prefix="b",
                                  wipe_data=day_data.get("wipe_data", {}))

    tab_buttons = ""
    split_divs  = ""
    for bi, (boss_name, boss_html) in enumerate(reversed(list(boss_htmls.items()))):
        boss_display = boss_name.rsplit(" (", 1)[0]  # strip difficulty suffix
        tab_id    = f"boss-{bi}"
        active    = "active" if bi == 0 else ""
        fb_html   = _filter_bar_html(tab_id)
        wipe_only = boss_name not in merged_boss_data
        btn_label = f'💀 {escape(boss_display)}' if wipe_only else escape(boss_display)
        btn_style = ' style="color:#e57373"' if wipe_only else ''
        tab_buttons += f'<button class="tab-btn {active}"{btn_style} onclick="switchTab(\'{tab_id}\',this)">{btn_label}</button>\n'
        split_divs  += f'<div id="tab-{tab_id}" class="tab-content {active}">{fb_html}{boss_html}</div>\n'

    tab_buttons += gear_tab_btn  # gear always last

    wcl_link = f'<a href="https://www.warcraftlogs.com/reports/{rc}" target="_blank" class="wcl-link">View on WarcraftLogs ↗</a>'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{escape(title)} — {date_str} — Raid Audit</title>
{WOWHEAD_SCRIPTS}
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: #1a1a2e; color: #e0e0e0; font-family: -apple-system, 'Segoe UI', sans-serif; padding: 20px; }}
h1 {{ color: #7289DA; font-size: 22px; margin-bottom: 4px; }}
.breadcrumb {{ font-size: 12px; color: #555; margin-bottom: 14px; }}
.breadcrumb a {{ color: #7289DA; text-decoration: none; }}
.breadcrumb a:hover {{ text-decoration: underline; }}
.raid-meta {{ display: flex; align-items: center; gap: 12px; margin-bottom: 18px; }}
.raid-date {{ color: #666; font-size: 13px; }}
.wcl-link {{ color: #7289DA; font-size: 12px; text-decoration: none; }}
.wcl-link:hover {{ text-decoration: underline; }}
/* ── Tabs ── */
.tab-bar {{ display: flex; gap: 4px; margin-bottom: 20px; border-bottom: 2px solid #2a2a4a; flex-wrap: wrap; }}
.tab-btn {{ background: #16213e; color: #888; border: none; padding: 10px 20px; border-radius: 8px 8px 0 0; font-size: 13px; font-weight: 600; cursor: pointer; border-bottom: 2px solid transparent; margin-bottom: -2px; transition: all 0.15s; }}
.tab-btn:hover {{ color: #bbb; background: #1e2d50; }}
.tab-btn.active {{ color: #7289DA; background: #1a1a2e; border-bottom: 2px solid #7289DA; }}
.tab-content {{ display: none; }}
.tab-content.active {{ display: block; }}
/* ── Boss section ── */
.boss-section {{ margin-bottom: 4px; }}
.boss-section-title {{ color: #a0b4ff; font-size: 14px; font-weight: 700; margin-bottom: 0; padding: 7px 12px; background: #111827; border-radius: 4px; border-left: 3px solid #7289DA; cursor: pointer; user-select: none; display: flex; justify-content: space-between; align-items: center; }}
.boss-section-title:hover {{ background: #1a2236; }}
.boss-toggle-arrow {{ font-size: 12px; transition: transform 0.2s; }}
.boss-section-title.collapsed .boss-toggle-arrow {{ transform: rotate(-90deg); }}
.boss-section-body {{ padding-top: 10px; margin-bottom: 20px; }}
.boss-section-body.collapsed {{ display: none; }}
/* ── Stats ── */
.stats {{ display: flex; gap: 16px; margin-bottom: 16px; flex-wrap: wrap; }}
.stat-box {{ background: #16213e; border-radius: 8px; padding: 10px 18px; }}
.stat-box .num {{ font-size: 24px; font-weight: bold; color: #7289DA; }}
.stat-box .label {{ font-size: 11px; color: #888; text-transform: uppercase; }}
.stat-box.warn .num {{ color: #ffc107; }}
/* ── Search ── */
.search-box {{ margin: 0 0 12px; }}
.search-box input {{ background: #16213e; border: 1px solid #2a2a4a; color: #e0e0e0; padding: 7px 14px; border-radius: 6px; width: 280px; font-size: 13px; }}
.search-box input::placeholder {{ color: #555; }}
/* ── Tables ── */
.table-wrap {{ overflow-x: auto; border-radius: 8px; margin-bottom: 28px; }}
table {{ border-collapse: collapse; min-width: 100%; }}
th {{ background: #16213e; color: #7289DA; padding: 8px 10px; text-align: left; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; white-space: nowrap; position: sticky; top: 0; z-index: 1; }}
th.player-header {{ position: sticky; left: 0; z-index: 2; background: #16213e; min-width: 140px; }}
td {{ padding: 6px 10px; border-bottom: 1px solid rgba(255,255,255,0.05); font-size: 13px; white-space: nowrap; }}
.player-cell {{ font-weight: bold; position: sticky; left: 0; background: #0d1f3c; z-index: 1; min-width: 140px; }}
tr:hover td {{ background: rgba(114,137,218,0.07); }}
tr:hover .player-cell {{ background: #152848; }}
td.divider, th.divider {{ border-left: 2px solid rgba(114,137,218,0.25); }}
.center {{ text-align: center; }}
.empty {{ color: #2a2a4a; text-align: center; }}
td, th {{ border-left: 1px solid rgba(255,255,255,0.07); }}
td:first-child, th:first-child {{ border-left: none; }}
/* ── Gear ── */
.no-craft {{ background: rgba(255,160,0,0.05); }}
.no-craft .player-cell {{ background: rgba(80,50,0,0.4); }}
.spark-yes {{ color: #4caf50; font-weight: bold; }}
.item-cell a {{ color: #a48cff; text-decoration: none; }}
tr.section-sep td {{ color: #555; font-size: 11px; padding: 4px 10px; background: #0d1117; letter-spacing: 0.5px; }}
.item-cell a:hover {{ text-decoration: underline; }}
/* ── Roles ── */
.role-sep td {{ background: #111827; color: #7289DA; font-size: 11px; font-weight: bold; padding: 4px 10px; }}
.pname {{ font-weight: bold; text-decoration: none; }}
.pname:hover {{ text-decoration: underline; opacity: 0.85; }}
.cname {{ color: #888; font-size: 11px; }}
/* ── Split overview row ── */
.split-ov-row td {{ padding: 0 !important; border: none; background: #0d1525; }}
.split-ov-bar {{ display: flex; align-items: center; flex-wrap: wrap; gap: 16px; padding: 10px 14px; border-top: 2px solid #2a3a6a; border-bottom: 1px solid #1e2a4a; }}
.sov-title {{ color: #a0b4ff; font-weight: 700; font-size: 14px; min-width: 60px; }}
.sov-dur {{ color: #666; font-size: 12px; }}
.sov-item {{ font-size: 12px; color: #ccc; }}
.sov-lbl {{ color: #666; margin-right: 4px; }}
/* ── Boss columns ── */
.death-h {{ color: #e57373; }}
.dmg-h {{ color: #ffb74d; }}
.heal-h {{ color: #81c784; }}
.parse-h {{ color: #E5CC80; }}
.uptime-h {{ color: #81c784; }}
.interrupt-h {{ color: #64b5f6; }}
.avoid-h {{ color: #ce93d8; }}
.detail-h {{ color: #7289DA; white-space: normal; min-width: 30px; cursor: pointer; user-select: none; }}
.detail-h:hover {{ color: #a0b4ff; }}
.mech-bad-h {{ color: #ff7070; }}
.mech-soak-h {{ color: #66bb6a; }}
.ext-h {{ color: #81c784; }}
.ext-num {{ color: #81c784; font-weight: bold; }}
.def-h {{ color: #64b5f6; }}
.def-num {{ color: #64b5f6; font-weight: bold; }}
.detail-cell {{ white-space: normal; min-width: 180px; font-size: 12px; color: #ccc; }}
.detail-col-hidden .detail-h span.detail-content,
.detail-col-hidden .detail-cell {{ display: none; }}
.detail-col-hidden .detail-h {{ min-width: 0; }}
.boss-death-row td {{ background: rgba(61,16,16,0.4); }}
.boss-death-row .player-cell {{ background: rgba(80,10,10,0.6); }}
.death-num {{ color: #e57373; font-weight: bold; }}
.bighit-row {{ color: #ff8a65; font-weight: bold; }}
/* ── Frontal failures ── */
.frontal-failures {{ margin-bottom: 12px; background: rgba(229,115,115,0.08); border: 1px solid rgba(229,115,115,0.3); border-radius: 8px; overflow: hidden; }}
.ff-header {{ display: flex; align-items: center; justify-content: space-between; padding: 8px 14px; cursor: pointer; user-select: none; }}
.ff-header:hover {{ background: rgba(229,115,115,0.12); }}
.ff-title {{ color: #e57373; font-weight: 700; font-size: 13px; }}
.ff-chevron {{ color: #e57373; font-size: 11px; transition: transform 0.2s; }}
.ff-header.open .ff-chevron {{ transform: rotate(180deg); }}
.ff-body {{ display: none; padding: 4px 0 8px; }}
.ff-body.open {{ display: block; }}
.ff-item {{ padding: 3px 14px; font-size: 12px; border-top: 1px solid rgba(255,255,255,0.04); }}
.ff-split {{ color: #888; margin-right: 6px; }}
.ff-time {{ color: #a0b4ff; font-weight: 600; margin-right: 6px; }}
.ff-label {{ color: #e57373; font-weight: 600; margin-right: 4px; }}
.ff-primary {{ color: #f4a742; font-weight: 700; }}
.ff-sep {{ color: #555; margin: 0 6px; }}
.ff-players {{ color: #e06c6c; }}
/* ── Avoidable Hits ── */
.avoid-failures {{ margin-bottom: 10px; background: rgba(244,167,66,0.07); border: 1px solid rgba(244,167,66,0.3); border-radius: 8px; overflow: hidden; }}
.av-header {{ display: flex; align-items: center; justify-content: space-between; padding: 8px 14px; cursor: pointer; user-select: none; }}
.av-header:hover {{ background: rgba(244,167,66,0.12); }}
.av-title {{ color: #f4a742; font-weight: 700; font-size: 13px; }}
.av-chevron {{ color: #f4a742; font-size: 11px; transition: transform 0.2s; }}
.av-header.open .av-chevron {{ transform: rotate(180deg); }}
.av-body {{ display: none; padding: 4px 0 8px; }}
.av-body.open {{ display: block; }}
.av-item {{ padding: 3px 14px; font-size: 12px; border-top: 1px solid rgba(255,255,255,0.04); }}
.av-label {{ color: #f4a742; font-weight: 600; margin-right: 4px; }}
.av-player {{ color: #e0e0e0; }}
.av-count {{ color: #e06c6c; font-weight: 700; margin-right: 8px; }}
/* ── Tank swap warning (Vorasius) ── */
.tank-swap-warn {{ margin-bottom: 12px; background: rgba(244,167,66,0.08); border: 1px solid rgba(244,167,66,0.4); border-radius: 8px; overflow: hidden; }}
.ts-header {{ display: flex; align-items: center; padding: 8px 14px; }}
.ts-title {{ color: #f4a742; font-weight: 700; font-size: 13px; }}
.ts-body {{ padding: 4px 0 8px; }}
.ts-item {{ padding: 3px 14px; font-size: 12px; border-top: 1px solid rgba(255,255,255,0.04); }}
.ts-player {{ color: #f4a742; font-weight: 700; }}
.ts-time {{ color: #a0b4ff; font-weight: 600; }}
/* ── Salhadaar wipe condition banner ── */
.salhadaar-wipe-warn {{ margin-bottom: 12px; background: rgba(229,115,115,0.1); border: 1px solid rgba(229,115,115,0.5); border-radius: 8px; overflow: hidden; }}
.sw-header {{ display: flex; align-items: center; padding: 8px 14px; }}
.sw-title {{ color: #e57373; font-weight: 700; font-size: 13px; }}
.sw-body {{ padding: 4px 0 8px; }}
.sw-item {{ padding: 3px 14px; font-size: 12px; border-top: 1px solid rgba(255,255,255,0.04); }}
.sw-time {{ color: #a0b4ff; font-weight: 600; margin-right: 6px; }}
.sw-event {{ color: #e57373; font-weight: 700; }}
/* ── Alndust Upheaval grouping panel ── */
.hw-title {{ font-size: 13px; font-weight: 600; color: #aaa; padding: 10px 0 6px; letter-spacing: .04em; }}
.hw-tbl {{ margin-bottom: 8px; }}
.hw-wave-hdr {{ text-align: center; font-size: 12px; color: #9ab; padding: 4px 8px; background: #0f1a28; }}
.hw-timing {{ font-size: 11px; color: #667; font-weight: 400; }}
.hw-player-col {{ min-width: 110px; }}
.hw-sub {{ font-size: 11px; color: #556; font-weight: 400; text-align: center; padding: 2px 6px; }}
.hw-sub-dmg {{ min-width: 60px; }}
.hw-name {{ font-size: 12px; white-space: nowrap; padding-right: 12px; }}
.hw-phase-cell {{ text-align: center; padding: 3px 6px; }}
.hw-dmg-cell {{ text-align: right; padding: 3px 8px; font-size: 12px; font-variant-numeric: tabular-nums; }}
.hw-phase-down {{ color: #e57373; font-size: 11px; font-weight: 600; }}
.hw-phase-up   {{ color: #64b5f6; font-size: 11px; font-weight: 600; }}
.hw-assist {{ color: #81c784; font-size: 11px; }}
.hw-slack  {{ color: #f4a742; font-size: 12px; cursor: help; }}
.horror-waves-wrap {{ margin: 16px 0 8px; overflow-x: auto; }}
.hw-tbl thead {{ position: static; }}
.hw-group-hdr td {{ font-size: 10px; font-weight: 700; letter-spacing: .08em; text-transform: uppercase; color: #556; background: #0b1520; padding: 4px 8px; border-top: 1px solid #1a2a3a; }}
.alndust-panel {{ margin-top: 12px; background: rgba(100,160,255,0.06); border: 1px solid rgba(100,160,255,0.25); border-radius: 8px; overflow: hidden; }}
.ag-header {{ display: flex; align-items: center; justify-content: space-between; padding: 8px 14px; cursor: pointer; user-select: none; }}
.ag-header:hover {{ background: rgba(100,160,255,0.1); }}
.ag-title {{ color: #7eb8ff; font-weight: 700; font-size: 13px; }}
.ag-chevron {{ color: #7eb8ff; font-size: 11px; transition: transform 0.2s; }}
.ag-header.open .ag-chevron {{ transform: rotate(180deg); }}
.ag-body {{ display: none; padding: 6px 12px 10px; }}
.ag-body.open {{ display: block; }}
.ag-wave {{ margin-bottom: 10px; padding-bottom: 10px; border-bottom: 1px solid rgba(255,255,255,0.05); }}
.ag-wave:last-child {{ margin-bottom: 0; border-bottom: none; }}
.ag-wave-label {{ color: #aac4ff; font-weight: 700; font-size: 12px; margin-bottom: 6px; }}
.ag-time {{ color: #5580aa; font-weight: 400; }}
.ag-group {{ display: flex; flex-wrap: wrap; gap: 5px; align-items: center; margin-bottom: 4px; }}
.ag-badge {{ font-size: 11px; font-weight: 700; padding: 2px 7px; border-radius: 10px; white-space: nowrap; }}
.ag-badge-down {{ background: rgba(229,115,115,0.2); color: #e57373; border: 1px solid rgba(229,115,115,0.4); }}
.ag-badge-up {{ background: rgba(129,199,132,0.15); color: #81c784; border: 1px solid rgba(129,199,132,0.3); }}
.ag-chip {{ color: #ccc; font-size: 12px; white-space: nowrap; padding: 2px 4px; }}
.ag-chip.ag-miss {{ background: rgba(229,115,115,0.2); border: 1px solid rgba(229,115,115,0.6); color: #e57373; font-weight: 700; padding: 2px 8px; border-radius: 12px; }}
.ag-chip.ag-wrong {{ background: rgba(244,167,66,0.15); border: 1px solid rgba(244,167,66,0.5); color: #f4a742; font-weight: 600; padding: 2px 8px; border-radius: 12px; }}
.ag-chip.ag-double-up {{ background: rgba(255,183,77,0.1); border: 1px solid rgba(255,183,77,0.4); color: #ffb74d; font-weight: 600; padding: 2px 8px; border-radius: 12px; }}
.ag-miss-count {{ color: #e57373; font-size: 11px; font-weight: 700; margin-left: 8px; }}
.ag-double-up-count {{ color: #ffb74d; font-size: 11px; font-weight: 700; margin-left: 8px; }}
.ag-chip.ag-chip-heal {{ background: rgba(129,199,132,0.12); border: 1px solid rgba(129,199,132,0.35); color: #81c784; padding: 2px 8px; border-radius: 12px; }}
.vm-arrow {{ color: #555; font-size: 13px; }}
.vm-expired {{ color: #555; font-size: 12px; font-style: italic; }}
.add-dmg-table {{ width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 4px; }}
.add-dmg-table th {{ color: #888; font-weight: 600; text-align: left; padding: 4px 8px; border-bottom: 1px solid #2a2a4a; }}
.add-dmg-table td {{ padding: 3px 8px; border-bottom: 1px solid #1a1a2e; color: #ccc; }}
.add-dmg-table tr:last-child td {{ border-bottom: none; }}
.add-dmg-val {{ text-align: right; color: #aaa; font-variant-numeric: tabular-nums; }}
/* ── Sort arrows ── */
.sort-arrow {{ opacity: 0.6; font-size: 11px; margin-left: 4px; }}
th[data-sortable]:hover .sort-arrow {{ opacity: 1; }}
/* ── Pull selector ── */
.pull-selector {{ display: flex; flex-wrap: wrap; gap: 4px; margin-bottom: 10px; }}
.pull-btn {{ background: #16213e; color: #888; border: 1px solid #2a2a4a; padding: 4px 12px; border-radius: 4px; font-size: 12px; cursor: pointer; transition: all 0.15s; }}
.pull-btn:hover {{ color: #bbb; }}
.pull-btn.active {{ background: #1a3a2a; color: #4caf50; border-color: #4caf50; font-weight: 600; }}
.pull-btn.wipe.active {{ background: #3a1a1a; color: #e57373; border-color: #e57373; }}
.pull-pane {{ display: none; }}
.pull-pane.active {{ display: block; }}
{_FILTER_BAR_CSS}
{_CROWN_MECHANICS_CSS}
</style>
<script>const whTooltips = {{colorLinks: true, iconizeLinks: true, iconSize: 'small'}};</script>
<script src="https://wow.zamimg.com/js/tooltips.js"></script>
</head>
<body>
<div class="breadcrumb"><a href="index.html">← Overview</a> / {escape(title)}</div>
<h1>{escape(title)}</h1>
<div class="raid-meta">
  <span class="raid-date">{date_str}</span>
  {wcl_link}
</div>
<div class="tab-bar">{tab_buttons}</div>

{gear_tab_html}
<!-- ── SPLIT TABS ── -->
{split_divs}
<script>
function switchTabByName(name) {{
  document.querySelectorAll('.tab-content').forEach(el => el.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(el => el.classList.remove('active'));
  const el = document.getElementById('tab-' + name);
  if (!el) return;
  el.classList.add('active');
  const btn = [...document.querySelectorAll('.tab-btn')].find(b => (b.getAttribute('onclick') || '').includes("'" + name + "'"));
  if (btn) btn.classList.add('active');
}}
function switchTab(name, btn) {{
  // Capture filter state from the current active tab before switching
  const prevEl = document.querySelector('.tab-content.active');
  const prevId = prevEl ? prevEl.id.replace('tab-', '') : null;
  let savedRoles = [], savedClasses = [], savedNames = [];
  if (prevId && prevId !== name) {{
    const prevFb = document.getElementById('fb-' + prevId);
    if (prevFb) {{
      savedRoles   = [...prevFb.querySelectorAll('[data-ftype="role"].active')].map(b => b.dataset.fval);
      savedClasses = [...prevFb.querySelectorAll('[data-ftype="class"].active')].map(b => b.dataset.fval);
    }}
    const prevChips = document.getElementById('chips-' + prevId);
    if (prevChips) savedNames = [...prevChips.querySelectorAll('.chip')].map(c => c.dataset.val);
  }}
  switchTabByName(name);
  // Sync filter state into new tab (clear first, then apply saved state)
  if (prevId && prevId !== name) {{
    const newFb = document.getElementById('fb-' + name);
    if (newFb) newFb.querySelectorAll('.filter-tag.active').forEach(b => b.classList.remove('active'));
    const newChips = document.getElementById('chips-' + name);
    if (newChips) newChips.innerHTML = '';
    if (newFb) {{
      savedRoles.forEach(v => {{ const b = newFb.querySelector(`[data-ftype="role"][data-fval="${{v}}"]`); if (b) b.classList.add('active'); }});
      savedClasses.forEach(v => {{ const b = newFb.querySelector(`[data-ftype="class"][data-fval="${{v}}"]`); if (b) b.classList.add('active'); }});
    }}
    savedNames.forEach(v => addChip(name, v));
    applyFilters(name);
  }}
  const tab = document.getElementById('tab-' + name);
}}
window.addEventListener('DOMContentLoaded', () => {{
  const h = location.hash.replace('#', '');
  if (h) switchTabByName(h);
  document.querySelectorAll('[id^="class-tags-"]').forEach(el => buildClassTags(el.id.replace('class-tags-', '')));
}});
function switchPull(btn, paneId) {{
  const body = btn.closest('.boss-section-body') || btn.closest('.pull-selector').parentElement;
  body.querySelectorAll('.pull-btn').forEach(b => b.classList.remove('active'));
  body.querySelectorAll('.pull-pane').forEach(p => p.classList.remove('active'));
  btn.classList.add('active');
  const pane = document.getElementById('pane-' + paneId);
  if (pane) {{ pane.classList.add('active'); }}
}}
function toggleBoss(titleEl) {{
  titleEl.classList.toggle('collapsed');
  titleEl.nextElementSibling.classList.toggle('collapsed');
}}
function filterGear(input) {{
  const filter = input.value.toLowerCase();
  for (let row of document.getElementById('gear-table').tBodies[0].rows)
    row.style.display = row.cells[0].textContent.toLowerCase().includes(filter) ? '' : 'none';
}}
function toggleDetails(tableId, btn) {{
  const tbl = document.getElementById(tableId);
  if (!tbl) return;
  const hidden = tbl.classList.toggle('detail-col-hidden');
  btn.textContent = hidden ? '▶ Details' : '▼ Details';
}}
const _htip = document.createElement('div');
_htip.id = 'htip';
Object.assign(_htip.style, {{ position:'fixed', display:'none', pointerEvents:'none', zIndex:'9999',
  background:'#1a2233', border:'1px solid #4a5568', borderRadius:'8px',
  padding:'10px 14px', fontSize:'12px', color:'#e0e0e0', maxWidth:'340px', lineHeight:'1.6', boxShadow:'0 4px 16px #0008' }});
document.body.appendChild(_htip);
document.addEventListener('mousemove', e => {{
  if (_htip.style.display !== 'none') {{
    const x = e.clientX + 16, y = e.clientY - 10;
    _htip.style.left = (x + _htip.offsetWidth > window.innerWidth ? e.clientX - _htip.offsetWidth - 8 : x) + 'px';
    _htip.style.top  = (y + _htip.offsetHeight > window.innerHeight ? e.clientY - _htip.offsetHeight - 8 : y) + 'px';
  }}
}});
function showHTip(el) {{ const d = el.getAttribute('data-htip'); if (!d) return; _htip.innerHTML = d; _htip.style.display = 'block'; }}
function hideHTip() {{ _htip.style.display = 'none'; }}
function sortBossTable(tableId, colIdx, thEl) {{
  const tbl = document.getElementById(tableId);
  if (!tbl) return;
  const dir = thEl.dataset.sortDir === 'desc' ? 'asc' : 'desc';
  tbl.querySelectorAll('th[data-sortable]').forEach(th => {{ th.dataset.sortDir = ''; const a = th.querySelector('.sort-arrow'); if (a) a.textContent = '▼'; }});
  thEl.dataset.sortDir = dir;
  const arrow = thEl.querySelector('.sort-arrow');
  if (arrow) arrow.textContent = dir === 'desc' ? '▼' : '▲';
  const tbody = tbl.tBodies[0];
  const rows = Array.from(tbody.rows);
  const segments = []; let cur = null;
  for (const row of rows) {{
    if (row.classList.contains('role-sep')) {{ if (cur) segments.push(cur); cur = {{ sep: row, rows: [] }}; }}
    else {{ if (!cur) cur = {{ sep: null, rows: [] }}; cur.rows.push(row); }}
  }}
  if (cur) segments.push(cur);
  function parseVal(cell) {{
    if (!cell) return null;
    const text = cell.textContent.trim().replace(/⇅|▲|▼/g, '').trim();
    if (!text || text === '—') return null;
    const m = text.replace('%','').match(/^([\d.]+)\s*([MkKBbG]?)$/);
    if (m) {{ let n = parseFloat(m[1]); const s = m[2].toUpperCase();
      if (s==='B') n*=1e9; else if (s==='M') n*=1e6; else if (s==='K') n*=1e3; return n; }}
    return text.toLowerCase();
  }}
  for (const seg of segments) {{
    seg.rows.sort((a, b) => {{
      const va = parseVal(a.cells[colIdx]), vb = parseVal(b.cells[colIdx]);
      if (va===null && vb===null) return 0; if (va===null) return 1; if (vb===null) return -1;
      if (typeof va==='number' && typeof vb==='number') return dir==='asc' ? va-vb : vb-va;
      if (va<vb) return dir==='asc'?-1:1; if (va>vb) return dir==='asc'?1:-1; return 0;
    }});
  }}
  for (const seg of segments) {{ if (seg.sep) tbody.appendChild(seg.sep); for (const row of seg.rows) tbody.appendChild(row); }}
}}
{_FILTER_BAR_JS}
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Raid page saved: {output_path}")


def build_player_profiles(days_data: list) -> dict:
    """Aggregate per-player stats across all reports/raids.
    Returns {pname: {chars, class, role, appearances: [{boss_key, difficulty, date_str, parse, deaths, dmg_taken, uptime_pct, interrupts, def_count, mechanics}]}}
    """
    from datetime import datetime, timezone
    profiles = {}
    for day_data in days_data:
        actor_lookup = {a["id"]: a for a in day_data.get("actors", [])}
        ri = day_data.get("report_info", {})
        start_ts = ri.get("startTime", 0)
        date_str = datetime.fromtimestamp(start_ts / 1000, tz=timezone.utc).strftime("%Y-%m-%d") if start_ts else "Unknown"
        difficulty = day_data.get("difficulty", "")

        for boss_key, fights in day_data.get("boss_data", {}).items():
            boss_display = boss_key.rsplit(" (", 1)[0]
            for fight in fights:
                fight_dur = fight.get("fight_dur_ms", 1) or 1
                for pid in fight.get("all_player_ids", set()):
                    actor = actor_lookup.get(pid, {})
                    char_name = actor.get("name", "")
                    cls = actor.get("subType", "Unknown")
                    if not char_name:
                        continue
                    pname, _ = lookup_roster(char_name)
                    role = (fight.get("spec_roles", {}).get(pid)
                            or PLAYER_ROLES.get(pname, "DPS"))

                    if pname not in profiles:
                        profiles[pname] = {"chars": set(), "class": cls, "role": role, "appearances": []}
                    p = profiles[pname]
                    p["chars"].add(char_name)
                    p["class"] = cls
                    p["role"] = role

                    rmap = fight.get("rankings_map", {}).get(pid, {})
                    parse = rmap.get("rankPercent", 0)
                    deaths = len(fight.get("deaths", {}).get(pid, []))
                    dmg_taken = fight.get("dmg_taken", {}).get(pid, 0)
                    uptime = fight.get("uptime_map", {}).get(pid, {})
                    uptime_pct = (min(uptime.get("activeTime", 0) / fight_dur * 100, 100)
                                  if isinstance(uptime, dict) else 0)
                    interrupts = fight.get("interrupts", {}).get(pid, 0)
                    def_count = len(fight.get("defensive_casts", {}).get(pid, []))
                    mechanics = dict(fight.get("mechanics_data", {}).get(pid, {}))

                    p["appearances"].append({
                        "char": char_name,
                        "boss": boss_display,
                        "difficulty": difficulty,
                        "date": date_str,
                        "parse": parse,
                        "deaths": deaths,
                        "dmg_taken": dmg_taken,
                        "uptime_pct": uptime_pct,
                        "interrupts": interrupts,
                        "def_count": def_count,
                        "mechanics": mechanics,
                    })
    return profiles


def write_player_pages(days_data: list, output_dir: str = "players") -> None:
    """Generate one player_{slug}.html profile page per known player."""
    from html import escape as _esc

    profiles = build_player_profiles(days_data)

    # Per-char class lookup
    char_classes: dict = {}
    for day_data in days_data:
        for actor in day_data.get("actors", []):
            name = actor.get("name", "")
            cls  = actor.get("subType", "Unknown")
            if name and cls != "Unknown":
                char_classes[name.lower()] = cls

    for pname, p in profiles.items():
        slug = _player_slug(pname)
        cls  = p.get("class", "Unknown")
        role = p.get("role", "DPS")
        cls_color   = CLASS_COLORS.get(cls, "#ccc")
        role_color  = {"Tank": "#64b5f6", "Healer": "#81c784", "DPS": "#e57373"}.get(role, "#ccc")

        # Char chips — colored by each char's own class, clickable for filter
        chars_html = ""
        for c in sorted(p["chars"]):
            char_cls   = char_classes.get(c.lower(), cls)
            char_color = CLASS_COLORS.get(char_cls, "#888")
            chars_html += (
                f'<span class="char-chip" style="color:{char_color}" '
                f'onclick="filterChar(this,\'{_esc(c.replace(chr(39), ""))}\')\"'
                f' title="Filter by {_esc(c)}">{_esc(c)}</span>'
            )

        # Build table rows with data-char attribute
        rows_html = ""
        for a in sorted(p["appearances"], key=lambda x: (x["boss"], x["date"])):
            parse = a["parse"]
            if parse >= 99:   pc = "#E5CC80"
            elif parse >= 95: pc = "#FF8000"
            elif parse >= 75: pc = "#A335EE"
            elif parse >= 50: pc = "#0070DD"
            elif parse >= 25: pc = "#1EFF00"
            else:              pc = "#aaa"
            parse_str  = f'<span style="color:{pc};font-weight:bold">{parse:.0f}%</span>' if parse else "—"
            dmg        = a["dmg_taken"]
            dmg_str    = (f"{dmg/1_000_000:.1f}M" if dmg >= 1_000_000 else (f"{dmg/1000:.0f}k" if dmg > 0 else "—"))
            uptime_str = f'{a["uptime_pct"]:.0f}%' if a["uptime_pct"] > 0 else "—"
            death_str  = f'<span style="color:#e57373;font-weight:bold">{a["deaths"]}</span>' if a["deaths"] else "—"
            int_str    = str(a["interrupts"]) if a["interrupts"] else "—"
            def_str    = str(a["def_count"]) if a["def_count"] else "—"
            mech_parts = [f'{k}:{v}' for k, v in a["mechanics"].items() if v]
            mech_str   = ", ".join(mech_parts) if mech_parts else "—"
            char_name  = a.get("char", "")
            char_cls   = char_classes.get(char_name.lower(), cls)
            char_color = CLASS_COLORS.get(char_cls, "#888")
            rows_html += (
                f'<tr data-char="{_esc(char_name)}">'
                f'<td>{_esc(a["boss"])}</td>'
                f'<td style="color:#888">{_esc(a["difficulty"])}</td>'
                f'<td style="color:#666">{_esc(a["date"])}</td>'
                f'<td class="center" style="color:{char_color};font-size:11px">{_esc(char_name)}</td>'
                f'<td class="center">{parse_str}</td>'
                f'<td class="center">{death_str}</td>'
                f'<td class="center">{dmg_str}</td>'
                f'<td class="center">{uptime_str}</td>'
                f'<td class="center">{int_str}</td>'
                f'<td class="center" style="color:#64b5f6">{def_str}</td>'
                f'<td style="color:#e57373;font-size:11px">{mech_str}</td>'
                f'</tr>'
            )

        def _sth(col, lbl):
            return (f'<th onclick="sortPerfTable({col},this)" style="cursor:pointer;user-select:none">'
                    f'{lbl} <span class="sarr">▼</span></th>')

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_esc(pname)} — The Rupture</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0d1117;color:#c9d1d9;font-family:'Segoe UI',sans-serif;font-size:13px;padding:24px}}
a{{color:#7289DA;text-decoration:none}} a:hover{{text-decoration:underline}}
h1{{font-size:28px;font-weight:700;margin-bottom:4px}}
h2{{font-size:13px;color:#7289DA;margin:24px 0 10px;letter-spacing:0.5px;text-transform:uppercase}}
.role-badge{{display:inline-block;padding:2px 10px;border-radius:12px;font-size:11px;font-weight:bold;color:#0d1117;background:{role_color};margin-left:10px;vertical-align:middle}}
.chars-list{{display:flex;flex-wrap:wrap;gap:8px;margin-top:4px}}
.char-chip{{background:#1a2233;border:1px solid #2a3a6a;border-radius:20px;padding:3px 14px;font-size:12px;font-weight:600;cursor:pointer;transition:border-color .15s,background .15s}}
.char-chip:hover{{background:#1e2d4a;border-color:#7289DA}}
.char-chip.active{{background:#2a3a6a;border-color:#7289DA;box-shadow:0 0 0 1px #7289DA}}
.filter-note{{color:#555;font-size:11px;margin-top:6px}}
table{{width:100%;border-collapse:collapse;background:#0d1525;margin-top:8px}}
th{{background:#111827;color:#7289DA;padding:8px 10px;text-align:left;font-size:11px;letter-spacing:0.5px;text-transform:uppercase;border-bottom:1px solid #2a3a6a}}
th:hover{{color:#a0b4ff}}
td{{padding:7px 10px;border-bottom:1px solid #151f2e;font-size:12px}}
tr:hover td{{background:#111827}}
.center{{text-align:center}}
.sarr{{color:#555;font-size:10px}}
.back{{margin-bottom:20px;display:inline-block;color:#7289DA;font-size:13px}}
</style>
</head>
<body>
<a class="back" href="../index.html">← Back to Raids</a>
<h1 style="color:{cls_color}">{_esc(pname)}<span class="role-badge">{_esc(role)}</span></h1>

<h2>Characters</h2>
<div class="chars-list">{chars_html}</div>
<div class="filter-note" id="filter-note"></div>

<h2>Boss Performance</h2>
<table id="perf-table">
<thead><tr>
{_sth(0,'Boss')}{_sth(1,'Difficulty')}{_sth(2,'Date')}
<th class="center">Char</th>
{_sth(4,'Parse%')}{_sth(5,'Deaths')}{_sth(6,'Dmg Taken')}{_sth(7,'Uptime%')}{_sth(8,'Interrupts')}{_sth(9,'Def Used')}
<th>Mechanics</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>
<script>
let _activeChar = null;
function filterChar(chip, charName) {{
  const same = _activeChar === charName;
  document.querySelectorAll('.char-chip').forEach(c => c.classList.remove('active'));
  _activeChar = same ? null : charName;
  if (!same) chip.classList.add('active');
  _applyFilter();
  const note = document.getElementById('filter-note');
  note.textContent = _activeChar ? 'Showing: ' + _activeChar + ' — click again to clear' : '';
}}
function _applyFilter() {{
  document.querySelectorAll('#perf-table tbody tr').forEach(row => {{
    row.style.display = (!_activeChar || row.dataset.char === _activeChar) ? '' : 'none';
  }});
}}
function sortPerfTable(col, th) {{
  const tbody = document.querySelector('#perf-table tbody');
  const rows = Array.from(tbody.rows);
  const dir = th.dataset.dir === 'asc' ? 'desc' : 'asc';
  document.querySelectorAll('#perf-table th').forEach(t => {{
    t.dataset.dir = '';
    const a = t.querySelector('.sarr'); if (a) a.textContent = '▼';
  }});
  th.dataset.dir = dir;
  const arr = th.querySelector('.sarr'); if (arr) arr.textContent = dir === 'asc' ? '▲' : '▼';
  rows.sort((a, b) => {{
    const ca = a.cells[col], cb = b.cells[col];
    const va = (ca?.textContent || '').trim().replace(/%|,/g,'');
    const vb = (cb?.textContent || '').trim().replace(/%|,/g,'');
    const na = parseFloat(va), nb = parseFloat(vb);
    if (!isNaN(na) && !isNaN(nb)) return dir === 'asc' ? na - nb : nb - na;
    return dir === 'asc' ? va.localeCompare(vb) : vb.localeCompare(va);
  }});
  rows.forEach(r => tbody.appendChild(r));
  _applyFilter();
}}
</script>
</body>
</html>"""

        os.makedirs(output_dir, exist_ok=True)
        out_path = os.path.join(output_dir, f"player_{slug}.html")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)
    print(f"[OK] Player pages saved: {len(profiles)} player(s).")


_VOIDSPIRE_BOSS_ORDER = [
    "Chimaerus, the Undreamt God",
    "Imperator Averzian",
    "Vorasius",
    "Fallen-King Salhadaar",
    "Vaelgor & Ezzorak",
    "Lightblinded Vanguard",
    "Crown of the Cosmos",
]
_BOSS_DEDICATED_PAGES = {
    "Chimaerus, the Undreamt God": "boss_chimaerus_heroic.html",
    "Crown of the Cosmos":         "boss_crown_heroic.html",
}


def write_bosses_html(days_data: list, output_path: str, guild_name: str = "") -> None:
    """Write a per-boss hub page (bosses.html) aggregating kills/wipes across all reports."""
    from html import escape
    diff_order = {"Mythic": 0, "Heroic": 1, "Normal": 2}

    # Aggregate per boss_base per difficulty
    # Structure: {boss_base: {diff: {"kills": int, "wipes": list[dict], "page": str}}}
    boss_agg: dict = {}

    for day in days_data:
        bd       = day.get("boss_data", {})
        wd       = day.get("wipe_data", {})
        day_diff = day.get("difficulty", "Normal")

        def _page(d):
            fn = _raid_filename(d)
            return fn

        for key, fights in bd.items():
            parts = key.rsplit(" (", 1)
            base  = parts[0]
            diff  = parts[1].rstrip(")") if len(parts) > 1 else day_diff
            entry = boss_agg.setdefault(base, {}).setdefault(diff, {"kills": 0, "wipes": [], "page": ""})
            entry["kills"] += len(fights)
            if not entry["page"]:
                entry["page"] = _page(day)

        for key, wipes in wd.items():
            parts = key.rsplit(" (", 1)
            base  = parts[0]
            diff  = parts[1].rstrip(")") if len(parts) > 1 else day_diff
            entry = boss_agg.setdefault(base, {}).setdefault(diff, {"kills": 0, "wipes": [], "page": ""})
            entry["wipes"].extend(wipes)
            if not entry["page"]:
                entry["page"] = _page(day)

    def _boss_card(boss_base: str) -> str:
        diffs = boss_agg.get(boss_base, {})
        if not diffs:
            # Not attempted
            return (
                f'<div class="boss-card boss-card-unstarted">'
                f'<div class="bc-name">{escape(boss_base)}</div>'
                f'<div class="bc-status bc-unstarted">Not attempted</div>'
                f'</div>'
            )

        # Pick best difficulty
        best_diff = min(diffs.keys(), key=lambda d: diff_order.get(d, 9))
        info      = diffs[best_diff]
        kills     = info["kills"]
        wipes     = info["wipes"]
        page      = _BOSS_DEDICATED_PAGES.get(boss_base, info["page"])
        diff_cls  = {"Mythic": "badge-mythic", "Heroic": "badge-heroic", "Normal": "badge-normal"}.get(best_diff, "badge-normal")

        if kills > 0:
            best_kill_ms = None
            for day in days_data:
                for key, fights in day.get("boss_data", {}).items():
                    if key.startswith(boss_base):
                        for f in fights:
                            ms = f.get("fight_dur_ms", 0)
                            if ms and (best_kill_ms is None or ms < best_kill_ms):
                                best_kill_ms = ms
            kill_time = ""
            if best_kill_ms:
                m, s   = divmod(best_kill_ms // 1000, 60)
                kill_time = f' &middot; Best: {m}:{s:02d}'
            status_html = (
                f'<div class="bc-status bc-killed">&#10003; Killed'
                f'<span class="bc-detail"> &middot; {kills} kill{"s" if kills != 1 else ""}{kill_time}</span>'
                f'</div>'
            )
        elif wipes:
            best_pct  = min(w.get("boss_pct", 100) for w in wipes)
            status_html = (
                f'<div class="bc-status bc-prog">Progressing</div>'
                f'<div class="bc-detail">{len(wipes)} wipe{"s" if len(wipes) != 1 else ""} &middot; Best: {best_pct:.1f}%</div>'
            )
        else:
            status_html = '<div class="bc-status bc-unstarted">No data</div>'

        link_open  = f'<a class="boss-card boss-card-{best_diff.lower()}" href="{escape(page)}">' if page else f'<div class="boss-card boss-card-{best_diff.lower()}">'
        link_close = '</a>' if page else '</div>'

        return (
            f'{link_open}'
            f'<div class="bc-top"><span class="badge-diff {diff_cls}">{escape(best_diff)}</span></div>'
            f'<div class="bc-name">{escape(boss_base)}</div>'
            f'{status_html}'
            f'<div class="bc-link">{"View Analysis &rarr;" if page else ""}</div>'
            f'{link_close}'
        )

    ordered   = _VOIDSPIRE_BOSS_ORDER
    extras    = [b for b in boss_agg if b not in ordered]
    all_bosses = ordered + extras
    cards_html = "\n".join(_boss_card(b) for b in all_bosses)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Boss Progression — {escape(guild_name)}</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: #1a1a2e; color: #e0e0e0; font-family: -apple-system, 'Segoe UI', sans-serif; padding: 28px 24px; }}
.page-header {{ display: flex; align-items: center; gap: 12px; margin-bottom: 4px; }}
h1 {{ color: #e0e0e0; font-size: 24px; }}
.back {{ color: #7289DA; font-size: 13px; text-decoration: none; display: inline-block; margin-bottom: 20px; }}
.back:hover {{ text-decoration: underline; }}
.subtitle {{ color: #555; font-size: 13px; margin-bottom: 28px; }}
.boss-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(220px, 1fr)); gap: 16px; }}
.boss-card {{ background: #16213e; border: 1px solid #2a2a4a; border-radius: 12px; padding: 18px; text-decoration: none; color: inherit; display: block; transition: border-color 0.15s, transform 0.15s; }}
.boss-card:hover {{ transform: translateY(-2px); }}
.boss-card-mythic  {{ border-color: #3a1a5a; }}
.boss-card-mythic:hover  {{ border-color: #c77dff; }}
.boss-card-heroic  {{ border-color: #1a2a4a; }}
.boss-card-heroic:hover  {{ border-color: #64b5f6; }}
.boss-card-normal  {{ border-color: #1a2a2a; }}
.boss-card-unstarted {{ border-color: #222; opacity: 0.5; }}
.bc-top {{ margin-bottom: 10px; }}
.badge-diff {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }}
.badge-normal  {{ background: #1a3a1a; color: #4caf50; }}
.badge-heroic  {{ background: #1a2a4a; color: #64b5f6; }}
.badge-mythic  {{ background: #2a1a3a; color: #c77dff; }}
.bc-name {{ font-size: 15px; font-weight: 700; color: #e0e0e0; margin-bottom: 10px; line-height: 1.3; }}
.bc-status {{ font-size: 13px; font-weight: 600; margin-bottom: 4px; }}
.bc-killed {{ color: #81c784; }}
.bc-prog   {{ color: #ffb74d; }}
.bc-unstarted {{ color: #555; }}
.bc-detail {{ font-size: 12px; color: #888; margin-bottom: 8px; }}
.bc-link {{ color: #7289DA; font-size: 12px; font-weight: 600; margin-top: 10px; }}
.boss-card:hover .bc-link {{ text-decoration: underline; }}
</style>
</head>
<body>
<a class="back" href="index.html">&#8592; Back to Raids</a>
<div class="page-header"><h1>&#9760; Boss Progression — {escape(guild_name)}</h1></div>
<div class="subtitle">{len(all_bosses)} bosses tracked</div>
<div class="boss-grid">
{cards_html}
</div>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Bosses page saved: {output_path}")


def write_index_html(days_data: list, output_path: str, guild_name: str = "") -> None:
    """Write the overview index.html.
    Normal splits are expanded into individual cards (numbered Run 1, Run 2, …).
    Heroic/Mythic get one card each.
    """
    diff_cls_map = {"Normal": "badge-normal", "Heroic": "badge-heroic", "Mythic": "badge-mythic"}

    # ── Build card descriptors (oldest→newest for NM numbering) ──
    sorted_asc = sorted(days_data, key=lambda d: d["report_info"].get("startTime", 0))
    nm_cards   = []   # Normal-split cards only, in chronological order
    all_cards  = []   # every card

    for day_data in sorted_asc:
        ri   = day_data["report_info"]
        rc   = day_data["report_code"]
        bd   = day_data["boss_data"]
        diff = day_data.get("difficulty", "")
        date_str = datetime.fromtimestamp(ri["startTime"] / 1000, tz=timezone.utc).strftime("%B %d, %Y") if ri.get("startTime") else ""
        title    = ri.get("title", rc)
        filename = _raid_filename(day_data)
        if diff in ("Heroic", "Mythic"):
            label = "HC" if diff == "Heroic" else "Mythic"
            all_cards.append({"diff": diff, "label": label, "date_str": date_str, "title": title,
                               "filename": filename, "boss_count": len(bd), "nm_idx": None, "day_data": day_data})
        else:
            card = {"diff": "Normal", "label": "NM", "date_str": date_str, "title": title,
                    "filename": filename, "boss_count": len(bd), "nm_idx": None, "day_data": day_data}
            all_cards.append(card)
            nm_cards.append(card)

    # Sequential NM numbering ordered by date then player_split
    total_nm = len(nm_cards)
    nm_ordered = sorted(nm_cards, key=lambda c: (
        c["day_data"]["report_info"].get("startTime", 0),
        c["day_data"].get("player_split", 0)
    ))
    for i, c in enumerate(nm_ordered):
        c["nm_idx"] = i + 1
        c["label"]  = f"NM · Run {i + 1}"

    # Display order: newest first; within same date Heroic > Normal; higher split before lower
    diff_order = {"Mythic": 0, "Heroic": 1, "Normal": 2}
    all_cards.sort(key=lambda c: (-c["day_data"]["report_info"].get("startTime", 0),
                                   diff_order.get(c["diff"], 3),
                                   -c["day_data"].get("player_split", 0)))

    def _card(c: dict, is_first: bool) -> str:
        diff_cls     = diff_cls_map.get(c["diff"], "badge-normal")
        recent_badge = '<span class="badge-recent">Most Recent</span>' if is_first else ""
        sub = f'<span>Run {c["nm_idx"]} of {total_nm}</span>' if c["nm_idx"] else ""
        return (
            f'<a class="raid-card" href="{c["filename"]}">'
            f'  <div class="card-top">'
            f'    <div class="card-date">{escape(c["date_str"])}</div>'
            f'    <div class="card-badges"><span class="badge-diff {diff_cls}">{escape(c["label"])}</span>{recent_badge}</div>'
            f'  </div>'
            f'  <div class="card-title">{escape(c["title"])}</div>'
            f'  <div class="card-stats">'
            f'    <span>&#9876; {c["boss_count"]} boss{"es" if c["boss_count"] != 1 else ""} killed</span>'
            f'    {sub}'
            f'  </div>'
            f'  <div class="card-link">View Report &#8594;</div>'
            f'</a>'
        )

    mythic_cards = [c for c in all_cards if c["diff"] == "Mythic"]
    legacy_cards = [c for c in all_cards if c["diff"] != "Mythic"]
    total_cards  = len(all_cards)
    gear_link    = '<a href="gear_normal.html" class="gear-link" title="Crafted Gear — All Normal Runs">⚙</a>' if nm_cards else ""

    # Build Mythic section (prominent, always open)
    if mythic_cards:
        mythic_html = (
            '<div class="section-header section-mythic">&#9876; Mythic Progression</div>'
            '<div class="raid-grid mythic-grid">'
            + "\n".join(_card(c, i == 0) for i, c in enumerate(mythic_cards))
            + '</div>'
        )
    else:
        mythic_html = ""

    # Build Normal/Heroic archive section (collapsed by default if Mythic exists)
    if legacy_cards:
        legacy_inner = (
            '<div class="raid-grid">'
            + "\n".join(_card(c, i == 0 and not mythic_cards) for i, c in enumerate(legacy_cards))
            + '</div>'
        )
        if mythic_cards:
            legacy_html = (
                f'<details class="legacy-section">'
                f'<summary class="section-header section-legacy">Normal &amp; Heroic Archive'
                f'<span class="legacy-count">{len(legacy_cards)} run{"s" if len(legacy_cards) != 1 else ""}</span>'
                f'</summary>'
                f'{legacy_inner}'
                f'</details>'
            )
        else:
            # No Mythic yet — show all normally with a neutral header
            legacy_html = (
                '<div class="section-header section-legacy">Normal &amp; Heroic</div>'
                + legacy_inner
            )
    else:
        legacy_html = ""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Raid Audit — {escape(guild_name)}</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: #1a1a2e; color: #e0e0e0; font-family: -apple-system, 'Segoe UI', sans-serif; padding: 28px 24px; }}
.page-header {{ display: flex; align-items: center; gap: 12px; margin-bottom: 4px; }}
h1 {{ color: #7289DA; font-size: 24px; }}
.gear-link {{ color: #4caf50; font-size: 20px; text-decoration: none; line-height: 1; }}
.gear-link:hover {{ color: #81c784; }}
.roster-link {{ color: #7289DA; font-size: 20px; text-decoration: none; line-height: 1; }}
.boss-link {{ color: #d29922; font-size: 18px; text-decoration: none; line-height: 1; font-weight: 700; }}
.boss-link:hover {{ color: #e5cc80; }}
.subtitle {{ color: #555; font-size: 13px; margin-bottom: 28px; }}
.section-header {{ font-size: 13px; font-weight: 700; letter-spacing: 0.08em; text-transform: uppercase; margin-bottom: 14px; padding-bottom: 8px; border-bottom: 1px solid #2a2a4a; }}
.section-mythic {{ color: #c77dff; border-color: #4a1a6a; margin-bottom: 14px; }}
.section-legacy {{ color: #666; border-color: #2a2a4a; cursor: pointer; display: flex; align-items: center; gap: 10px; list-style: none; margin-bottom: 14px; }}
.section-legacy::before {{ content: "▶"; font-size: 10px; color: #555; transition: transform 0.15s; }}
details[open] > .section-legacy::before {{ content: "▼"; }}
.section-legacy::-webkit-details-marker {{ display: none; }}
.legacy-count {{ color: #555; font-size: 11px; font-weight: 400; letter-spacing: 0; text-transform: none; }}
.legacy-section {{ margin-top: 36px; }}
.raid-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(240px, 1fr)); gap: 18px; margin-bottom: 36px; }}
.mythic-grid .raid-card {{ border-color: #3a1a5a; }}
.mythic-grid .raid-card:hover {{ border-color: #c77dff; }}
.raid-card {{ background: #16213e; border: 1px solid #2a2a4a; border-radius: 12px; padding: 20px; text-decoration: none; color: inherit; display: block; transition: border-color 0.15s, transform 0.15s; }}
.raid-card:hover {{ border-color: #7289DA; transform: translateY(-2px); }}
.card-top {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 10px; }}
.card-date {{ color: #7289DA; font-size: 12px; font-weight: 600; }}
.card-badges {{ display: flex; gap: 6px; flex-wrap: wrap; justify-content: flex-end; }}
.badge-diff {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }}
.badge-normal  {{ background: #1a3a1a; color: #4caf50; }}
.badge-heroic  {{ background: #1a2a4a; color: #64b5f6; }}
.badge-mythic  {{ background: #2a1a3a; color: #c77dff; }}
.badge-recent  {{ background: #2a2416; color: #E5CC80; display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 700; }}
.card-title {{ color: #e0e0e0; font-size: 16px; font-weight: 700; margin-bottom: 12px; }}
.card-stats {{ display: flex; gap: 14px; color: #888; font-size: 12px; margin-bottom: 14px; flex-wrap: wrap; }}
.card-link {{ color: #7289DA; font-size: 13px; font-weight: 600; }}
.raid-card:hover .card-link {{ text-decoration: underline; }}
</style>
</head>
<body>
<div class="page-header"><h1>Raid Audit — {escape(guild_name)}</h1>{gear_link}<a href="roster.html" class="roster-link" title="Guild Roster">&#128101;</a><a href="bosses.html" class="boss-link" title="Boss Progression">&#9760;</a></div>
<div class="subtitle">{total_cards} entr{"ies" if total_cards != 1 else "y"} tracked</div>
{mythic_html}
{legacy_html}
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Index page saved: {output_path}")


def write_roster_html(days_data: list, output_path: str, guild_name: str = "") -> None:
    """Write a roster overview page listing all players grouped by role."""
    from html import escape as _esc

    profiles = build_player_profiles(days_data)

    # Build per-character class lookup from all actors across all reports
    char_classes: dict = {}  # char_name_lower -> class_name
    for day_data in days_data:
        for actor in day_data.get("actors", []):
            name = actor.get("name", "")
            cls  = actor.get("subType", "Unknown")
            if name and cls != "Unknown":
                char_classes[name.lower()] = cls

    # Build full player list: merge roster.txt with seen-in-reports data
    # PLAYER_ROLES has all players from roster.txt
    all_players = {}
    for pname, role in PLAYER_ROLES.items():
        prof = profiles.get(pname, {})
        cls = prof.get("class", "Unknown")
        cls_color = CLASS_COLORS.get(cls, "#888")

        # Chars from roster.txt for this player
        roster_chars = {}  # char_name -> "Main"/"Alt"
        for char_lower, (rname, mtype) in ROSTER.items():
            if rname == pname:
                # Find the properly-cased name from profiles or use as-is
                char_display = next(
                    (c for c in prof.get("chars", set()) if c.lower() == char_lower),
                    char_lower.capitalize()
                )
                roster_chars[char_display] = mtype

        # Chars seen in reports but not in roster.txt
        extra_chars = prof.get("chars", set()) - {c for c in prof.get("chars", set())
                                                   if c.lower() in {k for k, (n,_) in ROSTER.items() if n == pname}}

        all_players[pname] = {
            "role": role,
            "class": cls,
            "cls_color": cls_color,
            "roster_chars": roster_chars,
            "extra_chars": extra_chars,
            "has_profile": pname in profiles,
        }

    # Group by role
    role_order = {"Tank": 0, "Healer": 1, "DPS": 2}
    grouped = {"Tank": [], "Healer": [], "DPS": []}
    for pname, data in sorted(all_players.items(), key=lambda x: (role_order.get(x[1]["role"], 3), x[0].lower())):
        grouped.get(data["role"], grouped["DPS"]).append((pname, data))

    role_colors = {"Tank": "#64b5f6", "Healer": "#81c784", "DPS": "#e57373"}
    role_icons  = {"Tank": "🛡", "Healer": "💚", "DPS": "⚔"}

    cards_html = ""
    for role in ("Tank", "Healer", "DPS"):
        players = grouped[role]
        if not players:
            continue
        rc = role_colors[role]
        ri = role_icons[role]
        role_label = role if role == "DPS" else f"{role}s"
        cards_html += f'<div class="role-group"><div class="role-group-title" style="color:{rc}">{ri} {role_label}</div><div class="player-cards">'
        for pname, data in players:
            slug = _player_slug(pname)
            cc = data["cls_color"]

            # Build char pills — each char gets its own class color
            chars_html = ""
            for cname, mtype in sorted(data["roster_chars"].items(), key=lambda x: (0 if x[1]=="Main" else 1, x[0])):
                char_cls   = char_classes.get(cname.lower(), "Unknown")
                char_color = CLASS_COLORS.get(char_cls, "#888")
                tag_bg     = "#1e3a1e" if mtype == "Main" else "#1a2233"
                tag_color  = "#81c784" if mtype == "Main" else "#7289DA"
                tag_lbl    = "M" if mtype == "Main" else "A"
                chars_html += (f'<span class="char-pill">'
                               f'<span class="cname-text" style="color:{char_color}">{_esc(cname)}</span>'
                               f'<span class="char-type" style="background:{tag_bg};color:{tag_color}">{tag_lbl}</span>'
                               f'</span>')
            for cname in sorted(data["extra_chars"]):
                char_cls   = char_classes.get(cname.lower(), "Unknown")
                char_color = CLASS_COLORS.get(char_cls, "#888")
                chars_html += (f'<span class="char-pill">'
                               f'<span class="cname-text" style="color:{char_color}">{_esc(cname)}</span>'
                               f'<span class="char-type" style="background:#2a1a1a;color:#aaa">?</span>'
                               f'</span>')

            seen_badge = ""
            if not data["has_profile"]:
                seen_badge = '<span class="unseen-badge">no logs</span>'

            cards_html += (
                f'<div class="player-card">'
                f'<div class="card-top">'
                f'<a href="players/player_{slug}.html" class="card-name" style="color:{cc}">{_esc(pname)}</a>'
                f'{seen_badge}'
                f'</div>'
                f'<div class="card-chars">{chars_html}</div>'
                f'</div>'
            )
        cards_html += '</div></div>'

    title = f"Roster — {guild_name}" if guild_name else "Roster"
    total_players = len(all_players)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_esc(title)}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0d1117;color:#c9d1d9;font-family:'Segoe UI',sans-serif;font-size:13px;padding:24px 28px}}
a{{color:#7289DA;text-decoration:none}} a:hover{{text-decoration:underline}}
.page-header{{display:flex;align-items:center;gap:12px;margin-bottom:6px}}
h1{{color:#7289DA;font-size:22px;font-weight:700}}
.subtitle{{color:#555;font-size:12px;margin-bottom:28px}}
.role-group{{margin-bottom:32px}}
.role-group-title{{font-size:13px;font-weight:700;letter-spacing:1px;text-transform:uppercase;margin-bottom:12px;padding-bottom:6px;border-bottom:1px solid #1e2a3a}}
.player-cards{{display:flex;flex-wrap:wrap;gap:10px}}
.player-card{{background:#0d1525;border:1px solid #1e2a4a;border-radius:8px;padding:12px 14px;min-width:200px;max-width:260px;flex:1}}
.player-card:hover{{border-color:#2a3a6a;background:#111827}}
.card-top{{display:flex;align-items:baseline;gap:8px;margin-bottom:8px;flex-wrap:wrap}}
.card-name{{font-size:15px;font-weight:700}}
.card-class{{font-size:11px;opacity:0.7}}
.unseen-badge{{font-size:10px;color:#555;background:#111;border:1px solid #222;border-radius:10px;padding:1px 6px;margin-left:auto}}
.card-chars{{display:flex;flex-wrap:wrap;gap:5px}}
.char-pill{{display:inline-flex;align-items:center;gap:3px;background:#111827;border-radius:12px;padding:2px 8px 2px 6px;font-size:11px}}
.cname-text{{font-weight:600}}
.char-type{{font-size:9px;font-weight:700;border-radius:8px;padding:1px 4px}}
.back{{margin-bottom:20px;display:inline-block;color:#7289DA;font-size:13px}}
</style>
</head>
<body>
<a class="back" href="index.html">← Back to Raids</a>
<div class="page-header"><h1>👥 {_esc(title)}</h1></div>
<div class="subtitle">{total_players} players</div>
{cards_html}
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Roster page saved: {output_path}")


def write_boss_progression_html(days_data: list, output_path: str, guild_name: str = "") -> None:
    """Write Chimaerus Heroic boss progression page: aggregated Alndust + Horror wave stats across all kills."""
    from html import escape as _esc

    TARGET_BOSS = "Chimaerus, the Undreamt God"
    TARGET_DIFF = "Heroic"
    WCL_BASE    = "https://www.warcraftlogs.com/reports"

    # ── Step 1: collect all Chimaerus Heroic kills, normalize PIDs → names ──
    pulls = []
    for day_data in sorted(days_data, key=lambda d: d["report_info"].get("startTime", 0)):
        report_code  = day_data["report_code"]
        actor_lookup = {a["id"]: a for a in day_data.get("actors", [])}

        def pid_name(pid, al=actor_lookup):
            char = al.get(pid, {}).get("name", f"#{pid}")
            player, _ = lookup_roster(char)
            return player

        boss_data = day_data.get("boss_data", {})
        for bname, fights in boss_data.items():
            if TARGET_BOSS not in bname or TARGET_DIFF not in bname:
                continue
            for fight in fights:
                # Normalize alndust waves
                alndust = []
                for wave in fight.get("alndust_groups", []):
                    alndust.append({
                        "wave":       wave["wave"],
                        "down":       [pid_name(p) for p in wave.get("down_pids",      [])],
                        "up":         [pid_name(p) for p in wave.get("up_pids",        [])],
                        "missed":     [pid_name(p) for p in wave.get("missed_pids",    [])],
                        "wrong":      [pid_name(p) for p in wave.get("wrong_pids",     [])],
                        "double_up":  [pid_name(p) for p in wave.get("double_up_pids", [])],
                    })
                # Normalize horror waves
                horror = []
                for wave in fight.get("horror_waves", []):
                    pp_named = {}
                    for pid, stats in wave.get("per_player", {}).items():
                        pp_named[pid_name(int(pid))] = stats
                    horror.append({
                        "wave":         wave["wave"],
                        "wave_dur_ms":  wave.get("wave_dur_ms", 0),
                        "kill_time_s":  wave.get("kill_time_s"),
                        "per_player":   pp_named,
                    })
                dur_s = (fight.get("fight_dur_ms") or 0) // 1000
                pulls.append({
                    "report_code": report_code,
                    "fight_id":    fight["fight_id"],
                    "dur_s":       dur_s,
                    "dur_fmt":     f"{dur_s//60}:{dur_s%60:02d}",
                    "alndust":     alndust,
                    "horror":      horror,
                    "actor_lookup": actor_lookup,
                })

    if not pulls:
        return

    num_pulls = len(pulls)

    def pull_label(pi):
        p = pulls[pi]
        return f"Kill {pi+1} ({p['dur_fmt']})"

    def wcl_link(pi):
        p = pulls[pi]
        return f"{WCL_BASE}/{p['report_code']}?fight={p['fight_id']}"

    # ── Step 2: Aggregate Alndust per player ──
    # Collect all players who participated in any alndust wave
    alndust_players: set = set()
    player_alndust: dict = {}  # name → {wrong: [(pi, wave)], missed: [...], double_up: [...]}

    for pi, pull in enumerate(pulls):
        for wave in pull["alndust"]:
            wn = wave["wave"]
            for name in wave["down"] + wave["up"] + wave["missed"] + wave["wrong"] + wave["double_up"]:
                alndust_players.add(name)
            for key, names in [("wrong", wave["wrong"]), ("missed", wave["missed"]), ("double_up", wave["double_up"])]:
                for name in names:
                    rec = player_alndust.setdefault(name, {"wrong": [], "missed": [], "double_up": []})
                    rec[key].append((pi, wn))

    # ── Step 3: Aggregate Horror wave stats per player ──
    horror_players: set = set()
    player_horror: dict = {}  # name → {pulls: [{wave, dmg, dps, active_pct}]}

    for pi, pull in enumerate(pulls):
        for wave in pull["horror"]:
            dur  = wave["wave_dur_ms"]
            for name, stats in wave["per_player"].items():
                horror_players.add(name)
                dmg    = stats.get("dmg", 0)
                active = stats.get("active_ms", 0)
                dps    = int(dmg / (dur / 1000)) if dur else 0
                pct    = round(active / dur * 100, 1) if dur and active else 0.0
                player_horror.setdefault(name, []).append({
                    "pull": pi, "wave": wave["wave"], "dmg": dmg, "dps": dps, "active_pct": pct
                })

    def fmt_dmg(v):
        if v >= 1_000_000: return f"{v/1_000_000:.2f}M"
        if v >= 1_000:     return f"{v/1_000:.1f}k"
        return str(v) if v else "—"

    # ── Step 4: Build HTML ──
    # Pull timeline bar
    timeline_cells = ""
    for pi, p in enumerate(pulls):
        timeline_cells += (f'<a href="{wcl_link(pi)}" target="_blank" class="tl-cell tl-kill">'
                           f'Kill {pi+1}<br><span class="tl-sub">{p["dur_fmt"]}</span></a>')

    # ── Alndust table ──
    def tip(items, pi_wn_list):
        """Render a count cell with hover tooltip listing specific pulls/waves."""
        if not pi_wn_list:
            return '<td class="prog-cell prog-ok" data-val="0">—</td>'
        count = len(pi_wn_list)
        lines = ""
        for (pi, wn) in pi_wn_list:
            lines += (f'<div class="tip-row">{_esc(pull_label(pi))} · Wave {wn} '
                      f'<a href="{wcl_link(pi)}" target="_blank" class="tip-wcl">WCL ↗</a></div>')
        badge_cls = "prog-bad" if count >= 2 else "prog-warn"
        return (f'<td class="prog-cell {badge_cls}" data-val="{count}">'
                f'<div class="tip-wrap">{count}'
                f'<div class="tip-box">{lines}</div></div></td>')

    def player_color(player_name):
        """Get class color by matching any char belonging to this player across all pulls."""
        for pull in reversed(pulls):
            for actor in pull["actor_lookup"].values():
                char = actor.get("name", "")
                if lookup_roster(char)[0] == player_name:
                    return CLASS_COLORS.get(actor.get("subType", ""), "#ccc")
        return "#ccc"

    alndust_rows = ""
    sorted_alndust = sorted(alndust_players, key=lambda n: (
        -(len(player_alndust.get(n, {}).get("wrong",      [])) +
          len(player_alndust.get(n, {}).get("missed",     [])) +
          len(player_alndust.get(n, {}).get("double_up",  []))),
        n.lower()
    ))
    for name in sorted_alndust:
        rec      = player_alndust.get(name, {"wrong": [], "missed": [], "double_up": []})
        total    = len(rec["wrong"]) + len(rec["missed"]) + len(rec["double_up"])
        row_cls  = "prog-row-clean" if total == 0 else ""
        color = player_color(name)
        alndust_rows += (f'<tr class="{row_cls}">'
                         f'<td class="prog-name" style="color:{color}">{_esc(name)}</td>'
                         + tip("wrong",     rec["wrong"])
                         + tip("missed",    rec["missed"])
                         + tip("double_up", rec["double_up"])
                         + f'</tr>')

    # ── Horror table ──
    horror_rows = ""
    # Compute per-player averages
    horror_stats = {}
    for name in horror_players:
        entries     = player_horror.get(name, [])
        avg_dmg     = int(sum(e["dmg"]        for e in entries) / len(entries)) if entries else 0
        avg_dps     = int(sum(e["dps"]        for e in entries) / len(entries)) if entries else 0
        avg_active  = round(sum(e["active_pct"] for e in entries) / len(entries), 1) if entries else 0.0
        low_pulls   = [(e["pull"], e["wave"]) for e in entries if e["dmg"] < 200_000 and e["active_pct"] < 5]
        horror_stats[name] = {"avg_dmg": avg_dmg, "avg_dps": avg_dps, "avg_active": avg_active,
                               "low_pulls": low_pulls, "entries": entries}

    sorted_horror = sorted(horror_players, key=lambda n: -horror_stats[n]["avg_dmg"])
    for name in sorted_horror:
        s     = horror_stats[name]
        color = player_color(name)
        low_cell = ""
        if s["low_pulls"]:
            lines = ""
            for (pi, wn) in s["low_pulls"]:
                lines += (f'<div class="tip-row">{_esc(pull_label(pi))} · Wave {wn} '
                          f'<a href="{wcl_link(pi)}" target="_blank" class="tip-wcl">WCL ↗</a></div>')
            low_cell = (f'<td class="prog-cell prog-warn">'
                        f'<div class="tip-wrap">{len(s["low_pulls"])}'
                        f'<div class="tip-box">{lines}</div></div></td>')
        else:
            low_cell = f'<td class="prog-cell prog-ok" data-val="0">—</td>'

        horror_rows += (f'<tr>'
                        f'<td class="prog-name" style="color:{color}">{_esc(name)}</td>'
                        f'<td class="prog-cell" data-val="{s["avg_dmg"]}">{fmt_dmg(s["avg_dmg"])}</td>'
                        f'<td class="prog-cell" data-val="{s["avg_dps"]}">{fmt_dmg(s["avg_dps"])}</td>'
                        f'<td class="prog-cell" data-val="{s["avg_active"]}">{s["avg_active"]:.1f}%</td>'
                        f'{low_cell}'
                        f'</tr>')

    title = f"Boss Progression — {TARGET_BOSS} ({TARGET_DIFF})"
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_esc(title)}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0d1117;color:#c9d1d9;font-family:'Segoe UI',sans-serif;font-size:13px;padding:24px 28px}}
a{{color:#7289DA;text-decoration:none}} a:hover{{text-decoration:underline}}
.back{{display:inline-block;margin-bottom:20px;color:#7289DA;font-size:13px}}
h1{{color:#e2e8f0;font-size:20px;font-weight:700;margin-bottom:4px}}
.subtitle{{color:#555;font-size:12px;margin-bottom:24px}}
.section-title{{font-size:13px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:#7289DA;
  margin:28px 0 12px;padding-bottom:6px;border-bottom:1px solid #1e2a3a}}

/* Pull timeline */
.timeline{{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:8px}}
.tl-cell{{display:flex;flex-direction:column;align-items:center;justify-content:center;
  min-width:90px;padding:8px 12px;border-radius:8px;font-size:12px;font-weight:600;
  text-align:center;text-decoration:none;border:1px solid}}
.tl-kill{{background:#0d2010;border-color:#2a5c2a;color:#81c784}}
.tl-kill:hover{{background:#122816;text-decoration:none}}
.tl-sub{{font-size:10px;font-weight:400;opacity:0.7;margin-top:2px}}

/* Tables */
.prog-wrap{{overflow-x:auto;margin-bottom:8px}}
table.prog-tbl{{border-collapse:collapse;width:100%}}
table.prog-tbl th{{background:#0d1525;color:#7289DA;font-size:11px;font-weight:700;
  text-transform:uppercase;letter-spacing:0.5px;padding:8px 12px;text-align:left;
  border-bottom:2px solid #1e2a4a;border-right:1px solid #1a2640;white-space:nowrap;
  cursor:pointer;user-select:none}}
table.prog-tbl th:last-child{{border-right:none}}
table.prog-tbl th:hover{{background:#111e33;color:#a0b4e0}}
table.prog-tbl td{{padding:6px 12px;border-bottom:1px solid #111827;
  border-right:1px solid #151e2e;white-space:nowrap}}
table.prog-tbl td:last-child{{border-right:none}}
.prog-name{{font-weight:600;font-size:13px;min-width:140px}}
.prog-cell{{text-align:center;min-width:80px;font-size:12px;color:#c9d1d9}}
.prog-ok{{color:#4a5568}}
.prog-warn{{color:#f4a742;font-weight:600}}
.prog-bad{{color:#e57373;font-weight:700}}
.prog-row-clean td{{opacity:0.5}}
.sort-arrow{{font-size:9px;margin-left:4px;opacity:0.5}}

/* Tooltip */
.tip-wrap{{position:relative;display:inline-block;cursor:default}}
.tip-box{{display:none;position:absolute;z-index:100;bottom:calc(100% + 6px);left:50%;
  transform:translateX(-50%);background:#1a2233;border:1px solid #2a3a5a;border-radius:8px;
  padding:10px 12px;min-width:220px;max-width:320px;box-shadow:0 4px 20px rgba(0,0,0,0.6);
  text-align:left}}
.tip-wrap:hover .tip-box{{display:block}}
.tip-row{{font-size:11px;color:#c9d1d9;padding:2px 0;white-space:nowrap}}
.tip-wcl{{color:#7289DA;font-size:10px;margin-left:6px}}
</style>
</head>
<body>
<a class="back" href="index.html">← Back to Raids</a>
<h1>⚔ {_esc(TARGET_BOSS)} — {TARGET_DIFF} Progression</h1>
<div class="subtitle">{num_pulls} kill(s) tracked · Data from all reports · Hover counts for details</div>

<div class="section-title">Pull Timeline</div>
<div class="timeline">{timeline_cells}</div>

<div class="section-title">Alndust — Mechanic Failures (across all pulls)</div>
<div class="prog-wrap">
<table class="prog-tbl" id="tbl-alndust">
<thead><tr>
  <th onclick="progSort('tbl-alndust',0,this)">Player <span class="sort-arrow">▼</span></th>
  <th onclick="progSort('tbl-alndust',1,this)" title="Went to wrong group">Wrong Group <span class="sort-arrow">▼</span></th>
  <th onclick="progSort('tbl-alndust',2,this)" title="Missed going underground entirely">Missed <span class="sort-arrow">▼</span></th>
  <th onclick="progSort('tbl-alndust',3,this)" title="Went underground 2+ waves in a row">Double-up <span class="sort-arrow">▼</span></th>
</tr></thead>
<tbody>{alndust_rows}</tbody>
</table>
</div>

<div class="section-title">Colossal Horror — Wave Contribution (avg across all pulls)</div>
<div class="prog-wrap">
<table class="prog-tbl" id="tbl-horror">
<thead><tr>
  <th onclick="progSort('tbl-horror',0,this)">Player <span class="sort-arrow">▼</span></th>
  <th onclick="progSort('tbl-horror',1,this)">Avg DMG <span class="sort-arrow">▼</span></th>
  <th onclick="progSort('tbl-horror',2,this)">Avg DPS <span class="sort-arrow">▼</span></th>
  <th onclick="progSort('tbl-horror',3,this)">Avg Active <span class="sort-arrow">▼</span></th>
  <th onclick="progSort('tbl-horror',4,this)" title="Pulls where DMG &lt;200k and Active &lt;5%">Low Contrib <span class="sort-arrow">▼</span></th>
</tr></thead>
<tbody>{horror_rows}</tbody>
</table>
</div>
<script>
(function(){{
  var asc = {{}};
  window.progSort = function(tblId, colIdx, th) {{
    var key = tblId + '_' + colIdx;
    asc[key] = !asc[key];
    var tbody = document.getElementById(tblId).querySelector('tbody');
    var rows  = Array.from(tbody.querySelectorAll('tr'));
    rows.sort(function(a, b) {{
      var ac = a.cells[colIdx], bc = b.cells[colIdx];
      var av = ac ? (ac.getAttribute('data-val') || ac.innerText.trim()) : '';
      var bv = bc ? (bc.getAttribute('data-val') || bc.innerText.trim()) : '';
      var an = parseFloat(av.replace(/[^0-9.\-]/g,'')), bn = parseFloat(bv.replace(/[^0-9.\-]/g,''));
      if (!isNaN(an) && !isNaN(bn)) return asc[key] ? an - bn : bn - an;
      return asc[key] ? av.localeCompare(bv) : bv.localeCompare(av);
    }});
    rows.forEach(function(r) {{ tbody.appendChild(r); }});
    document.querySelectorAll('#' + tblId + ' .sort-arrow').forEach(function(s) {{ s.textContent = '▼'; s.style.opacity='0.5'; }});
    var arrow = th.querySelector('.sort-arrow');
    if (arrow) {{ arrow.textContent = asc[key] ? '▲' : '▼'; arrow.style.opacity='1'; }}
  }};
}})();
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Boss progression page saved: {output_path}")


def write_crown_progression_html(days_data: list, output_path: str, guild_name: str = "") -> None:
    """Write Crown of the Cosmos Heroic boss progression page.

    Shows per-wipe Silverstrike intermission tables (arrow order, multi-hits)
    and P3 circle leave-order tables. Aggregates multi-hit counts per player
    across all wipes.
    """
    from html import escape as _esc

    TARGET_BOSS = "Crown of the Cosmos"
    TARGET_DIFF = "Heroic"
    WCL_BASE    = "https://www.warcraftlogs.com/reports"

    # ── collect all Crown wipes ──────────────────────────────────────────────
    wipes = []
    for day_data in sorted(days_data, key=lambda d: d["report_info"].get("startTime", 0)):
        report_code  = day_data["report_code"]
        actor_lookup = {a["id"]: a for a in day_data.get("actors", [])}

        def normalize_name(raw_name):
            # raw_name may be a character name — resolve to player name
            player, _ = lookup_roster(raw_name)
            return player

        def actor_color(player_name, al=actor_lookup):
            for a in al.values():
                if lookup_roster(a.get("name", ""))[0] == player_name:
                    return CLASS_COLORS.get(a.get("subType", ""), "#ccc")
            return "#ccc"

        wd = day_data.get("wipe_data", {})
        for bname, bwipes in wd.items():
            if TARGET_BOSS not in bname or TARGET_DIFF not in bname:
                continue
            for w in bwipes:
                cm = w.get("crown_mechanics", {})
                if not cm:
                    continue
                dur_s = (w.get("fight_dur_ms") or 0) // 1000

                # Normalise arrow names to player names
                silverstrike = []
                for im in cm.get("silverstrike", []):
                    arrows = []
                    for a in im["arrows"]:
                        arrows.append({**a, "name": normalize_name(a["name"])})
                    multi_named = {}
                    for pid_str, cnt in im.get("multi_hit", {}).items():
                        # Find any arrow entry with this pid
                        pid = int(pid_str)
                        for a in im["arrows"]:
                            if a["pid"] == pid:
                                multi_named[normalize_name(a["name"])] = cnt
                                break
                    silverstrike.append({
                        "intermission": im["intermission"],
                        "arrows":       arrows,
                        "multi_hit":    multi_named,
                    })

                p3_circles = []
                for s in cm.get("p3_circles", []):
                    players = [{**p, "name": normalize_name(p["name"])} for p in s["players"]]
                    p3_circles.append({"players": players, "flags": s["flags"]})

                wipes.append({
                    "report_code": report_code,
                    "fight_id":    w["fight_id"],
                    "wipe_num":    w["wipe_num"],
                    "dur_s":       dur_s,
                    "dur_fmt":     f"{dur_s//60}:{dur_s%60:02d}",
                    "boss_pct":    w.get("boss_pct", 0),
                    "silverstrike": silverstrike,
                    "p3_circles":  p3_circles,
                    "actor_color": actor_color,
                })

    if not wipes:
        return

    # ── aggregate multi-hit counts per player across all wipes ───────────────
    multi_totals: dict = {}  # player_name → total multi-hit count
    for w in wipes:
        for im in w["silverstrike"]:
            for name, cnt in im["multi_hit"].items():
                multi_totals[name] = multi_totals.get(name, 0) + (cnt - 1)  # extra hits only

    # All players who appeared in any arrow
    all_arrow_players: set = set()
    for w in wipes:
        for im in w["silverstrike"]:
            for a in im["arrows"]:
                all_arrow_players.add(a["name"])

    # ── per-player color lookup (first actor_color found across wipes) ────────
    def player_color(player_name):
        for w in wipes:
            c = w["actor_color"](player_name)
            if c != "#ccc":
                return c
        return "#ccc"

    # ── Build summary table rows (sorted by most multi-hits) ─────────────────
    sorted_players = sorted(all_arrow_players,
                            key=lambda n: (-multi_totals.get(n, 0), n.lower()))
    summary_rows = ""
    for name in sorted_players:
        extra = multi_totals.get(name, 0)
        color = player_color(name)
        cls   = "prog-warn" if extra > 2 else ("prog-caution" if extra > 0 else "prog-ok")
        summary_rows += (f'<tr>'
                         f'<td class="prog-name" style="color:{color}">{_esc(name)}</td>'
                         f'<td class="prog-cell {cls}" data-val="{extra}">'
                         f'{"—" if extra == 0 else extra}</td>'
                         f'</tr>')

    # ── Build per-wipe cards ──────────────────────────────────────────────────
    wipe_cards = ""
    for w in wipes:
        wcl_url = f"{WCL_BASE}/{w['report_code']}?fight={w['fight_id']}"
        pct_lbl = f"{w['boss_pct']:.1f}% boss HP"
        header  = (f'<div class="wipe-header">'
                   f'<span class="wipe-title">Wipe {w["wipe_num"]} &nbsp;·&nbsp; {w["dur_fmt"]} &nbsp;·&nbsp; {_esc(pct_lbl)}</span>'
                   f'<a href="{wcl_url}" target="_blank" class="wipe-wcl">WCL ↗</a>'
                   f'</div>')

        im_tables = ""
        for im in w["silverstrike"]:
            # Group arrows by time clusters (same t_s = same round)
            rounds: dict = {}  # t_s → list of arrow entries
            for a in im["arrows"]:
                rounds.setdefault(a["t_s"], []).append(a)

            round_rows = ""
            round_num = 0
            for t_s in sorted(rounds.keys()):
                round_num += 1
                for i, a in enumerate(rounds[t_s]):
                    is_multi = im["multi_hit"].get(a["name"], 0) > 1
                    multi_cls = " arrow-multi" if is_multi else ""
                    seq_lbl   = f"#{a['seq']}" if a["seq"] > 1 else ""
                    ricochet  = i > 0  # first hit at this timestamp = primary; rest = ricochet
                    rico_cls  = " arrow-rico" if ricochet else ""
                    round_rows += (f'<tr class="arrow-row{multi_cls}{rico_cls}">'
                                   f'<td class="arrow-t">{t_s}s</td>'
                                   f'<td class="arrow-round">{"" if ricochet else f"↓{round_num}"}</td>'
                                   f'<td class="arrow-name" style="color:{player_color(a["name"])}">'
                                   f'{_esc(a["name"])}{f" <span class=arrow-seq>{seq_lbl}</span>" if seq_lbl else ""}'
                                   f'</td>'
                                   f'<td class="arrow-role">{_esc(a["role"])}</td>'
                                   f'</tr>')

            multi_names = ", ".join(f'<span style="color:{player_color(n)}">{_esc(n)}</span> ×{c}'
                                    for n, c in im["multi_hit"].items() if c > 1)
            multi_warning = (f'<div class="multi-warn">⚠ Multi-hit: {multi_names}</div>'
                             if multi_names else "")

            im_tables += (f'<div class="im-block">'
                          f'<div class="im-title">Intermission {im["intermission"]}</div>'
                          f'<table class="arrow-table"><thead>'
                          f'<tr><th>Time</th><th>Arrow</th><th>Player</th><th>Role</th></tr>'
                          f'</thead><tbody>{round_rows}</tbody></table>'
                          f'{multi_warning}</div>')

        # P3 circles (if any)
        p3_html = ""
        for ci, circle_set in enumerate(w["p3_circles"]):
            circle_rows = ""
            for i, p in enumerate(circle_set["players"]):
                order_cls = ""
                if i == 0 and p["role"] in ("Tank", "Melee"):
                    order_cls = " circle-bad"
                elif i == len(circle_set["players"]) - 1 and p["role"] not in ("Tank",):
                    order_cls = " circle-bad"
                hold_s = f'{p.get("hold_ms",0)//1000}s' if p.get("hold_ms") else "?"
                circle_rows += (f'<tr class="circle-row{order_cls}">'
                                 f'<td class="circle-pos">{i+1}</td>'
                                 f'<td class="circle-name" style="color:{player_color(p["name"])}">{_esc(p["name"])}</td>'
                                 f'<td class="circle-role">{_esc(p["role"])}</td>'
                                 f'<td class="circle-hold">{hold_s}</td>'
                                 f'</tr>')
            flag_html = "".join(f'<div class="circle-flag">⚠ {_esc(fl)}</div>' for fl in circle_set["flags"])
            p3_html += (f'<div class="im-block">'
                        f'<div class="im-title">P3 Circles — Set {ci+1}</div>'
                        f'<table class="arrow-table"><thead>'
                        f'<tr><th>Leave #</th><th>Player</th><th>Role</th><th>Held</th></tr>'
                        f'</thead><tbody>{circle_rows}</tbody></table>'
                        f'{flag_html}</div>')

        if not im_tables and not p3_html:
            im_tables = '<div class="no-data">No mechanic data for this wipe.</div>'

        wipe_cards += (f'<div class="wipe-card">'
                       f'{header}'
                       f'<div class="wipe-body">{im_tables}{p3_html}</div>'
                       f'</div>')

    title = f"Boss Progression — {TARGET_BOSS} ({TARGET_DIFF})"
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_esc(title)}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0d1117;color:#c9d1d9;font-family:'Segoe UI',sans-serif;font-size:13px;padding:24px 28px}}
a{{color:#7289DA;text-decoration:none}} a:hover{{text-decoration:underline}}
.back{{display:inline-block;margin-bottom:20px;color:#7289DA;font-size:13px}}
h1{{font-size:20px;font-weight:700;margin-bottom:18px;color:#e6edf3}}
h2{{font-size:15px;font-weight:600;margin:24px 0 10px;color:#e6edf3;border-bottom:1px solid #30363d;padding-bottom:6px}}
/* Summary table */
.summary-table{{border-collapse:collapse;width:320px;margin-bottom:32px}}
.summary-table th{{background:#161b22;color:#8b949e;font-size:11px;font-weight:600;text-transform:uppercase;
  letter-spacing:.05em;padding:6px 10px;border:1px solid #30363d;text-align:left}}
.summary-table td{{padding:5px 10px;border:1px solid #21262d}}
.prog-name{{font-weight:600;min-width:120px}}
.prog-cell{{text-align:center;min-width:80px}}
.prog-ok{{color:#3fb950}}
.prog-caution{{color:#d29922}}
.prog-warn{{color:#f85149;font-weight:700}}
/* Wipe cards */
.wipe-card{{background:#161b22;border:1px solid #30363d;border-radius:8px;margin-bottom:20px;overflow:hidden}}
.wipe-header{{display:flex;align-items:center;justify-content:space-between;padding:10px 14px;
  background:#1c2128;border-bottom:1px solid #30363d}}
.wipe-title{{font-weight:600;color:#e6edf3;font-size:13px}}
.wipe-wcl{{font-size:11px;color:#7289DA}}
.wipe-body{{padding:12px 14px;display:flex;flex-wrap:wrap;gap:16px}}
/* Intermission blocks */
.im-block{{flex:0 0 auto}}
.im-title{{font-size:12px;font-weight:700;color:#8b949e;text-transform:uppercase;
  letter-spacing:.05em;margin-bottom:6px}}
.arrow-table{{border-collapse:collapse;font-size:12px}}
.arrow-table th{{background:#0d1117;color:#8b949e;font-size:10px;font-weight:600;
  text-transform:uppercase;padding:4px 8px;border:1px solid #30363d;text-align:left}}
.arrow-table td{{padding:3px 8px;border:1px solid #21262d}}
.arrow-t{{color:#8b949e;min-width:40px}}
.arrow-round{{color:#58a6ff;font-weight:700;min-width:40px}}
.arrow-name{{font-weight:600;min-width:110px}}
.arrow-role{{color:#8b949e;min-width:60px}}
.arrow-seq{{color:#d29922;font-size:10px;margin-left:2px}}
.arrow-rico{{opacity:0.55}}
.arrow-multi td{{background:#3d1a1a}}
.multi-warn{{margin-top:6px;font-size:11px;color:#f85149}}
.no-data{{color:#8b949e;font-style:italic;padding:8px 0}}
/* P3 circles */
.circle-bad td{{background:#3d1a1a;color:#f85149}}
.circle-pos{{color:#8b949e;min-width:60px}}
.circle-name{{font-weight:600;min-width:110px}}
.circle-role{{color:#8b949e;min-width:60px}}
.circle-hold{{color:#8b949e;min-width:50px}}
.circle-flag{{margin-top:6px;font-size:11px;color:#f85149}}
</style>
</head>
<body>
<a href="index.html" class="back">← Back to Index</a>
<h1>{_esc(title)}</h1>
<h2>Silverstrike Multi-Hit Summary — {len(wipes)} wipe(s)</h2>
<table class="summary-table" id="summary-tbl">
<thead><tr>
  <th>Player</th>
  <th>Extra Hits</th>
</tr></thead>
<tbody>{summary_rows}</tbody>
</table>
<h2>Per-Wipe Details</h2>
{wipe_cards}
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Crown progression page saved: {output_path}")


def write_gear_html(days_data: list, output_path: str, guild_name: str = "") -> None:
    """Write a consolidated gear page for all Normal runs."""
    normal_days = [d for d in days_data if d.get("difficulty", "") in ("Normal", "")]
    if not normal_days:
        return

    # Merge + deduplicate records across all normal days
    seen: set = set()
    merged: list = []
    for day_data in sorted(normal_days, key=lambda d: d["report_info"].get("startTime", 0)):
        for rec in day_data["records"]:
            key = (rec["player"], rec["item_id"])
            if key not in seen:
                seen.add(key)
                merged.append(rec)

    gear_header, gear_rows, _, total_players, total_crafted, no_craft = _build_gear_html(merged)
    nm_runs = len(normal_days)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Crafted Gear — {escape(guild_name)}</title>
{WOWHEAD_SCRIPTS}
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ background: #1a1a2e; color: #e0e0e0; font-family: -apple-system, 'Segoe UI', sans-serif; padding: 28px 24px; }}
.page-header {{ display: flex; align-items: center; gap: 16px; margin-bottom: 4px; }}
h1 {{ color: #4caf50; font-size: 22px; }}
.back-link {{ color: #7289DA; font-size: 13px; text-decoration: none; }}
.back-link:hover {{ text-decoration: underline; }}
.subtitle {{ color: #555; font-size: 13px; margin-bottom: 24px; }}
.stats {{ display: flex; gap: 16px; margin-bottom: 16px; flex-wrap: wrap; }}
.stat-box {{ background: #16213e; border-radius: 8px; padding: 10px 18px; }}
.stat-box .num {{ font-size: 24px; font-weight: bold; color: #4caf50; }}
.stat-box .label {{ font-size: 11px; color: #888; text-transform: uppercase; }}
.stat-box.warn .num {{ color: #ffc107; }}
.search-box {{ margin: 0 0 12px; }}
.search-box input {{ background: #16213e; border: 1px solid #2a2a4a; color: #e0e0e0; padding: 7px 14px; border-radius: 6px; width: 280px; font-size: 13px; }}
.search-box input::placeholder {{ color: #555; }}
.table-wrap {{ overflow-x: auto; border-radius: 8px; }}
table {{ border-collapse: collapse; min-width: 100%; }}
th {{ background: #16213e; color: #4caf50; padding: 8px 10px; text-align: left; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; white-space: nowrap; position: sticky; top: 0; z-index: 1; }}
th.player-header {{ position: sticky; left: 0; z-index: 2; background: #16213e; min-width: 140px; }}
td {{ padding: 6px 10px; border-bottom: 1px solid rgba(255,255,255,0.05); font-size: 13px; white-space: nowrap; }}
td.player-cell {{ position: sticky; left: 0; background: #0d1117; font-weight: 600; z-index: 1; }}
tr:hover td {{ background: rgba(255,255,255,0.04); }}
tr.no-craft td {{ opacity: 0.45; }}
td.empty {{ color: #333; }}
td.item-cell a {{ color: #a0b4ff; text-decoration: none; }}
td.item-cell a:hover {{ text-decoration: underline; }}
td.divider {{ border-left: 1px solid #2a2a4a; }}
td.center {{ text-align: center; }}
.spark-yes {{ color: #ffc107; font-weight: 700; }}
tr.section-sep td {{ color: #555; font-size: 11px; padding: 4px 10px; background: #0d1117; letter-spacing: 0.5px; }}
</style>
</head>
<body>
<div class="page-header">
  <h1>⚙ Crafted Gear — All Normal Runs</h1>
  <a href="index.html" class="back-link">← Back to Index</a>
</div>
<div class="subtitle">Compiled across {nm_runs} Normal run{"s" if nm_runs != 1 else ""}</div>
<div class="stats">
  <div class="stat-box"><div class="num">{total_players}</div><div class="label">Players</div></div>
  <div class="stat-box"><div class="num">{total_crafted}</div><div class="label">Crafted Items</div></div>
  <div class="stat-box warn"><div class="num">{no_craft}</div><div class="label">No Crafted Gear</div></div>
</div>
<div class="search-box"><input type="text" placeholder="Search player..." onkeyup="filterGear(this)"></div>
<div class="table-wrap">
  <table id="gear-table">
    <thead><tr>{gear_header}</tr></thead>
    <tbody>{gear_rows}</tbody>
  </table>
</div>
<script>
function filterGear(input) {{
  const filter = input.value.toLowerCase();
  for (let row of document.getElementById('gear-table').tBodies[0].rows)
    row.style.display = row.cells[0].textContent.toLowerCase().includes(filter) ? '' : 'none';
}}
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"[OK] Gear overview saved: {output_path}")


# ─── Roster Mapping (The Rupture) ────────────────────────────────────────────

# Maps character name (lowercase) -> (Player name, "Main" or "Alt")
ROSTER = {}
PLAYER_ROLES = {}  # player_name -> "Tank"/"Healer"/"DPS"


def load_roster(path="roster.txt"):
    """Load player roster from roster.txt. Format: PlayerName | Role | CharName:Main/Alt, ..."""
    global ROSTER, PLAYER_ROLES
    ROSTER = {}
    PLAYER_ROLES = {}
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = [p.strip() for p in line.split("|")]
                if len(parts) < 3:
                    continue
                player_name, role, chars_raw = parts[0], parts[1], parts[2]
                PLAYER_ROLES[player_name] = role
                for char_entry in chars_raw.split(","):
                    char_entry = char_entry.strip()
                    if not char_entry:
                        continue
                    if ":" in char_entry:
                        char_name, main_alt = char_entry.rsplit(":", 1)
                    else:
                        char_name, main_alt = char_entry, "Main"
                    ROSTER[char_name.strip().lower()] = (player_name, main_alt.strip())
    except FileNotFoundError:
        print("[WARN] roster.txt not found — player role/name lookup will be unavailable.")


load_roster()


def lookup_roster(char_name: str):
    """Look up a character in the roster. Returns (player_name, 'Main'/'Alt') or (char_name, 'Unknown')."""
    return ROSTER.get(char_name.lower(), (char_name, "Unknown"))


# ─── Boss Mechanic Definitions — loaded from boss_mechanics.py ───────────────

# (imported at top of file from boss_mechanics.py)

# ─── XLSX Output ─────────────────────────────────────────────────────────────

# Class background colors (lighter/muted versions for row backgrounds)
XLSX_CLASS_BG = {
    "DeathKnight": "77C41E3A", "DemonHunter": "77A330C9", "Druid": "77FF7C0A",
    "Evoker": "7733937F", "Hunter": "77AAD372", "Mage": "773FC7EB",
    "Monk": "7700FF98", "Paladin": "77F48CBA", "Priest": "77FFFFFF",
    "Rogue": "77FFF468", "Shaman": "770070DD", "Warlock": "778788EE",
    "Warrior": "77C69B6D",
}


def _build_boss_sheet(ws, boss_name: str, fights: list, report_info: dict, actor_lookup: dict):
    """One sheet per boss: rows = chars from both splits, cols = Split | Deaths | Death Times | Dmg Hits | Big Hits | Notes."""
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor="1a1a2e")
    data_font    = Font(name="Arial", size=10, color="FFFFFF")
    bold_font    = Font(name="Arial", size=10, color="FFFFFF", bold=True)
    death_font   = Font(name="Arial", size=10, color="CC0000", bold=True)
    border       = Border(bottom=Side(style="thin", color="333333"))
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left", vertical="center")
    wrap         = Alignment(horizontal="left", vertical="top", wrap_text=True)
    notes_fill   = PatternFill("solid", fgColor="1E2736")
    no_death_fill = PatternFill("solid", fgColor="111827")  # dark = clean
    death_fill   = PatternFill("solid", fgColor="3D1010")   # dark red = died

    def _parse_color(pct):
        """Return ARGB fill color for parse percentile (WarcraftLogs style)."""
        if not isinstance(pct, int): return "111827"
        if pct >= 99: return "E5CC80"   # gold/legendary
        if pct >= 95: return "FF8000"   # orange/epic
        if pct >= 75: return "A335EE"   # purple/rare
        if pct >= 50: return "0070DD"   # blue/uncommon
        if pct >= 25: return "1EFF00"   # green/common
        return "666666"                  # grey/low

    guild_name  = report_info.get("guild", {}).get("name", "") if report_info.get("guild") else ""
    report_date = ""
    if report_info.get("startTime"):
        report_date = datetime.fromtimestamp(report_info["startTime"] / 1000, tz=timezone.utc).strftime("%Y-%m-%d")

    ROLE_ORDER = {"Tank": 0, "Healer": 1, "DPS": 2}

    # Pre-fill entire sheet with dark background
    dk = PatternFill("solid", fgColor="0D1117")
    for r in range(1, 300):
        for c in range(1, 55):
            ws.cell(row=r, column=c).fill = dk

    # Title
    ws.merge_cells("A1:M1")
    ws["A1"].value = f"{boss_name} — {guild_name} ({report_date})"
    ws["A1"].font  = Font(name="Arial", bold=True, size=13, color="7289DA")
    ws.row_dimensions[1].height = 22

    ws["A2"].value = f"Deaths filtered: only first {WIPE_DEATH_THRESHOLD} deaths shown (wipe cascade excluded)"
    ws["A2"].font  = Font(name="Arial", size=9, color="888888", italic=True)

    # Row 3: fight overview (aggregated across all splits)
    def _k(v):
        if v >= 1_000_000_000: return f"{v/1_000_000_000:.1f}B"
        if v >= 1_000_000: return f"{v/1_000_000:.1f}M"
        if v > 0: return f"{v/1000:.0f}k"
        return "—"

    total_dur_ms   = sum(f.get("fight_dur_ms", 0) for f in fights)
    total_dmg_done = sum(e.get("total", 0) for f in fights for e in f.get("uptime_map", {}).values())
    total_healing  = sum(v for f in fights for v in f.get("healing_map", {}).values())
    total_dmg_taken = sum(v for f in fights for v in f.get("dmg_taken", {}).values())
    total_avoidable = sum(e.get("hits", 0) for f in fights for e in f.get("avoidable_damage", {}).values())
    dur_m, dur_s   = divmod(total_dur_ms // 1000, 60)

    overview_items = [
        ("⏱ Duration", f"{dur_m}:{dur_s:02d}"),
        ("⚔ Dmg Done", _k(total_dmg_done)),
        ("💚 Healing", _k(total_healing)),
        ("🛡 Dmg Taken", _k(total_dmg_taken)),
        ("☠ Avoid Hits", str(total_avoidable) if total_avoidable else "—"),
    ]
    ov_fill  = PatternFill("solid", fgColor="16213e")
    ov_label = Font(name="Arial", size=9, color="888888")
    ov_value = Font(name="Arial", size=11, bold=True, color="7289DA")
    for i, (label, value) in enumerate(overview_items):
        col = 2 + i * 2
        ws.cell(row=3, column=col,     value=label).font  = ov_label
        ws.cell(row=3, column=col,     value=label).fill  = ov_fill
        ws.cell(row=3, column=col + 1, value=value).font  = ov_value
        ws.cell(row=3, column=col + 1, value=value).fill  = ov_fill
        ws.cell(row=3, column=col + 1).alignment = center
    ws.row_dimensions[3].height = 22

    # Dynamic columns: base + boss-specific mechanics + Notes
    boss_name_base = boss_name.rsplit(" (", 1)[0]
    mech_defs = BOSS_MECHANICS.get(boss_name_base, [])
    has_interrupts = boss_name_base in BOSS_HAS_INTERRUPTS

    hr = 5
    base_cols   = ["", "Player", "Char", "Split", "Parse %", "Damage", "Healing",
                   "Deaths", "Killed by (time)", "Dmg Taken", "Uptime %",
                   "Interrupts", "Avoid Hits", ">10% HP Hits", "Big Hit Details"]
    base_widths = [3, 16, 18, 8, 9, 12, 12, 10, 28, 12, 10, 11, 11, 13, 40]
    mech_cols   = [m["label"] for m in mech_defs]
    cols   = base_cols + mech_cols + ["Notes"]
    widths = base_widths + [11] * len(mech_cols) + [40]

    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        c = ws.cell(row=hr, column=ci, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[hr].height = 30
    ws.freeze_panes = "A6"
    # Hide "Big Hit Details" column (col 15) by default — unhide to expand
    ws.column_dimensions[get_column_letter(15)].hidden = True

    # Collect all chars across both splits, tag with split number
    rows = []
    for fi, fight in enumerate(fights, 1):
        split_num    = fi
        avoidable    = fight.get("avoidable_damage", {})
        dmg_taken    = fight.get("dmg_taken", {})
        uptime_map   = fight.get("uptime_map", {})
        healing_map  = fight.get("healing_map", {})
        rankings_map = fight.get("rankings_map", {})
        interrupts_map  = fight.get("interrupts", {})
        mechanics_data  = fight.get("mechanics_data", {})
        fight_dur    = fight.get("fight_dur_ms", 0)
        seen_pids    = set()

        def _fmt_k(val):
            if val >= 1_000_000: return f"{val/1_000_000:.1f}M"
            if val > 0:          return f"{val/1000:.0f}k"
            return "—"

        def _fmt_dmg_taken(pid):
            return _fmt_k(dmg_taken.get(pid, 0))

        def _fmt_dps_dmg(pid):
            return _fmt_k(uptime_map.get(pid, {}).get("total", 0))

        def _fmt_healing(pid):
            return _fmt_k(healing_map.get(pid, 0))

        def _fmt_uptime(pid, role):
            if role not in ("DPS", "Healer"):
                return "—"
            active = uptime_map.get(pid, {}).get("activeTime", 0)
            if fight_dur > 0 and active > 0:
                return f"{min(active / fight_dur * 100, 100):.0f}%"
            return "—"

        def _fmt_parse(pid, char_name):
            pct = rankings_map.get(char_name.lower(), None)
            return pct if pct is not None else "—"

        def _fmt_interrupts(pid):
            count = interrupts_map.get(pid, 0)
            return count if count > 0 else "—"

        for pid, death_times in fight.get("deaths", {}).items():
            actor = actor_lookup.get(pid, {})
            char_name = actor.get("name", f"ID-{pid}")
            cls = actor.get("subType", "Unknown")
            player_name, _ = lookup_roster(char_name)
            role = PLAYER_ROLES.get(player_name, "DPS")
            av = avoidable.get(pid, {})
            details_str = "\n".join(f"{d['ability']}: {d['amount_k']} @ {d['time']}" for d in av.get("details", []))
            rows.append({
                "player": player_name, "char": char_name, "cls": cls,
                "split": split_num, "role": role,
                "deaths": len(death_times),
                "death_times": "\n".join(f"{d['ability']} @ {d['time']} ({d.get('fight_pct', 0)}%)" for d in death_times),
                "parse": _fmt_parse(pid, char_name),
                "dps_dmg": _fmt_dps_dmg(pid),
                "healing": _fmt_healing(pid),
                "dmg_taken": _fmt_dmg_taken(pid),
                "uptime": _fmt_uptime(pid, role),
                "interrupts": _fmt_interrupts(pid),
                "dmg_hits": av.get("hits", "—"),
                "big_hits": av.get("big_hits", "—"),
                "details": details_str,
                "mechanics": mechanics_data.get(pid, {}),
            })
            seen_pids.add(pid)
        # Also include players with 0 deaths from cast events (they were in the fight)
        for pid in fight.get("all_player_ids", set()):
            if pid not in seen_pids:
                actor = actor_lookup.get(pid, {})
                char_name = actor.get("name", f"ID-{pid}")
                cls = actor.get("subType", "Unknown")
                player_name, _ = lookup_roster(char_name)
                role = PLAYER_ROLES.get(player_name, "DPS")
                av = avoidable.get(pid, {})
                details_str = "\n".join(f"{d['ability']}: {d['amount_k']} @ {d['time']}" for d in av.get("details", []))
                rows.append({
                    "player": player_name, "char": char_name, "cls": cls,
                    "split": split_num, "role": role,
                    "deaths": 0, "death_times": "",
                    "parse": _fmt_parse(pid, char_name),
                    "dps_dmg": _fmt_dps_dmg(pid),
                    "healing": _fmt_healing(pid),
                    "dmg_taken": _fmt_dmg_taken(pid),
                    "uptime": _fmt_uptime(pid, role),
                    "interrupts": _fmt_interrupts(pid),
                    "dmg_hits": av.get("hits", "—"),
                    "big_hits": av.get("big_hits", "—"),
                    "details": details_str,
                    "mechanics": mechanics_data.get(pid, {}),
                })

    # Sort: role → player name → split
    rows.sort(key=lambda r: (ROLE_ORDER.get(r["role"], 2), r["player"].lower(), r["split"]))

    row = hr + 1
    current_role = None
    for r in rows:
        if r["role"] != current_role:
            current_role = r["role"]
            for ci in range(1, len(cols) + 1):
                sc = ws.cell(row=row, column=ci)
                sc.fill = PatternFill("solid", fgColor="111827")
                sc.border = border
            ws.cell(row=row, column=2).value = f"── {current_role}s ──"
            ws.cell(row=row, column=2).font = Font(name="Arial", size=9, bold=True, color="7289DA")
            ws.row_dimensions[row].height = 30
            row += 1

        cls_hex = CLASS_COLORS.get(r["cls"], "#CCCCCC").lstrip("#")
        cls_bar_fill = PatternFill("solid", fgColor=cls_hex)
        died = r["deaths"] > 0
        row_fill = death_fill if died else no_death_fill
        dmg_hits      = r.get("dmg_hits", "—")
        big_hits      = r.get("big_hits", "—")
        details       = r.get("details", "")
        interrupt_val = r.get("interrupts", "—")
        big_fill      = PatternFill("solid", fgColor="5D2D1A") if isinstance(big_hits, int) and big_hits > 0 else row_fill
        interrupt_fill = (
            PatternFill("solid", fgColor="1A3A1A") if isinstance(interrupt_val, int) and interrupt_val > 0
            else PatternFill("solid", fgColor="5D1A1A") if has_interrupts
            else row_fill
        )
        char_font = Font(name="Arial", size=10, color=cls_hex, bold=False)

        parse_val  = r.get("parse", "—")
        parse_fill = PatternFill("solid", fgColor=_parse_color(parse_val)) if isinstance(parse_val, int) else row_fill
        parse_font = Font(name="Arial", size=10, color="000000" if isinstance(parse_val, int) else "FFFFFF", bold=isinstance(parse_val, int))

        def _mech_fill(mech, count):
            if not isinstance(count, int) or count == 0:
                return row_fill
            return PatternFill("solid", fgColor="1A3D1A") if mech["type"] == "soak" else PatternFill("solid", fgColor="5D1A1A")

        mech_data  = r.get("mechanics", {})
        mech_vals  = [mech_data.get(m["label"], 0) or "—" for m in mech_defs]
        mech_fills = [_mech_fill(m, mech_data.get(m["label"], 0)) for m in mech_defs]

        base_vals  = ["", r["player"], r["char"], f"Split {r['split']}",
                      parse_val, r.get("dps_dmg", "—"), r.get("healing", "—"),
                      r["deaths"] or "—", r["death_times"] or "—",
                      r.get("dmg_taken", "—"), r.get("uptime", "—"), interrupt_val,
                      dmg_hits, big_hits, details or "—"]
        base_fills = [cls_bar_fill, row_fill, row_fill, row_fill,
                      parse_fill, row_fill, row_fill,
                      row_fill, row_fill, row_fill, row_fill, interrupt_fill, row_fill, big_fill, big_fill]
        base_aligns = [left, left, left, center, center, center, center, center, wrap, center, center, center, center, center, wrap]
        base_fonts  = [data_font, bold_font, char_font, data_font,
                       parse_font, data_font, data_font,
                       death_font if died else data_font, data_font,
                       data_font, data_font, data_font, data_font, data_font, data_font]

        vals   = base_vals  + mech_vals  + [""]
        fills  = base_fills + mech_fills + [notes_fill]
        aligns = base_aligns + [center] * len(mech_defs) + [wrap]
        fonts  = base_fonts  + [data_font] * len(mech_defs) + [data_font]

        for ci, (val, fill, align, font) in enumerate(zip(vals, fills, aligns, fonts), 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = fill
            c.font = font
            c.alignment = align
            c.border = border

        ws.row_dimensions[row].height = 30
        row += 1


def write_xlsx(records: list, report_info: dict, report_code: str, output_path: str, split_data: dict = None, boss_data: dict = None, actors: list = None):
    """Write crafted gear data to XLSX with Mains, Alts, and Split tabs."""
    wb = Workbook()
    
    guild_name = "Unknown Guild"
    if report_info.get("guild"):
        guild_name = report_info["guild"].get("name", "Unknown Guild")
    report_title = report_info.get("title", report_code)
    report_date = ""
    if report_info.get("startTime"):
        report_date = datetime.fromtimestamp(report_info["startTime"] / 1000, tz=timezone.utc).strftime("%Y-%m-%d")

    # Group records by player and enrich with roster data
    players = {}
    for rec in records:
        char_name = rec["player"]
        player_name, role = lookup_roster(char_name)
        key = (player_name, char_name)
        if key not in players:
            players[key] = {"player": player_name, "char": char_name, "class": rec["class"], "role": role, "items": []}
        if rec["item_id"] != 0:
            players[key]["items"].append(rec)

    # Split into mains and alts
    mains = {k: v for k, v in players.items() if v["role"] == "Main"}
    alts = {k: v for k, v in players.items() if v["role"] != "Main"}

    # Find max items for column count
    all_items_counts = [len(v["items"]) for v in players.values()]
    max_items = max(all_items_counts) if all_items_counts else 2
    max_items = max(max_items, 2)

    # Role sort order
    ROLE_ORDER = {"Tank": 0, "Healer": 1, "DPS": 2}

    def sort_key(k):
        pdata = players.get(k) or mains.get(k) or alts.get(k)
        player_name = pdata["player"]
        role_rank = ROLE_ORDER.get(PLAYER_ROLES.get(player_name, "DPS"), 2)
        return (role_rank, -len(pdata["items"]), player_name.lower())

    # Styling
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    black_font = Font(name="Arial", size=10, color="FFFFFF")
    black_bold = Font(name="Arial", size=10, color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="333333"))
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    spark_font = Font(name="Arial", size=10, color="00FF88", bold=True)
    no_spark_font = Font(name="Arial", size=10, color="FFFFFF")
    no_craft_fill = PatternFill("solid", fgColor="3D3D3D")
    no_craft_font = Font(name="Arial", size=10, color="999999")

    def get_class_fill(cls_name):
        color = XLSX_CLASS_BG.get(cls_name, "77666666")
        return PatternFill("solid", fgColor=color)

    def prefill_dark(ws, rows=300, cols=60):
        """Pre-fill entire sheet area with dark background so no white cells show."""
        dk = PatternFill("solid", fgColor="0D1117")
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c).fill = dk

    def build_sheet(ws, title, player_data):
        # Title
        total_cols = 2 + max_items * 3
        prefill_dark(ws, rows=200, cols=total_cols + 5)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws["A1"].value = f"{title} — {guild_name} — {report_title} ({report_date})"
        ws["A1"].font = Font(name="Arial", bold=True, size=13, color="7289DA")
        ws.row_dimensions[1].height = 26

        ws["A2"].value = f"Report: https://www.warcraftlogs.com/reports/{report_code}"
        ws["A2"].font = Font(name="Arial", size=9, color="888888", italic=True)

        # Headers: Player | Char | Slot 1 | Ilvl 1 | Spark? 1 | Slot 2 | Ilvl 2 | Spark? 2 ...
        hr = 4
        headers = ["Player", "Char"]
        for i in range(max_items):
            headers += [f"Slot {i+1}", f"Ilvl", f"Spark?"]
        
        col_widths = [16, 18] + [12, 8, 8] * max_items
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=hr, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w
            # Divider before each item group
            if ci > 2 and (ci - 3) % 3 == 0 and (ci - 3) // 3 > 0:
                cell.border = Border(bottom=Side(style="thin", color="333333"), left=Side(style="medium", color="7289DA"))
        ws.row_dimensions[hr].height = 20
        ws.freeze_panes = "A5"

        # Data rows — sorted by role then items
        row = hr + 1
        current_role = None
        for key in sorted(player_data.keys(), key=sort_key):
            pdata = player_data[key]
            cls = pdata["class"]
            cls_fill = get_class_fill(cls)
            has_items = len(pdata["items"]) > 0
            player_name = pdata["player"]
            player_role = PLAYER_ROLES.get(player_name, "DPS")

            # Role separator
            if player_role != current_role:
                current_role = player_role
                sep_cell = ws.cell(row=row, column=1, value=f"── {current_role}s ──")
                sep_cell.font = Font(name="Arial", size=9, bold=True, color="7289DA")
                for ci in range(1, total_cols + 1):
                    ws.cell(row=row, column=ci).fill = PatternFill("solid", fgColor="111122")
                    ws.cell(row=row, column=ci).border = border
                row += 1

            if not has_items:
                fill = no_craft_fill
                name_font = no_craft_font
                txt_font = no_craft_font
            else:
                fill = cls_fill
                name_font = black_bold
                txt_font = black_font

            # Player name
            c = ws.cell(row=row, column=1, value=player_name)
            c.font = name_font
            c.fill = fill
            c.border = border
            c.alignment = left

            # Char name
            c = ws.cell(row=row, column=2, value=pdata["char"])
            c.font = txt_font
            c.fill = fill
            c.border = border
            c.alignment = left

            # Items: Slot | Ilvl | Spark?
            for i in range(max_items):
                base_col = 3 + i * 3
                is_divider = i > 0

                if i < len(pdata["items"]):
                    item = pdata["items"][i]
                    spark_yes = "Yes" in str(item["spark_used"])
                    
                    vals = [item["slot"], item["item_level"], "Yes" if spark_yes else "No"]
                    for j, val in enumerate(vals):
                        cell = ws.cell(row=row, column=base_col + j, value=val)
                        cell.fill = fill
                        cell.border = border
                        if j == 1:
                            cell.alignment = center
                            cell.font = txt_font
                        elif j == 2:
                            cell.alignment = center
                            cell.font = spark_font if spark_yes else txt_font
                        else:
                            cell.font = txt_font
                        # Divider
                        if j == 0 and is_divider:
                            cell.border = Border(bottom=Side(style="thin", color="333333"), left=Side(style="medium", color="7289DA"))
                else:
                    for j in range(3):
                        cell = ws.cell(row=row, column=base_col + j, value="—")
                        cell.font = no_craft_font if not has_items else Font(name="Arial", size=10, color="AAAAAA")
                        cell.fill = fill
                        cell.border = border
                        cell.alignment = center
                        if j == 0 and is_divider:
                            cell.border = Border(bottom=Side(style="thin", color="333333"), left=Side(style="medium", color="7289DA"))

            row += 1

        # Auto-filter
        ws.auto_filter.ref = f"A{hr}:{get_column_letter(total_cols)}{row - 1}"

    # Build crafted gear sheets
    ws_mains = wb.active
    ws_mains.title = "Mains"
    build_sheet(ws_mains, "Mains — Crafted Gear", mains)

    ws_alts = wb.create_sheet("Alts")
    build_sheet(ws_alts, "Alts — Crafted Gear", alts)

    # Build split consumable/defensive tabs
    if split_data:
        for split_name, split_fights in split_data.items():
            ws_split = wb.create_sheet(split_name)
            _build_split_sheet(ws_split, split_name, split_fights, report_info, actors_list=None)

    # Build boss analysis tabs (deaths + notes)
    if boss_data:
        actor_lookup = {a["id"]: a for a in (actors or [])}
        for boss_name, fights in boss_data.items():
            safe_name = boss_name[:28]  # Excel sheet name limit
            ws_boss = wb.create_sheet(safe_name)
            _build_boss_sheet(ws_boss, boss_name, fights, report_info, actor_lookup)

    wb.save(output_path)
    print(f"[OK] XLSX report saved to: {output_path}")


def _build_split_sheet(ws, title, fights_data, report_info, actors_list):
    """Build a split tab showing consumable/defensive usage per player per fight.
    
    fights_data: list of { "fight_name": str, "difficulty": str, "player_casts": { sourceID: [cast_info] } }
    """
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    data_font = Font(name="Arial", size=10, color="FFFFFF")
    border = Border(bottom=Side(style="thin", color="333333"))
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    cat_fills = {
        "Health": PatternFill("solid", fgColor="77CC4444"),
        "Combat Pot": PatternFill("solid", fgColor="77AA44CC"),
        "Defensive": PatternFill("solid", fgColor="774488CC"),
    }

    guild_name = ""
    if report_info.get("guild"):
        guild_name = report_info["guild"].get("name", "")
    report_title = report_info.get("title", "")
    report_date = ""
    if report_info.get("startTime"):
        report_date = datetime.fromtimestamp(report_info["startTime"] / 1000, tz=timezone.utc).strftime("%Y-%m-%d")

    # Pre-fill entire sheet with dark background
    dk = PatternFill("solid", fgColor="0D1117")
    num_cols = 1 + len(fights_data) * 4 + 10
    for r in range(1, 200):
        for c in range(1, num_cols + 1):
            ws.cell(row=r, column=c).fill = dk

    full_title = f"{title} — {guild_name} — {report_title} ({report_date})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(fights_data) * 3)
    ws["A1"].value = full_title
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="7289DA")
    ws.row_dimensions[1].height = 26

    ws["A2"].value = f"Report: {report_title} ({report_date})"
    ws["A2"].font = Font(name="Arial", size=9, color="888888", italic=True)

    # Collect all unique players across all fights in this split
    all_player_ids = set()
    for fd in fights_data:
        all_player_ids.update(fd.get("player_casts", {}).keys())

    # Get actor lookup
    actors = {}
    if report_info.get("masterData") and report_info["masterData"].get("actors"):
        for a in report_info["masterData"]["actors"]:
            actors[a["id"]] = a

    # Sort players by roster role
    ROLE_ORDER = {"Tank": 0, "Healer": 1, "DPS": 2}
    def player_sort(pid):
        actor = actors.get(pid, {})
        name = actor.get("name", "")
        player_name, _ = lookup_roster(name)
        role = PLAYER_ROLES.get(player_name, "DPS")
        return (ROLE_ORDER.get(role, 2), player_name.lower())

    sorted_players = sorted(all_player_ids, key=player_sort)

    # Two-row header: row 3 = boss names (merged 3 cols), row 4 = sub-columns
    hr_boss = 3
    hr_sub  = 4

    # "Player" cell spans both header rows
    ws.merge_cells(start_row=hr_boss, start_column=1, end_row=hr_sub, end_column=1)
    c = ws.cell(row=hr_boss, column=1, value="Player")
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
    ws.column_dimensions["A"].width = 18

    sub_labels = ["Health Pots", "Healthstone", "Combat Pots", "Defensives"]
    col_widths  = [18, 14, 18, 28]
    num_cols    = len(sub_labels)

    for fi, fd in enumerate(fights_data):
        fname   = fd.get("fight_name", "Boss")
        base_ci = 2 + fi * num_cols  # first column of this boss group

        # Merge boss name across num_cols columns
        ws.merge_cells(start_row=hr_boss, start_column=base_ci, end_row=hr_boss, end_column=base_ci + num_cols - 1)
        bc = ws.cell(row=hr_boss, column=base_ci, value=fname)
        bc.font = header_font
        bc.fill = PatternFill("solid", fgColor="111827")
        bc.alignment = Alignment(horizontal="center", vertical="center")
        bc.border = Border(bottom=Side(style="thin", color="333333"),
                           left=Side(style="medium", color="7289DA"))

        # Sub-column headers
        for j, (label, w) in enumerate(zip(sub_labels, col_widths)):
            ci = base_ci + j
            sc = ws.cell(row=hr_sub, column=ci, value=label)
            sc.font = header_font
            sc.fill = header_fill
            sc.alignment = Alignment(horizontal="center", vertical="center")
            sc.border = border
            if j == 0:
                sc.border = Border(bottom=Side(style="thin", color="333333"),
                                   left=Side(style="medium", color="7289DA"))
            ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[hr_boss].height = 20
    ws.row_dimensions[hr_sub].height  = 18
    ws.freeze_panes = "A5"

    # Data rows
    row = hr_sub + 1
    for pid in sorted_players:
        actor = actors.get(pid, {})
        char_name = actor.get("name", f"ID-{pid}")
        cls = actor.get("subType", "Unknown")
        cls_bg = XLSX_CLASS_BG.get(cls, "77666666")
        cls_fill = PatternFill("solid", fgColor=cls_bg)
        player_name, _ = lookup_roster(char_name)

        c = ws.cell(row=row, column=1, value=f"{player_name}\n({char_name})")
        c.font = Font(name="Arial", size=10, color="FFFFFF", bold=True)
        c.fill = cls_fill
        c.border = border
        c.alignment = Alignment(vertical="center", wrap_text=True)

        for fi, fd in enumerate(fights_data):
            base_col = 2 + fi * num_cols
            casts = fd.get("player_casts", {}).get(pid, [])

            # Group by category
            health      = [c for c in casts if c["category"] == "Health"]
            healthstone = [c for c in casts if c["category"] == "Healthstone"]
            combat      = [c for c in casts if c["category"] == "Combat Pot"]
            defensives  = [c for c in casts if c["category"] == "Defensive"]

            for j, (items, cat) in enumerate([(health, "Health"), (healthstone, "Healthstone"), (combat, "Combat Pot"), (defensives, "Defensive")]):
                cell = ws.cell(row=row, column=base_col + j)
                cell.fill = cls_fill
                cell.border = border
                
                if items:
                    lines = [f"{it['spell']} @ {it['time']}" for it in items]
                    cell.value = "\n".join(lines)
                    cell.font = data_font
                    cell.alignment = wrap
                else:
                    cell.value = "—"
                    cell.font = Font(name="Arial", size=10, color="AAAAAA")
                    cell.alignment = center

                # Divider
                if j == 0:
                    cell.border = Border(bottom=Side(style="thin", color="333333"), left=Side(style="medium", color="7289DA"))

        # Height = tallest cell: max casts in any single category across all fights
        max_lines = 1
        for fd in fights_data:
            player_casts = fd.get("player_casts", {}).get(pid, [])
            for cat in ("Health", "Healthstone", "Combat Pot", "Defensive"):
                count = sum(1 for c in player_casts if c["category"] == cat)
                max_lines = max(max_lines, count)
        ws.row_dimensions[row].height = max(18, 15 * max_lines)
        row += 1

def load_config(path="wcl_config.txt"):
    """Load credentials from a config file.
    Supports multiple REPORT_URL= lines (same key, repeated).
    Inline comments are stripped: REPORT_URL=https://... # my comment
    """
    config = {}
    report_urls = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                key, val = line.split("=", 1)
                key = key.strip()
                # Extract split_start from inline comment before stripping it
                split_start = 1
                if " #" in val:
                    comment = val.split(" #", 1)[1]
                    for part in comment.split():
                        if part.startswith("split_start="):
                            try:
                                split_start = int(part.split("=", 1)[1])
                            except ValueError:
                                pass
                val = val.split(" #")[0].strip()
                if not val:
                    continue
                if key.startswith("REPORT_URL"):
                    report_urls.append({"url": val, "split_start": split_start})
                else:
                    config[key] = val
    except FileNotFoundError:
        pass
    config["REPORT_URLS"] = report_urls
    return config


def _extract_report_code(url: str) -> str:
    if "warcraftlogs.com" in url:
        parts = url.rstrip("/").split("/")
        return parts[-1].split("#")[0].split("?")[0]
    return url


def _cache_path(report_code: str) -> str:
    os.makedirs("data", exist_ok=True)
    return os.path.join("data", f"raid_{report_code}.json")


def _fix_int_keys(boss_data: dict) -> dict:
    """After JSON load, restore integer actor-ID keys in fields that use actor IDs as keys.
    rankings_map uses char-name strings as keys — excluded intentionally.
    """
    int_key_fields = (
        "uptime_map", "dmg_taken", "healing_map", "interrupts",
        "avoidable_damage", "deaths", "mechanics_data",
        "defensive_casts", "external_casts", "horror_damage",
    )
    for fights in boss_data.values():
        for fight in fights:
            for field in int_key_fields:
                if field in fight and isinstance(fight[field], dict):
                    fight[field] = {int(k): v for k, v in fight[field].items()}
            # horror_waves: list of wave dicts, each with per_player {str->int keys after JSON}
            for wave in fight.get("horror_waves", []):
                if "per_player" in wave and isinstance(wave["per_player"], dict):
                    wave["per_player"] = {int(k): v for k, v in wave["per_player"].items()}
    return boss_data


def process_report(token: str, report_code: str, fight_input: str = "all") -> dict:
    """Fetch and process one WCL report. Loads from cache if available."""
    cache = _cache_path(report_code)
    if os.path.exists(cache):
        print(f"\n[CACHE] Loading {report_code} from cache (delete data/raid_{report_code}.json to re-fetch).")
        with open(cache, encoding="utf-8") as f:
            result = json.load(f)
        result["boss_data"] = _fix_int_keys(result["boss_data"])
        result["wipe_data"] = _fix_int_keys(result["wipe_data"])
        return result

    print(f"\n[...] Fetching report info for: {report_code}")
    report_info = fetch_report_info(token, report_code)
    guild_name = report_info.get("guild", {}).get("name", "") if report_info.get("guild") else ""
    print(f"[OK] Report: {report_info.get('title', 'Untitled')} ({guild_name})")

    all_raw_fights  = report_info.get("fights", [])
    all_fights      = [f for f in all_raw_fights if f.get("encounterID", 0) > 0 and f.get("kill")]
    all_wipe_fights = [f for f in all_raw_fights if f.get("encounterID", 0) > 0 and not f.get("kill")]
    selected_fights = all_fights

    if all_fights:
        print(f"\nBoss kills found ({len(all_fights)}):")
        for f in all_fights:
            diff = {3: "Normal", 4: "Heroic", 5: "Mythic"}.get(f.get("difficulty", 0), f"D{f.get('difficulty','?')}")
            print(f"  [{f['id']:>3}] {f['name']} — {diff} (Kill)")
        if fight_input == "interactive":
            fight_input = input("\nFight IDs to analyze (comma-separated, or 'all'): ").strip()
        if fight_input.lower() != "all" and fight_input != "":
            selected_ids = {int(x.strip()) for x in fight_input.split(",") if x.strip().isdigit()}
            selected_fights = [f for f in all_fights if f["id"] in selected_ids]

    all_actors  = report_info.get("masterData", {}).get("actors", [])
    actors      = [a for a in all_actors if a.get("type") == "Player"]
    npc_actors  = [a for a in all_actors if a.get("type") == "NPC"]
    print(f"[OK] Found {len(actors)} player actors in report.")
    ability_names = {ab["gameID"]: ab["name"] for ab in report_info.get("masterData", {}).get("abilities", [])}
    actor_lookup  = {a["id"]: a for a in actors}
    # Colossal Horror actor IDs for this report (Chimaerus add)
    _COLOSSAL_HORROR_GAME_IDS = {245556, 249341, 257691}
    horror_actor_ids = {a["id"] for a in npc_actors if a.get("gameID") in _COLOSSAL_HORROR_GAME_IDS}
    # Abyssal Voidshaper + Obscurion Endwalker (Heroic/Mythic transform) actor IDs for Averzian
    _AVERZIAN_ADD_GAME_IDS = {252918, 256942}  # Abyssal Voidshaper + Obscurion Endwalker
    averzian_add_actor_ids = {a["id"] for a in npc_actors if a.get("gameID") in _AVERZIAN_ADD_GAME_IDS}

    # Gear
    all_combatant_events = []
    fight_spec_roles: dict = {}   # fid → {pid: role}
    for fight in selected_fights:
        fid, fname = fight["id"], fight["name"]
        print(f"  [...] Fetching gear from fight {fid}: {fname}...")
        try:
            events = fetch_combatant_info_events(token, report_code, fid)
            all_combatant_events.extend(events)
            fight_spec_roles[fid] = {
                e["sourceID"]: SPEC_ROLES.get(e.get("specID"), "DPS")
                for e in events if e.get("sourceID") and e.get("specID")
            }
            print(f"         Got {len(events)} players' gear.")
        except Exception as e:
            print(f"    [WARN] Gear fetch failed for fight {fid}: {e}")

    player_max_hp = {}
    for ce in all_combatant_events:
        pid, stamina = ce.get("sourceID"), ce.get("stamina", 0)
        if pid and stamina:
            player_max_hp[pid] = max(player_max_hp.get(pid, 0), stamina * 20)

    print("\n[...] Analyzing crafted gear...")
    records = analyze_players(actors, all_combatant_events)
    crafted_count = sum(1 for r in records if r["item_id"] != 0)
    player_count  = len(set(r["player"] for r in records))
    no_craft_count= sum(1 for r in records if r["item_id"] == 0)
    print(f"[OK] Found {crafted_count} crafted items across {player_count} players.")
    if no_craft_count > 0:
        print(f"[!!] {no_craft_count} player(s) have NO detected crafted gear.")

    # Split detection — group by encounter, order of kills = split number
    encounter_kills = {}
    for fight in selected_fights:
        eid = fight.get("encounterID", 0)
        encounter_kills.setdefault(eid, []).append(fight)
    max_splits = max((len(v) for v in encounter_kills.values()), default=1)
    split_fights = {i + 1: [] for i in range(max_splits)}
    for kills in encounter_kills.values():
        for i, kill in enumerate(sorted(kills, key=lambda f: f["id"])):
            split_fights[i + 1].append(kill)

    print(f"\n[...] Detected {max_splits} split(s).")

    # Cast events per split
    split_data = {}
    for split_num, sfl in split_fights.items():
        split_name = f"Split {split_num}"
        print(f"\n[...] Fetching spell usage for {split_name}...")
        split_fight_data = []
        for fight in sfl:
            fid   = fight["id"]
            fname = fight["name"]
            diff  = {3: "Normal", 4: "Heroic", 5: "Mythic"}.get(fight.get("difficulty", 0), "")
            print(f"  [...] Fetching casts for {fname} ({diff})...")
            try:
                cast_events  = fetch_cast_events(token, report_code, fid)
                player_casts = analyze_fight_casts(cast_events, fight.get("startTime", 0), actor_lookup)
                total        = sum(len(v) for v in player_casts.values())
                print(f"         Found {total} tracked spell uses across {len(player_casts)} players.")
                split_fight_data.append({"fight_name": f"{fname} ({diff})", "fight_id": fid, "player_casts": player_casts})
            except Exception as e:
                print(f"    [WARN] Casts fetch failed for fight {fid}: {e}")
        split_data[split_name] = split_fight_data

    # Boss analysis — one fight dict per fight, tagged with split_num
    boss_data = {}
    encounter_kill_count = {}
    print(f"\n[...] Fetching death + damage events per boss...")
    player_roles_map = {pid: PLAYER_ROLES.get(lookup_roster(actor_lookup[pid]["name"])[0], "DPS")
                        for pid in actor_lookup if "name" in actor_lookup[pid]}

    for fight in selected_fights:
        fid  = fight["id"]
        fname= fight["name"]
        eid  = fight.get("encounterID", 0)
        diff = {3: "Normal", 4: "Heroic", 5: "Mythic"}.get(fight.get("difficulty", 0), "")
        boss_key = f"{fname} ({diff})"
        encounter_kill_count[eid] = encounter_kill_count.get(eid, 0) + 1
        split_num = encounter_kill_count[eid]
        try:
            death_events     = fetch_death_events(token, report_code, fid)
            damage_events    = fetch_damage_taken_events(token, report_code, fid)
            interrupt_events = fetch_interrupt_events(token, report_code, fid)
            cast_events      = fetch_boss_cast_events(token, report_code, fid)
            fight_start      = fight.get("startTime", 0)
            fight_dur_ms     = fight.get("endTime", 0) - fight_start
            deaths           = analyze_deaths(death_events, fight_start, ability_names,
                                              fight_end_ms=fight_dur_ms, damage_events=damage_events)
            avoidable        = analyze_avoidable_damage(damage_events, actor_lookup, fight_start,
                                                        ability_names, player_max_hp, player_roles_map)
            dmg_taken        = aggregate_damage_taken(damage_events, actor_lookup)
            uptime_map       = fetch_uptime_table(token, report_code, fid)
            healing_map      = fetch_healing_table(token, report_code, fid)
            rankings_map     = fetch_rankings(token, report_code, fid)
            interrupts       = analyze_interrupts(interrupt_events, actor_lookup)
            mech_defs        = BOSS_MECHANICS.get(fname, [])
            mechanics_data   = analyze_boss_mechanics(damage_events, actor_lookup, mech_defs)
            frontal_failures = analyze_frontal_failures(damage_events, actor_lookup, mech_defs, fight_start,
                                                        cast_events=cast_events)
            # Extract per-player defensive + external casts for this fight from split_data
            defensive_casts: dict = {}
            external_casts:  dict = {}
            for sdata in split_data.values():
                for fd in sdata:
                    if fd.get("fight_id") == fid:
                        for pid_raw, casts in fd.get("player_casts", {}).items():
                            pid = int(pid_raw) if isinstance(pid_raw, str) else pid_raw
                            defs = [{"spell": c["spell"], "time": c["time"]}
                                    for c in casts
                                    if c.get("category") in ("Defensive", "Healthstone", "Health")]
                            exts = [{"spell": c["spell"], "time": c["time"]}
                                    for c in casts if c.get("category") == "External"]
                            if defs:
                                defensive_casts[pid] = defs
                            if exts:
                                external_casts[pid] = exts
            all_pids = set()
            for sdata in split_data.values():
                for fd in sdata:
                    if fd.get("fight_id") == fid:
                        all_pids.update(fd.get("player_casts", {}).keys())
            all_pids.update(dmg_taken.keys())
            all_pids.update(uptime_map.keys())
            all_pids.update(deaths.keys())
            all_pids = {pid for pid in all_pids if pid in actor_lookup}
            alndust_groups = (analyze_alndust_groups(damage_events, all_pids, fight_start,
                                                      report_code=report_code,
                                                      actor_lookup=actor_lookup,
                                                      split_num=split_num)
                              if "Chimaerus" in fname else [])
            if "Vaelgor" in fname:
                gloom_debuffs = fetch_gloom_debuff_events(token, report_code, fid)
                gloom_soaks   = analyze_gloom_soaks(gloom_debuffs, all_pids, fight_start)
            else:
                gloom_soaks = []
            if "Averzian" in fname:
                void_marked_events = fetch_void_marked_events(token, report_code, fid)
                void_marked_data   = analyze_void_marked(void_marked_events, all_pids, fight_start)
                # add_damage: list of per-wave dicts from WCL table endpoint
                add_damage = fetch_add_instance_damage(
                    token, report_code,
                    fight_start, fight_start + fight_dur_ms,
                    252918)  # Abyssal Voidshaper game ID
            else:
                void_marked_data = []
                add_damage       = {}
            # Per-player damage to Colossal Horror + per-wave analysis (Chimaerus only)
            horror_damage: dict = {}
            horror_waves: list  = []
            if "Chimaerus" in fname and horror_actor_ids:
                horror_done_events = fetch_damage_done_events(token, report_code, fid,
                                                              target_ids=horror_actor_ids)
                for ev in horror_done_events:
                    if ev.get("type") != "damage":
                        continue
                    src = ev.get("sourceID")
                    if src in actor_lookup:
                        horror_damage[src] = horror_damage.get(src, 0) + ev.get("amount", 0)
                if alndust_groups:
                    horror_waves = analyze_chimaerus_horror_waves(
                        token, report_code, fid,
                        alndust_groups,
                        horror_actor_ids, actor_lookup, fight_start,
                        fight_start + fight_dur_ms,
                    )
            # Vorasius: tank swap failures
            vorasius_tank_swaps = (analyze_vorasius_tank_swaps(
                damage_events, actor_lookup, fight_start, fight_spec_roles.get(fid, {}))
                if "Vorasius" in fname else [])
            # All bosses: per-mechanic hit timestamps (for table tooltips)
            mechanic_timestamps = analyze_mechanic_timestamps(
                damage_events, actor_lookup, fight_start, mech_defs)
            # Salhadaar: Shadow Fracture interrupt+CC wave table + Void Infusion wipe flags
            salhadaar_fracture_waves = []
            salhadaar_wipe_events    = []
            if "Salhadaar" in fname:
                sal_cc_events = fetch_salhadaar_cc_events(
                    token, report_code, fid, fight_start, fight_start + fight_dur_ms)
                salhadaar_fracture_waves = analyze_salhadaar_fracture_waves(
                    interrupt_events, actor_lookup, fight_start,
                    ability_names=ability_names, spell_names_map=SPELL_NAMES,
                    cc_events=sal_cc_events)
                salhadaar_wipe_events = detect_salhadaar_wipe_events(
                    cast_events, fight_start, ability_names=ability_names)
            boss_data.setdefault(boss_key, []).append({
                "fight_id": fid, "fight_dur_ms": fight_dur_ms, "split_num": split_num,
                "deaths": deaths, "all_player_ids": all_pids,
                "avoidable_damage": avoidable, "dmg_taken": dmg_taken,
                "uptime_map": uptime_map, "healing_map": healing_map,
                "rankings_map": rankings_map, "interrupts": interrupts,
                "mechanics_data": mechanics_data, "frontal_failures": frontal_failures,
                "spec_roles": fight_spec_roles.get(fid, {}),
                "defensive_casts": defensive_casts,
                "external_casts":  external_casts,
                "alndust_groups":  alndust_groups,
                "gloom_soaks":     gloom_soaks,
                "void_marked":     void_marked_data,
                "add_damage":      add_damage,
                "horror_damage":   horror_damage,
                "horror_waves":    horror_waves,
                "vorasius_tank_swaps":     vorasius_tank_swaps,
                "mechanic_timestamps":     mechanic_timestamps,
                "salhadaar_fracture_waves": salhadaar_fracture_waves,
                "salhadaar_wipe_events":    salhadaar_wipe_events,
            })
            print(f"  [OK] {fname} (Split {split_num}): {len(deaths)} death(s), {sum(interrupts.values())} interrupts.")
        except Exception as e:
            print(f"  [WARN] Could not fetch events for fight {fid}: {e}")

    # ── Wipe analysis ──
    wipe_data: dict = {}
    if all_wipe_fights:
        print(f"\n[...] Fetching wipe events ({len(all_wipe_fights)} wipe(s))...")
        encounter_wipe_count: dict = {}
        for fight in sorted(all_wipe_fights, key=lambda f: f["id"]):
            fid   = fight["id"]
            fname = fight["name"]
            eid   = fight.get("encounterID", 0)
            diff  = {3: "Normal", 4: "Heroic", 5: "Mythic"}.get(fight.get("difficulty", 0), "")
            boss_key = f"{fname} ({diff})"
            encounter_wipe_count[eid] = encounter_wipe_count.get(eid, 0) + 1
            wipe_num     = encounter_wipe_count[eid]
            boss_pct     = fight.get("bossPercentage", 0)
            fight_start  = fight.get("startTime", 0)
            fight_dur_ms = fight.get("endTime", 0) - fight_start
            try:
                death_events     = fetch_death_events(token, report_code, fid)
                damage_events    = fetch_damage_taken_events(token, report_code, fid)
                interrupt_events = fetch_interrupt_events(token, report_code, fid)
                cast_events      = fetch_boss_cast_events(token, report_code, fid)
                deaths           = analyze_deaths(death_events, fight_start, ability_names,
                                                  fight_end_ms=fight_dur_ms, damage_events=damage_events)
                avoidable        = analyze_avoidable_damage(damage_events, actor_lookup, fight_start,
                                                            ability_names, player_max_hp, player_roles_map)
                dmg_taken        = aggregate_damage_taken(damage_events, actor_lookup)
                uptime_map       = fetch_uptime_table(token, report_code, fid)
                healing_map      = fetch_healing_table(token, report_code, fid)
                interrupts       = analyze_interrupts(interrupt_events, actor_lookup)
                mech_defs        = BOSS_MECHANICS.get(fname, [])
                mechanics_data   = analyze_boss_mechanics(damage_events, actor_lookup, mech_defs)
                frontal_failures = analyze_frontal_failures(damage_events, actor_lookup, mech_defs,
                                                            fight_start, cast_events=cast_events)
                try:
                    wipe_ci = fetch_combatant_info_events(token, report_code, fid)
                    wipe_spec_roles = {
                        e["sourceID"]: SPEC_ROLES.get(e.get("specID"), "DPS")
                        for e in wipe_ci if e.get("sourceID") and e.get("specID")
                    }
                except Exception:
                    wipe_spec_roles = {}
                try:
                    wipe_friendly_casts = fetch_cast_events(token, report_code, fid,
                                                            start_time=fight_start,
                                                            end_time=fight.get("endTime", 0))
                    wipe_player_casts = analyze_fight_casts(wipe_friendly_casts, fight_start, actor_lookup)
                    wipe_defensive_casts = {
                        pid: [{"spell": c["spell"], "time": c["time"]}
                              for c in casts
                              if c.get("category") in ("Defensive", "Healthstone", "Health")]
                        for pid, casts in wipe_player_casts.items()
                        if any(c.get("category") in ("Defensive", "Healthstone", "Health") for c in casts)
                    }
                    wipe_external_casts = {
                        pid: [{"spell": c["spell"], "time": c["time"]}
                              for c in casts if c.get("category") == "External"]
                        for pid, casts in wipe_player_casts.items()
                        if any(c.get("category") == "External" for c in casts)
                    }
                except Exception:
                    wipe_defensive_casts = {}
                    wipe_external_casts  = {}
                all_pids = {pid for pid in (set(dmg_taken) | set(uptime_map) | set(deaths)) if pid in actor_lookup}
                # Crown-specific mechanic analysis
                crown_mechanics = {}
                if "Crown" in fname:
                    try:
                        crown_debuffs   = fetch_crown_debuff_events(token, report_code, fid)
                        crown_add_buffs = fetch_crown_add_buff_events(token, report_code, fid)
                        crown_casts     = fetch_crown_cast_events(token, report_code, fid)
                        crown_mechanics = analyze_crown_mechanics(
                            crown_debuffs, damage_events, actor_lookup,
                            wipe_spec_roles, fight_start,
                            add_buff_events=crown_add_buffs,
                            cast_events=crown_casts)
                    except Exception as ce:
                        print(f"  [WARN] Crown mechanics fetch failed for fight {fid}: {ce}")
                # Vorasius: tank swap failures
                w_vorasius_tank_swaps = (analyze_vorasius_tank_swaps(
                    damage_events, actor_lookup, fight_start, wipe_spec_roles)
                    if "Vorasius" in fname else [])
                # All bosses: per-mechanic hit timestamps
                w_mechanic_timestamps = analyze_mechanic_timestamps(
                    damage_events, actor_lookup, fight_start, mech_defs)
                # Salhadaar: fracture wave table + wipe flags
                w_sal_fracture_waves = []
                w_sal_wipe_events    = []
                if "Salhadaar" in fname:
                    w_sal_cc_events = fetch_salhadaar_cc_events(
                        token, report_code, fid, fight_start, fight.get("endTime", 0))
                    w_sal_fracture_waves = analyze_salhadaar_fracture_waves(
                        interrupt_events, actor_lookup, fight_start,
                        ability_names=ability_names, spell_names_map=SPELL_NAMES,
                        cc_events=w_sal_cc_events)
                    w_sal_wipe_events = detect_salhadaar_wipe_events(
                        cast_events, fight_start, ability_names=ability_names)
                wipe_data.setdefault(boss_key, []).append({
                    "fight_id": fid, "fight_dur_ms": fight_dur_ms,
                    "wipe_num": wipe_num, "boss_pct": boss_pct,
                    "deaths": deaths, "all_player_ids": all_pids,
                    "avoidable_damage": avoidable, "dmg_taken": dmg_taken,
                    "uptime_map": uptime_map, "healing_map": healing_map,
                    "rankings_map": {},  # wipes have no rankings
                    "interrupts": interrupts, "mechanics_data": mechanics_data,
                    "frontal_failures": frontal_failures,
                    "spec_roles": wipe_spec_roles,
                    "defensive_casts": wipe_defensive_casts,
                    "external_casts":  wipe_external_casts,
                    "crown_mechanics": crown_mechanics,
                    "vorasius_tank_swaps":      w_vorasius_tank_swaps,
                    "mechanic_timestamps":      w_mechanic_timestamps,
                    "salhadaar_fracture_waves": w_sal_fracture_waves,
                    "salhadaar_wipe_events":    w_sal_wipe_events,
                })
                dur_s = fight_dur_ms // 1000
                print(f"  [OK] {fname} Wipe {wipe_num} ({boss_pct}% boss HP, {dur_s//60}:{dur_s%60:02d}): {len(deaths)} death(s).")
            except Exception as e:
                print(f"  [WARN] Could not fetch wipe events for fight {fid}: {e}")

    result = {
        "report_code": report_code, "report_info": report_info,
        "records": records, "split_data": split_data,
        "boss_data": boss_data, "wipe_data": wipe_data, "actors": actors,
        "max_splits": max_splits,
    }
    cache = _cache_path(report_code)
    with open(cache, "w", encoding="utf-8") as f:
        json.dump(result, f, default=list)
    print(f"[CACHE] Saved to {cache}")
    return result


def split_report_by_difficulty(day_data: dict) -> list:
    """If a report has multiple difficulties, return one day_data per difficulty."""
    diffs_present = [d for d in ("Normal", "Heroic", "Mythic")
                     if any(f"({d})" in k for k in day_data.get("boss_data", {}))]
    if len(diffs_present) <= 1:
        return [day_data]
    results = []
    for diff in diffs_present:
        filtered_boss  = {k: [dict(f, split_num=1) for f in v]
                          for k, v in day_data["boss_data"].items() if f"({diff})" in k}
        filtered_wipes = {k: v for k, v in day_data.get("wipe_data", {}).items() if f"({diff})" in k}
        results.append({**day_data, "boss_data": filtered_boss, "wipe_data": filtered_wipes, "difficulty": diff})
    return results


def split_report_by_player_group(day_data: dict) -> list:
    """If a report has multiple player-group splits, return one day_data per split."""
    bd = day_data.get("boss_data", {})
    split_nums = {f.get("split_num", 1) for fights in bd.values() for f in fights}
    if len(split_nums) <= 1:
        return [day_data]
    results = []
    for snum in sorted(split_nums):
        filtered_boss = {k: [f for f in v if f.get("split_num", 1) == snum]
                         for k, v in bd.items()}
        filtered_boss = {k: v for k, v in filtered_boss.items() if v}
        results.append({**day_data, "boss_data": filtered_boss,
                        "wipe_data": day_data.get("wipe_data", {}), "player_split": snum})
    return results


def _raid_filename(day_data: dict) -> str:
    rc     = day_data["report_code"]
    diff   = day_data.get("difficulty", "")
    psplit = day_data.get("player_split")
    return f"raid_{rc}{'_' + diff.lower() if diff else ''}{'_split' + str(psplit) if psplit else ''}.html"


def main():
    print("=" * 60)
    print("  WarcraftLogs Crafted Gear Audit — Midnight Season 1")
    print("=" * 60)

    config = load_config()

    if config.get("CLIENT_ID") and config.get("CLIENT_SECRET"):
        client_id, client_secret = config["CLIENT_ID"], config["CLIENT_SECRET"]
        print(f"\n[OK] Loaded credentials from wcl_config.txt (Client ID: {client_id[:8]}...)")
    else:
        client_id     = input("\nWarcraftLogs Client ID: ").strip()
        client_secret = input("WarcraftLogs Client Secret: ").strip()

    if not client_id or not client_secret:
        print("[ERROR] Both Client ID and Client Secret are required.")
        sys.exit(1)

    print("\n[...] Authenticating with WarcraftLogs...")
    token = get_access_token(client_id, client_secret)
    print("[OK] Authenticated successfully.")

    report_configs = config.get("REPORT_URLS", [])
    if not report_configs:
        report_configs = [{"url": input("\nReport code or URL: ").strip(), "split_start": 1}]
    else:
        print(f"[OK] Loaded {len(report_configs)} report URL(s) from config.")

    fight_mode = "interactive" if len(report_configs) == 1 else "all"

    days_data = []
    for rc in report_configs:
        url         = rc["url"] if isinstance(rc, dict) else rc
        split_start = rc.get("split_start", 1) if isinstance(rc, dict) else 1
        code        = _extract_report_code(url)
        try:
            result = process_report(token, code, fight_input=fight_mode)
            for r in split_report_by_difficulty(result):
                for dd in split_report_by_player_group(r):
                    if "player_split" in dd:
                        dd["player_split"] = dd["player_split"] + split_start - 1
                    days_data.append(dd)
        except Exception as e:
            print(f"[ERROR] Failed to process report {code}: {e}")

    if not days_data:
        print("[ERROR] No reports processed successfully.")
        sys.exit(1)

    # Most recent report first
    days_data.sort(key=lambda d: d["report_info"].get("startTime", 0), reverse=True)

    # Guild name (from most recent report)
    ri_first   = days_data[0]["report_info"]
    guild_name = ri_first.get("guild", {}).get("name", "The Rupture") if ri_first.get("guild") else "The Rupture"

    # Per-raid pages
    for day_data in days_data:
        write_raid_html(day_data, _raid_filename(day_data))

    # Overview index
    write_index_html(days_data, "index.html", guild_name=guild_name)
    write_bosses_html(days_data, "bosses.html", guild_name=guild_name)
    write_gear_html(days_data, "gear_normal.html", guild_name=guild_name)
    write_roster_html(days_data, "roster.html", guild_name=guild_name)
    write_boss_progression_html(days_data, "boss_chimaerus_heroic.html", guild_name=guild_name)
    write_crown_progression_html(days_data, "boss_crown_heroic.html", guild_name=guild_name)
    write_player_pages(days_data)

    # XLSX: most recent day only
    first = days_data[0]
    write_xlsx(first["records"], first["report_info"], first["report_code"],
               "craft_audit.xlsx", split_data=first["split_data"],
               boss_data=first["boss_data"], actors=first["actors"])

    print(f"\nDone! Open index.html in your browser, or open craft_audit.xlsx.\n")


if __name__ == "__main__":
    main()